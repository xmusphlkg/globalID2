from __future__ import annotations

from datetime import date
import hashlib
import json
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

from src.data.crawlers.is_history import (
    IcelandHistoryCrawler,
    IcelandHistoryRawFile,
    IcelandHistoryWorkbookSpec,
)
from src.data.processors.is_history import (
    HISTORY_SOURCE_ID,
    LEGACY_PAIR_LOOKUP,
    LEGACY_SOURCE_ID,
    MONTHLY_SERIES,
    ANNUAL_SERIES,
    IcelandHistoryPreparedResult,
    IcelandHistoryProcessor,
    _Workbook,
    _raw_file_from_manifest,
)
from scripts.import_iceland_history import _retire_ineligible_compatibility_rows


ICELANDIC_MONTHS = [
    "Janúar",
    "Febrúar",
    "Mars",
    "Apríl",
    "Maí",
    "Júní",
    "Júlí",
    "Ágúst",
    "September",
    "Október",
    "Nóvember",
    "Desember",
]


def _raw_file(
    path: Path,
    *,
    key: str,
    source_kind: str,
    disease_key: str = "",
) -> IcelandHistoryRawFile:
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return IcelandHistoryRawFile(
        key=key,
        source_kind=source_kind,
        filename=path.name,
        path=str(path),
        source_url=f"https://example.test/{path.name}",
        sha256=digest,
        size_bytes=path.stat().st_size,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disease_key=disease_key,
    )


def _history_registry_entry(definition, source_kind: str) -> dict:
    return {
        "id": definition.source_series_code(source_kind),
        "source_id": HISTORY_SOURCE_ID,
        "concept_id": definition.concept_id,
        "local_codes": [definition.local_code(source_kind)],
        "local_labels": [definition.label_is],
        "frequency": "annual" if source_kind == "registry_annual" else "monthly",
        "measure": "case_notifications",
        "unit": "count",
        "mapping_relation": "exact",
        "comparability": "conditional",
        "aggregation_policy": "non_additive",
        "status": "historical",
    }


def _legacy_registry_entry(definition) -> dict:
    return {
        "id": definition.source_series_code,
        "source_id": LEGACY_SOURCE_ID,
        "concept_id": definition.concept_id,
        "local_codes": [definition.local_code],
        "local_labels": [definition.label_is],
        "frequency": "monthly",
        "measure": "registered_diagnoses",
        "status": "historical",
    }


def _write_ontology(tmp_path: Path, entries: list[dict]) -> Path:
    path = tmp_path / "ontology.json"
    path.write_text(json.dumps({"source_series": entries}), encoding="utf-8")
    return path


def test_crawler_catalogue_has_22_primary_and_one_validation_file() -> None:
    primary = IcelandHistoryCrawler.catalogue()
    complete = IcelandHistoryCrawler.catalogue(include_validation=True)

    assert len(primary) == 22
    assert len(complete) == 23
    assert {item.source_kind for item in primary} == {
        "registry_annual",
        "registry_disease_monthly",
        "legacy_icd_monthly",
    }
    assert sum(item.source_kind == "registry_disease_monthly" for item in primary) == 14
    assert complete[-1].validation_only is True


def test_crawler_writes_hash_manifest_and_validates_signature(tmp_path: Path) -> None:
    payload = b"PK\x03\x04synthetic-ooxml"
    spec = IcelandHistoryWorkbookSpec(
        key="synthetic",
        source_kind="registry_annual",
        filename="synthetic.xlsx",
        url="https://example.test/synthetic.xlsx",
    )
    crawler = IcelandHistoryCrawler(raw_dir=tmp_path, delay=0)
    crawler.get = lambda _url: SimpleNamespace(content=payload)  # type: ignore[method-assign]

    result = crawler.download_history(specs=[spec], discover=False)
    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))

    expected_hash = hashlib.sha256(payload).hexdigest()
    assert result.raw_files[0].sha256 == expected_hash
    assert manifest["files"][0]["sha256"] == expected_hash
    assert manifest["files"][0]["path"] == "synthetic.xlsx"
    assert not Path(manifest["files"][0]["path"]).is_absolute()
    assert Path(result.raw_files[0].path).read_bytes() == payload


def test_processor_replays_relocated_legacy_absolute_manifest_path(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "historical.xlsx"
    workbook.write_bytes(b"relocated workbook")
    manifest = tmp_path / "raw_manifest.json"
    descriptor = _raw_file_from_manifest(
        {
            "key": "annual",
            "source_kind": "registry_annual",
            "filename": workbook.name,
            "path": "/old-machine/archive/historical.xlsx",
            "source_url": "https://example.test/historical.xlsx",
            "sha256": hashlib.sha256(workbook.read_bytes()).hexdigest(),
            "size_bytes": workbook.stat().st_size,
        },
        manifest,
    )

    assert Path(descriptor.path) == workbook.resolve()


def test_processor_writes_portable_paths_to_normalized_manifest(
    tmp_path: Path,
) -> None:
    raw_path = tmp_path / "data" / "raw" / "is" / "history" / "history.xlsx"
    raw_path.parent.mkdir(parents=True)
    raw_path.write_bytes(b"reviewed workbook")
    output_dir = tmp_path / "data" / "current" / "is" / "history"
    result = IcelandHistoryPreparedResult(
        rows=[],
        series_rows=[],
        quarantine=[],
        manifest={
            "files": [{"filename": raw_path.name, "path": str(raw_path.resolve())}]
        },
        raw_hashes={},
    )

    outputs = IcelandHistoryProcessor.write_outputs(result, output_dir)

    manifest = json.loads(outputs["manifest.json"].read_text(encoding="utf-8"))
    stored_path = Path(manifest["files"][0]["path"])
    assert not stored_path.is_absolute()
    assert (output_dir / stored_path).resolve() == raw_path.resolve()
    assert result.manifest["files"][0]["path"] == str(raw_path.resolve())


class _HistoryDBResult:
    def __init__(self, rows=()):
        self._rows = list(rows)

    def fetchone(self):
        return self._rows[0] if self._rows else None

    def fetchall(self):
        return list(self._rows)


class _HistoryImportDB:
    def __init__(self, *, current_identities=()):
        self.current_identities = list(current_identities)
        self.upsert_params = None
        self.upsert_sql = ""

    async def execute(self, statement, params=None):
        sql = str(statement)
        if "SELECT id FROM countries" in sql:
            return _HistoryDBResult([(7,)])
        if "SELECT name, id FROM diseases" in sql:
            return _HistoryDBResult([("D236", 236)])
        if "legacy_projection" in sql and "SELECT timezone" in sql:
            return _HistoryDBResult(self.current_identities)
        if "INSERT INTO disease_records" in sql:
            self.upsert_sql = sql
            self.upsert_params = params
            return _HistoryDBResult()
        raise AssertionError(f"Unexpected SQL: {sql}")


class _RetirementResult:
    rowcount = 2


class _RetirementDB:
    def __init__(self) -> None:
        self.sql = ""
        self.params = None

    async def execute(self, statement, params=None):
        self.sql = str(statement)
        self.params = params
        return _RetirementResult()


@pytest.mark.asyncio
async def test_retirement_counts_composite_key_rows_without_assuming_id_column() -> None:
    prepared = SimpleNamespace(
        series_rows=[
            {
                "SourceId": HISTORY_SOURCE_ID,
                "SourceSeriesCode": "SER_UNSAFE",
                "Measure": "case_notifications",
            },
            {
                "SourceId": HISTORY_SOURCE_ID,
                "SourceSeriesCode": "SER_SAFE",
                "Measure": "case_notifications",
            },
        ],
        rows=[
            {
                "SourceId": HISTORY_SOURCE_ID,
                "SourceSeriesCode": "SER_SAFE",
            }
        ],
    )
    db = _RetirementDB()

    retired = await _retire_ineligible_compatibility_rows(db, prepared)

    assert retired == 2
    assert "RETURNING id" not in db.sql
    assert db.params == {"series_codes": ["SER_UNSAFE"]}


@pytest.mark.asyncio
async def test_history_projection_never_overwrites_current_annual_mrsa_2019() -> None:
    processor = IcelandHistoryProcessor(require_registered_series=False)
    db = _HistoryImportDB(current_identities=[(date(2019, 1, 1), 236)])
    row = {
        "Date": "2019-01-01",
        "Cases": "31",
        "DiseaseFull": "D236",
        "SourceKind": "registry_annual",
        "Measure": "case_notifications",
        "Source": "Iceland Directorate of Health Historical Registry",
        "SourceSeriesCode": "SER_IS_HISTORY_MRSA_ANNUAL",
    }

    result = await processor.import_rows(db, [row])

    assert result.inserted_or_updated == 0
    assert result.skipped_current_precedence == 1
    assert db.upsert_params is None


@pytest.mark.asyncio
async def test_history_projection_conflict_clause_also_protects_concurrent_current() -> None:
    processor = IcelandHistoryProcessor(require_registered_series=False)
    db = _HistoryImportDB()
    row = {
        "Date": "2019-01-01",
        "Cases": "31",
        "DiseaseFull": "D236",
        "SourceKind": "registry_annual",
        "Measure": "case_notifications",
        "Source": "Iceland Directorate of Health Historical Registry",
        "SourceSeriesCode": "SER_IS_HISTORY_MRSA_ANNUAL",
    }

    result = await processor.import_rows(db, [row])

    assert result.inserted_or_updated == 1
    assert "current_annual_dashboard_only" in db.upsert_sql
    assert "WHERE COALESCE" in db.upsert_sql


def test_annual_parser_preserves_zero_skips_dash_and_quarantines_unknown(
    tmp_path: Path,
) -> None:
    definition = next(item for item in ANNUAL_SERIES if item.key == "hepatitis_a")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "2020-2021"
    sheet.cell(5, 3, 2020)
    sheet.cell(5, 4, 2021)
    sheet.cell(6, 3, "Fjöldi")
    sheet.cell(6, 4, "Fjöldi")
    sheet.cell(7, 3, "Number")
    sheet.cell(7, 4, "Number")
    sheet.cell(9, 1, "Lifrarbólga A")
    sheet.cell(9, 2, "Hepatitis A")
    sheet.cell(9, 3, 0)
    sheet.cell(9, 4, "-")
    sheet.cell(10, 1, "Ókortlagður sjúkdómur")
    sheet.cell(10, 2, "Unreviewed condition")
    sheet.cell(10, 3, 2)
    sheet.cell(10, 4, 1)
    # A duplicate rate section after three blank rows must never be parsed.
    sheet.cell(20, 1, "Lifrarbólga A")
    sheet.cell(20, 3, 0.75)
    path = tmp_path / "annual.xlsx"
    workbook.save(path)
    ontology = _write_ontology(
        tmp_path, [_history_registry_entry(definition, "registry_annual")]
    )

    result = IcelandHistoryProcessor(ontology_path=ontology).prepare_files(
        [_raw_file(path, key="annual", source_kind="registry_annual")]
    )

    assert len(result.series_rows) == 1
    assert result.series_rows[0]["Cases"] == "0"
    assert result.series_rows[0]["ValueStatus"] == "reported_zero"
    assert result.series_rows[0]["Date"] == "2020-01-01"
    assert result.manifest["counts"]["annual.not_applicable_dash_cells"] == 1
    assert result.manifest["quarantine"]["by_reason"] == {
        "unreviewed_annual_disease": 2
    }


def test_irregular_monthly_layout_formula_dash_and_blank_semantics(
    tmp_path: Path,
) -> None:
    definition = MONTHLY_SERIES["hepatitis_c"]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Greiningarmánuður"
    sheet.cell(3, 3, 2020)
    sheet.cell(3, 8, 2021)
    sheet.cell(4, 3, "Fjöldi")
    sheet.cell(4, 5, "Fjöldi/100.000")
    sheet.cell(4, 8, "Fjöldi")
    sheet.cell(4, 10, "Fjöldi/100.000")
    sheet.cell(5, 3, "Number")
    sheet.cell(5, 5, "Number/100.000")
    sheet.cell(5, 8, "Number")
    sheet.cell(5, 10, "Number/100.000")
    for month, label in enumerate(ICELANDIC_MONTHS, 1):
        row = month + 5
        sheet.cell(row, 1, label)
        sheet.cell(row, 3, 1)
        sheet.cell(row, 5, 0.25)
        sheet.cell(row, 8, 2)
        sheet.cell(row, 10, 0.5)
    sheet.cell(6, 3, "-")
    sheet.cell(7, 3).value = None
    sheet.cell(8, 3, "=1+2")
    sheet.cell(8, 5, "=C8/3")
    path = tmp_path / "hepatitis_c.xlsx"
    workbook.save(path)
    ontology = _write_ontology(
        tmp_path,
        [_history_registry_entry(definition, "registry_disease_monthly")],
    )

    result = IcelandHistoryProcessor(ontology_path=ontology).prepare_files(
        [
            _raw_file(
                path,
                key="hepatitis-c",
                source_kind="registry_disease_monthly",
                disease_key="hepatitis_c",
            )
        ]
    )

    assert len(result.series_rows) == 23
    january = next(row for row in result.series_rows if row["Date"] == "2020-01-01")
    march = next(row for row in result.series_rows if row["Date"] == "2020-03-01")
    assert january["Cases"] == "0"
    assert january["ValueStatus"] == "dash_zero"
    assert march["Cases"] == "3"
    assert march["Incidence"] == "1"
    assert march["ValueStatus"] == "formula_evaluated"
    assert not any(row["Date"] == "2020-02-01" for row in result.series_rows)
    assert result.manifest["counts"]["disease_monthly.blank_unknown_cells"] == 1


def test_legacy_icd_is_series_only_and_identity_uses_code_and_label(
    tmp_path: Path,
) -> None:
    definition = LEGACY_PAIR_LOOKUP[("B01-B01.9", "hlaupabóla")]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "2020"
    month_labels = ["Jan", "Feb", "Mar", "Apr", "Maí", "Jún", "Júl", "Ágú", "Sep", "Okt", "Nóv", "Des"]
    for column, label in enumerate(month_labels, 2):
        sheet.cell(5, column, label)
    sheet.cell(6, 1, "Hlaupabóla")
    for column in range(2, 14):
        sheet.cell(6, column, 1)
    sheet.cell(6, 2, "-")
    sheet.cell(6, 3).value = None
    sheet.cell(6, 4, "=1+1")
    sheet.cell(7, 1, "Ókortlögð greining")
    for column in range(2, 14):
        sheet.cell(7, column, 3)
    explanation = workbook.create_sheet("Skýringar")
    explanation.cell(1, 2, "ICD-10")
    explanation.cell(3, 1, "Hlaupabóla")
    explanation.cell(3, 2, "B01-B01.9")
    explanation.cell(4, 1, "Ókortlögð greining")
    explanation.cell(4, 2, "Z99")
    path = tmp_path / "legacy.xlsx"
    workbook.save(path)
    ontology = _write_ontology(tmp_path, [_legacy_registry_entry(definition)])

    result = IcelandHistoryProcessor(ontology_path=ontology).prepare_files(
        [_raw_file(path, key="legacy", source_kind="legacy_icd_monthly")]
    )

    assert result.rows == []
    assert len(result.series_rows) == 11
    assert {row["ICD10"] for row in result.series_rows} == {"B01-B01.9"}
    assert {row["Measure"] for row in result.series_rows} == {
        "registered_diagnoses"
    }
    assert {row["DiseaseCode"] for row in result.series_rows} == {
        "IS_LEGACY_ICD_B01_B01_9_HLAUPABOLA"
    }
    assert result.manifest["quarantine"]["by_reason"] == {
        "unreviewed_legacy_icd_series": 12
    }


def test_ontology_gate_fails_closed_on_unregistered_series(tmp_path: Path) -> None:
    definition = next(item for item in ANNUAL_SERIES if item.key == "hepatitis_a")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(3, 3, 2020)
    sheet.cell(3, 4, 2021)
    sheet.cell(4, 3, "Fjöldi")
    sheet.cell(4, 4, "Fjöldi")
    sheet.cell(6, 1, definition.label_is)
    sheet.cell(6, 3, 1)
    sheet.cell(6, 4, 2)
    path = tmp_path / "unregistered.xlsx"
    workbook.save(path)
    ontology = _write_ontology(tmp_path, [])

    result = IcelandHistoryProcessor(ontology_path=ontology).prepare_files(
        [_raw_file(path, key="annual", source_kind="registry_annual")]
    )

    assert result.series_rows == []
    assert result.rows == []
    assert result.manifest["quarantine"]["by_reason"] == {
        "source_series_not_registered": 2
    }


def test_projection_excludes_24_colliding_rows_but_keeps_complete_series(
    tmp_path: Path,
) -> None:
    definitions = [
        next(item for item in ANNUAL_SERIES if item.key == "hib"),
        next(item for item in ANNUAL_SERIES if item.key == "invasive_h_influenzae"),
    ]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for offset, year in enumerate(range(1997, 2009), 3):
        sheet.cell(3, offset, year)
        sheet.cell(4, offset, "Fjöldi")
    for row_number, definition in enumerate(definitions, 6):
        sheet.cell(row_number, 1, definition.label_is)
        sheet.cell(row_number, 2, definition.label_en)
        for column in range(3, 15):
            sheet.cell(row_number, column, 1)
    path = tmp_path / "colliding-annual.xlsx"
    workbook.save(path)
    ontology = _write_ontology(
        tmp_path,
        [
            _history_registry_entry(definition, "registry_annual")
            for definition in definitions
        ],
    )

    result = IcelandHistoryProcessor(ontology_path=ontology).prepare_files(
        [_raw_file(path, key="annual", source_kind="registry_annual")]
    )

    assert len(result.series_rows) == 24
    assert result.rows == []
    assert (
        result.manifest["counts"][
            "projection.multi_series_identity_rows_excluded"
        ]
        == 24
    )
    assert len(
        {
            (row["DiseaseFull"], row["Date"], row["SourceSeriesCode"])
            for row in result.series_rows
        }
    ) == 24
    assert len(
        {(row["DiseaseFull"], row["Date"]) for row in result.rows}
    ) == len(result.rows)


def test_xls_adapter_uses_xlrd_for_legacy_ole_workbooks(
    tmp_path: Path, monkeypatch
) -> None:
    class FakeSheet:
        nrows = 2
        ncols = 2

        @staticmethod
        def cell_value(row: int, column: int):
            return [["ICD-10", "jan"], ["B01", 4]][row][column]

    class FakeBook:
        @staticmethod
        def sheet_names():
            return ["1997"]

        @staticmethod
        def sheet_by_name(_name: str):
            return FakeSheet()

        @staticmethod
        def release_resources() -> None:
            return None

    monkeypatch.setattr(
        "src.data.processors.is_history.xlrd.open_workbook",
        lambda *_args, **_kwargs: FakeBook(),
    )
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"synthetic-ole-placeholder")

    workbook = _Workbook(path)
    try:
        assert workbook.sheet_names == ["1997"]
        assert workbook.dimensions("1997") == (2, 2)
        assert workbook.raw("1997", 2, 2) == 4
    finally:
        workbook.close()
