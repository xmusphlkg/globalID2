from __future__ import annotations

import json

from openpyxl import load_workbook
import pytest

from src.generation.direct_download_files import (
    build_country_download_rows,
    build_direct_download_files,
    partition_rows,
)
from src.generation.site_data_views import (
    build_country_data,
    build_country_site_data,
)
from src.generation.site_series_projection import apply_series_first_projection


def _context(
    dates: list[str] | None = None,
    *,
    generated_at: str = "2026-08-01T00:00:00+00:00",
    dataset_name: str = "Hepatitis A",
) -> dict:
    dates = dates or ["2010-01-01", "2015-01-01", "2020-01-01", "2026-01-01"]
    count = len(dates)
    source_info = {
        "primary_scope": "official",
        "primary_label": "Official source",
        "primary_url": "https://example.test/source",
        "primary_type": "api",
        "sources": [
            {
                "scope": "official",
                "label": "Official source",
                "url": "https://example.test/source",
                "type": "api",
            }
        ],
    }
    series = {
        "disease_id": "D007",
        "name_en": dataset_name,
        "name_zh": "甲型肝炎",
        "category": "Viral",
        "dates": dates,
        "cases": list(range(1, count + 1)),
        "weekly_equiv_cases": [1.0] * count,
        "deaths": [0] * count,
        "incidence_rates": [0.1] * count,
        "incidence_sources": ["official"] * count,
        "mortality_rates": [0] * count,
    }
    return {
        "generated_at": generated_at,
        "countries_simple": [{"code": "CN", "name": "China"}],
        "country_sources_by_code": {"CN": source_info},
        "country_download_entries": [
            {
                "kind": "country",
                "id": "cn",
                "code": "CN",
                "name": "China",
                "record_count": count,
                "date_range": {"start": min(dates), "end": max(dates)},
                "site_json_path": "/site-data/countries/cn.json",
            }
        ],
        "disease_download_entries": [
            {
                "kind": "disease",
                "id": "d007",
                "disease_id": "D007",
                "record_count": count,
                "date_range": {"start": min(dates), "end": max(dates)},
                "site_json_path": "/site-data/diseases/d007.json",
            }
        ],
        "country_exports": [
            {
                "code": "CN",
                "country_name": "China",
                "source_info": source_info,
                "country_data": {
                    "country_code": "CN",
                    "country_name": "China",
                    "disease_series": {"D007": series},
                },
            }
        ],
        "disease_exports": [
            {
                "disease_id": "D007",
                "disease_data": {
                    "disease_id": "D007",
                    "slug": "hepatitis-a",
                    "name_en": dataset_name,
                    "name_zh": "甲型肝炎",
                    "category": "Viral",
                    "country_series": {"CN": series},
                    "source_info": [source_info],
                },
            }
        ],
    }


def test_partition_boundaries_use_zero_and_five_year_anchors() -> None:
    rows = [
        {"date": value}
        for value in (
            "2010-01-01",
            "2014-12-01",
            "2015-01-01",
            "2019-12-01",
            "2020-01-01",
            "2025-12-01",
            "2026-01-01",
            "2029-12-01",
            "2030-01-01",
        )
    ]

    parts = partition_rows(rows)
    assert [(part.key, part.label) for part, _ in parts] == [
        ("2030-2034", "2030–now"),
        ("2026-2029", "2026–2029"),
        ("2020-2025", "2020–2025"),
        ("2015-2019", "2015–2019"),
        ("2010-2014", "2010–2014"),
    ]


def test_current_bridge_window_is_2026_now_until_2030_exists() -> None:
    parts = partition_rows([{"date": "2026-01-01"}, {"date": "2029-01-01"}])
    assert parts[0][0].key == "2026-2029"
    assert parts[0][0].label == "2026–now"
    assert parts[0][0].is_current is True


def test_build_writes_three_formats_for_every_partition(tmp_path) -> None:
    base = "https://raw.githubusercontent.com/example/data/main"
    manifest = build_direct_download_files(_context(), tmp_path, download_url_base=base)

    entry = manifest["diseases"][0]
    assert manifest["schema_version"] == 4
    assert manifest["includes_series_provenance"] is True
    current = entry["parts"][0]
    assert manifest["formats"] == ["csv", "json", "xlsx"]
    assert current["id"] == "2026-2029"
    for format_name in manifest["formats"]:
        meta = current["files"][format_name]
        assert meta["url"] == f"{base}/{meta['relative_path']}"
        assert (tmp_path / meta["relative_path"]).is_file()

    json_path = tmp_path / current["files"]["json"]["relative_path"]
    payload = json.loads(json_path.read_text())
    assert payload["metadata"]["record_count"] == 1
    assert payload["metadata"]["schema_version"] == 4
    assert "generated_at" not in payload["metadata"]
    assert "generated_at" not in payload["records"][0]

    xlsx_path = tmp_path / current["files"]["xlsx"]["relative_path"]
    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook.sheetnames == ["records", "metadata"]


def test_downloads_include_lossless_source_series_and_projection_provenance(
    tmp_path,
) -> None:
    context = _context(["2026-01-01"])
    series = context["country_exports"][0]["country_data"]["disease_series"][
        "D007"
    ]
    series.update(
        {
            "data_layer": "series_registry",
            "projection_policy": "single_series",
            "period_granularity": "monthly",
            "selected_series_codes": ["sti:hepatitis-a:monthly-diagnoses"],
            "available_series_count": 2,
            "data_provenance": {"note_en": "Monthly facts; quarterly publication."},
            "source_series": [
                {
                    "series_code": "sti:hepatitis-a:monthly-diagnoses",
                    "source_series_code": "monthly-diagnoses",
                    "source_system": "is_doh_sti",
                    "source_label": "Iceland Directorate of Health STI Dashboard",
                    "metric_type": "case_notifications",
                    "reporting_basis": "laboratory_and_registry_diagnoses",
                    "temporal_granularity": "monthly",
                    "unit": "count",
                    "geography_key": "country:IS:national",
                    "dimension_key": "all",
                    "mapping_relation": "exact",
                    "comparability": "within_series",
                    "aggregation_policy": "none",
                    "availability_status": "available",
                    "missing_value_policy": "missing_not_zero",
                    "definition_version": "2026-08-01",
                    "dates": ["2026-01-01"],
                    "values": [7],
                    "quality_statuses": ["validated"],
                },
                {
                    "series_code": "respiratory:hepatitis-a:hospitalizations",
                    "source_series_code": "hospitalizations",
                    "source_system": "is_doh_respiratory",
                    "source_label": "Iceland Directorate of Health Respiratory Dashboard",
                    "metric_type": "hospitalizations",
                    "reporting_basis": "hospital_admissions",
                    "temporal_granularity": "weekly",
                    "unit": "count",
                    "geography_key": "country:IS:national",
                    "dimension_key": "all",
                    "mapping_relation": "related",
                    "comparability": "not_comparable_with_diagnoses",
                    "aggregation_policy": "none",
                    "availability_status": "available",
                    "missing_value_policy": "missing_not_zero",
                    "definition_version": "2026-08-01",
                    "dates": ["2026-01-01"],
                    "values": [3],
                    "quality_statuses": ["validated"],
                },
            ],
        }
    )
    context["disease_exports"][0]["disease_data"]["country_series"]["CN"] = series

    manifest = build_direct_download_files(
        context,
        tmp_path,
        download_url_base="https://raw.githubusercontent.com/example/data/main",
    )

    entry = manifest["countries"][0]
    assert entry["source_series_count"] == 2
    assert entry["source_observation_count"] == 2
    assert entry["projection_record_count"] == 1
    part = entry["parts"][0]
    payload = json.loads(
        (tmp_path / part["files"]["json"]["relative_path"]).read_text()
    )
    projection_row = next(
        record for record in payload["records"]
        if record["record_kind"] == "public_projection"
    )
    source_rows = {
        record["series_code"]: record
        for record in payload["records"]
        if record["record_kind"] == "source_series_observation"
    }
    source_row = source_rows["sti:hepatitis-a:monthly-diagnoses"]
    assert source_row["series_code"] == "sti:hepatitis-a:monthly-diagnoses"
    assert source_row["reporting_basis"] == "laboratory_and_registry_diagnoses"
    assert source_row["temporal_granularity"] == "monthly"
    assert source_row["weekly_equiv_cases"] is None
    assert source_row["is_selected_series"] is True
    assert source_row["value"] == 7
    assert source_row["cases"] == 7
    hospital_row = source_rows["respiratory:hepatitis-a:hospitalizations"]
    assert hospital_row["value"] == 3
    assert hospital_row["cases"] is None
    assert hospital_row["weekly_equiv_cases"] is None
    assert projection_row["provenance_note"] == (
        "Monthly facts; quarterly publication."
    )
    assert payload["metadata"]["total_basis"] == "public_projection"
    assert payload["metadata"]["total_cases"] == 1


def test_iceland_exports_all_125_series_and_9396_observations_without_false_curves() -> None:
    """Regression for the 18 source-only Iceland registered-diagnosis series."""

    source_only_diseases = [
        "D031",
        "D047",
        "D054",
        "D063",
        "D109",
        "D112",
        "D135",
        "D199",
        "D224",
    ]
    source_records: list[dict] = []
    for series_index in range(125):
        is_case_series = series_index < 107
        if is_case_series:
            disease_id = "D033"
            observation_count = 71 if series_index < 11 else 70
            metric_type = "case_notifications"
            reporting_basis = "national_registry_notifications"
        else:
            source_only_index = series_index - 107
            disease_id = source_only_diseases[source_only_index // 2]
            observation_count = 106 if source_only_index < 5 else 105
            metric_type = "registered_diagnoses"
            reporting_basis = "primary_care_icd_registered_diagnoses"
        series_code = f"SER_IS_REGRESSION_{series_index:03d}"
        for observation_index in range(observation_count):
            year, month_index = divmod(observation_index, 12)
            month = month_index + 1
            value_date = f"{1997 + year:04d}-{month:02d}-01"
            source_records.append(
                {
                    "disease_id": disease_id,
                    "date": value_date,
                    "year_month": value_date[:7],
                    "cases": observation_index,
                    "deaths": 0,
                    "recoveries": 0,
                    "incidence_rate": None,
                    "incidence_rate_source": "missing_population",
                    "mortality_rate": None,
                    "data_quality": "validated",
                    "quality_status": "validated",
                    "series_code": series_code,
                    "source_system": (
                        "SRC_IS_DOH_HISTORY"
                        if is_case_series else "SRC_IS_DOH_LEGACY_ICD"
                    ),
                    "source_series_code": series_code.lower(),
                    "source_label": "Iceland Directorate of Health",
                    "definition_version": "IS_REGRESSION",
                    "case_definition": "Regression definition",
                    "case_definition_uri": None,
                    "metric_type": metric_type,
                    "reporting_basis": reporting_basis,
                    "temporal_granularity": "monthly",
                    "series_unit": "count",
                    "observation_unit": "count",
                    "mapping_relation": "exact",
                    "comparability": (
                        "conditional" if is_case_series else "not_comparable"
                    ),
                    "aggregation_policy": "non_additive",
                    "availability_status": "available",
                    "missing_value_policy": "missing_is_unknown",
                    "valid_from": None,
                    "valid_to": None,
                    "series_is_active": series_index == 0,
                    "geography_key": "country:IS:national",
                    "dimension_key": "all",
                }
            )

    assert len(source_records) == 9396
    projected = apply_series_first_projection([], source_records)
    diseases = {
        disease_id: {
            "name_en": disease_id,
            "name_zh": disease_id,
            "category": "Test",
            "slug": disease_id.lower(),
        }
        for disease_id in ["D033", *source_only_diseases]
    }
    country = build_country_data(
        "IS",
        "Iceland",
        projected,
        diseases,
        source_records=source_records,
    )

    retained_series = [
        source_series
        for disease in country["disease_series"].values()
        for source_series in disease["source_series"]
    ]
    assert len(retained_series) == 125
    assert sum(item["observation_count"] for item in retained_series) == 9396
    assert sum(
        item["observation_count"]
        for item in retained_series
        if item["metric_type"] == "registered_diagnoses"
    ) == 1895
    assert len(
        {
            item["series_code"]
            for item in retained_series
            if item["metric_type"] == "registered_diagnoses"
        }
    ) == 18
    for disease_id in source_only_diseases:
        assert country["disease_series"][disease_id]["dates"] == []
        assert country["disease_series"][disease_id]["projection_policy"] == (
            "no_eligible_public_projection"
        )

    compact = build_country_site_data(country)
    assert {item["id"] for item in compact["series"]} == {"D033"}

    rows = build_country_download_rows(
        country,
        {
            "primary_scope": "is_doh_history",
            "primary_label": "Iceland Directorate of Health",
            "primary_url": "https://island.is/s/sottvarnir",
            "primary_type": "official_dashboard",
            "sources": [],
        },
    )
    source_rows = [
        row for row in rows if row["record_kind"] == "source_series_observation"
    ]
    projection_rows = [
        row for row in rows if row["record_kind"] == "public_projection"
    ]
    assert len(source_rows) == 9396
    assert len({row["series_code"] for row in source_rows}) == 125
    registered_rows = [
        row for row in source_rows if row["metric_type"] == "registered_diagnoses"
    ]
    assert len(registered_rows) == 1895
    assert all(row["cases"] is None for row in registered_rows)
    assert {row["disease_id"] for row in projection_rows} == {"D033"}


def test_historical_partition_bytes_stay_stable_when_current_data_changes(tmp_path) -> None:
    base = "https://raw.githubusercontent.com/example/data/main"
    first = build_direct_download_files(_context(), tmp_path, download_url_base=base)
    historical = next(
        part for part in first["diseases"][0]["parts"] if part["id"] == "2020-2025"
    )
    before = {
        format_name: (tmp_path / historical["files"][format_name]["relative_path"]).read_bytes()
        for format_name in ("csv", "json", "xlsx")
    }

    second_context = _context(
        ["2010-01-01", "2015-01-01", "2020-01-01", "2026-01-01", "2027-01-01"],
        generated_at="2027-01-01T00:00:00+00:00",
    )
    second = build_direct_download_files(second_context, tmp_path, download_url_base=base)
    historical_after = next(
        part for part in second["diseases"][0]["parts"] if part["id"] == "2020-2025"
    )
    for format_name in ("csv", "json", "xlsx"):
        path = tmp_path / historical_after["files"][format_name]["relative_path"]
        assert path.read_bytes() == before[format_name]


def test_oversized_calendar_window_is_split_below_target(tmp_path) -> None:
    dates = [f"2026-{month:02d}-01" for month in range(1, 13)] + [
        f"2027-{month:02d}-01" for month in range(1, 13)
    ]
    manifest = build_direct_download_files(
        _context(dates, dataset_name="Hepatitis A " + "x" * 1000),
        tmp_path,
        download_url_base="https://raw.githubusercontent.com/example/data/main",
        max_file_bytes=12_000,
    )
    parts = manifest["diseases"][0]["parts"]
    assert len(parts) > 1
    assert all(
        file_meta["bytes"] < 12_000
        for part in parts
        for file_meta in part["files"].values()
    )


def test_direct_download_files_reject_non_raw_base(tmp_path) -> None:
    with pytest.raises(ValueError, match="GitHub Raw"):
        build_direct_download_files(
            _context(),
            tmp_path,
            download_url_base="/downloads",
        )
