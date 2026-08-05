from __future__ import annotations

import csv
from datetime import date
from io import BytesIO, StringIO
from unittest.mock import AsyncMock

import pytest
from openpyxl import Workbook, load_workbook

from scripts.us_hiv_data_quality_check import AIDS_LABEL, HIV_LABEL, check_history
from scripts.us_prepare_hiv_history import parse_atlas_history
from src.core.source_scopes import (
    canonicalize_task_source,
    get_expected_scopes_for_country,
    scope_from_data_source,
)
from src.data.crawlers.us import (
    NHSS_HIV_LABEL,
    USNHSSHIVCrawler,
    USNNDSSCrawler,
)
from src.data.processors.us import USWeeklyUpdater


def _workbook_payload() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Table A1a"
    sheet.cell(1, 1).value = (
        "Table A1a. HIV diagnoses among persons aged ≥13 years, "
        "by selected characteristics, 2022–2024—United States"
    )
    sheet.cell(2, 2).value = 2022
    sheet.cell(2, 4).value = 2023
    sheet.cell(2, 6).value = "2024 (provisional)"
    sheet.cell(3, 2).value = "No."
    sheet.cell(3, 3).value = "Rate"
    sheet.cell(3, 4).value = "No."
    sheet.cell(3, 5).value = "Rate"
    sheet.cell(3, 6).value = "No."
    sheet.cell(3, 7).value = "Rate"
    sheet.cell(8, 1).value = "Totalh"
    sheet.cell(8, 2).value = 37_685
    sheet.cell(8, 3).value = 13.3
    sheet.cell(8, 4).value = 38_831
    sheet.cell(8, 5).value = 13.6
    sheet.cell(8, 6).value = 38_434
    sheet.cell(8, 7).value = 13.3
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_nhss_current_workbook_parser_extracts_national_total() -> None:
    rows = USNHSSHIVCrawler.parse_current_workbook(
        _workbook_payload(),
        source_url="https://www.cdc.gov/example.xlsx",
    )

    assert [row["SurveillanceYear"] for row in rows] == ["2022", "2023", "2024"]
    assert [row["Cases"] for row in rows] == ["37685", "38831", "38434"]
    assert rows[-1]["Diseases"] == NHSS_HIV_LABEL
    assert rows[-1]["Frequency"] == "annual"
    assert rows[-1]["ReleaseYear"] == "2024"
    assert rows[-1]["IsProvisional"] == "true"
    assert rows[0]["IsProvisional"] == "false"


def test_nhss_release_workbook_url_is_discovered_from_stable_page(monkeypatch) -> None:
    class FakeResponse:
        text = (
            '<a href="/hiv-data/media/files/2027/04/'
            'hiv_surveillance_data_release_tables_2025.xlsx">Download</a>'
        )
        url = "https://www.cdc.gov/hiv-data/nhss/example.html"

    crawler = USNHSSHIVCrawler(delay=0)
    monkeypatch.setattr(crawler, "get", lambda _url: FakeResponse())

    assert crawler.discover_current_workbook_url() == (
        "https://www.cdc.gov/hiv-data/media/files/2027/04/"
        "hiv_surveillance_data_release_tables_2025.xlsx"
    )


def test_nhss_release_discovery_prefers_newest_workbook(monkeypatch) -> None:
    class FakeResponse:
        text = """
            <a href="/files/hiv_surveillance_data_release_tables_2023.xlsx">2023</a>
            <a href="/files/hiv_surveillance_data_release_tables_2025.xlsx">2025</a>
            <a href="/files/hiv_surveillance_data_release_tables_2024.xlsx">2024</a>
        """
        url = "https://www.cdc.gov/hiv-data/nhss/releases.html"

    crawler = USNHSSHIVCrawler(delay=0)
    monkeypatch.setattr(crawler, "get", lambda _url: FakeResponse())

    assert crawler.discover_current_workbook_url().endswith(
        "hiv_surveillance_data_release_tables_2025.xlsx"
    )


def test_nhss_current_workbook_rejects_non_contiguous_years() -> None:
    payload = _workbook_payload()
    workbook = load_workbook(BytesIO(payload))
    sheet = workbook["Table A1a"]
    sheet.cell(2, 4).value = None
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(RuntimeError, match="Non-contiguous surveillance years"):
        USNHSSHIVCrawler.parse_current_workbook(
            output.getvalue(),
            source_url="https://www.cdc.gov/example.xlsx",
        )


def test_nndss_crawler_requests_total_and_us_residents(monkeypatch) -> None:
    requested_urls: list[str] = []

    class FakeResponse:
        text = (
            "states,year,week,label,m1,sort_order\n"
            "TOTAL,2024,1,Anthrax,2,1\n"
            "U.S. Residents,2024,1,Anthrax,1,1\n"
        )

    crawler = USNNDSSCrawler(delay=0)

    def fake_get(url: str):
        requested_urls.append(url)
        return FakeResponse()

    monkeypatch.setattr(crawler, "get", fake_get)

    rows, _source_url = crawler.fetch_raw_pages()

    assert {row["states"] for row in rows} == {"TOTAL", "U.S. Residents"}
    assert len(requested_urls) == 1
    assert "'TOTAL'" in requested_urls[0]
    assert "'US RESIDENTS'" in requested_urls[0]
    assert "'U.S. RESIDENTS'" in requested_urls[0]


def test_us_updater_all_combines_weekly_nndss_and_annual_nhss(monkeypatch) -> None:
    monkeypatch.setattr(
        "src.data.processors.us.USNNDSSCrawler.fetch_raw_pages",
        lambda _self: (
            [
                {
                    "states": "TOTAL",
                    "year": "2024",
                    "week": "1",
                    "label": "Anthrax",
                    "m1": "1",
                    "sort_order": "1",
                },
                {
                    "states": "U.S. Residents",
                    "year": "2024",
                    "week": "1",
                    "label": "Anthrax",
                    "m1": "2",
                    "sort_order": "1",
                }
            ],
            "https://data.cdc.gov/nndss.csv",
        ),
    )
    monkeypatch.setattr(
        "src.data.processors.us.USNHSSHIVCrawler.fetch_national_annual_rows",
        lambda _self: (
            [
                {
                    "Date": "2024-12-31",
                    "Diseases": NHSS_HIV_LABEL,
                    "DiseasesCN": NHSS_HIV_LABEL,
                    "Cases": "38434",
                    "Deaths": "",
                    "Source": "US CDC NHSS",
                    "ReportingArea": "TOTAL",
                    "SurveillanceYear": "2024",
                    "RawDiseaseLabel": NHSS_HIV_LABEL,
                    "IsProvisional": "true",
                    "UpdateMode": "current_release_xlsx",
                    "Frequency": "annual",
                    "Measure": "hiv_diagnoses",
                    "PopulationScope": "persons_age_13_plus",
                    "__source_file": "https://www.cdc.gov/hiv.xlsx",
                }
            ],
            "https://www.cdc.gov/hiv.xlsx",
        ),
    )

    fetched = USWeeklyUpdater().fetch_latest(source="all")

    assert len(fetched.rows) == 2
    assert [
        (row["Source"], row["ReportingArea"], row["Cases"])
        for row in fetched.rows
    ] == [
        ("US CDC NNDSS", "U.S. Residents", "2"),
        ("US CDC NHSS", "TOTAL", "38434"),
    ]
    assert {
        (row["Source"], row["ReportingArea"], row["Cases"])
        for row in fetched.series_rows
    } == {
        ("US CDC NNDSS", "TOTAL", "1"),
        ("US CDC NNDSS", "U.S. Residents", "2"),
        ("US CDC NHSS", "TOTAL", "38434"),
    }
    assert fetched.rows[0]["PopulationScope"] == (
        "us_residents_excluding_territories"
    )
    assert next(
        row
        for row in fetched.series_rows
        if row["Source"] == "US CDC NNDSS" and row["ReportingArea"] == "TOTAL"
    )["PopulationScope"] == (
        "nndss_total_including_us_residents_territories_and_non_us_residents"
    )
    assert set(fetched.latest_by_source) == {"US CDC NNDSS", "US CDC NHSS"}
    assert fetched.latest_by_source["US CDC NHSS"].isoformat() == "2024-12-31"


def test_us_updater_refuses_total_fallback_when_residents_are_missing(
    monkeypatch,
) -> None:
    monkeypatch.setattr(
        "src.data.processors.us.USNNDSSCrawler.fetch_raw_pages",
        lambda _self: (
            [
                {
                    "states": "TOTAL",
                    "year": "2024",
                    "week": "1",
                    "label": "Anthrax",
                    "m1": "1",
                    "sort_order": "1",
                }
            ],
            "https://data.cdc.gov/nndss.csv",
        ),
    )

    with pytest.raises(RuntimeError, match="US RESIDENTS rows are missing"):
        USWeeklyUpdater().fetch_latest(source="nndss_api")


@pytest.mark.parametrize("resident_label", ["US RESIDENTS", "U.S. Residents"])
def test_us_updater_accepts_historical_and_current_resident_aliases(
    monkeypatch,
    resident_label: str,
) -> None:
    monkeypatch.setattr(
        "src.data.processors.us.USNNDSSCrawler.fetch_raw_pages",
        lambda _self: (
            [
                {
                    "states": resident_label,
                    "year": "2024",
                    "week": "1",
                    "label": "Anthrax",
                    "m1": "1",
                    "sort_order": "1",
                },
                {
                    "states": "TOTAL",
                    "year": "2024",
                    "week": "1",
                    "label": "Anthrax",
                    "m1": "2",
                    "sort_order": "1",
                },
            ],
            "https://data.cdc.gov/nndss.csv",
        ),
    )

    fetched = USWeeklyUpdater().fetch_latest(source="nndss_api")

    assert [row["ReportingArea"] for row in fetched.rows] == [resident_label]
    assert {row["ReportingArea"] for row in fetched.series_rows} == {
        resident_label,
        "TOTAL",
    }


def test_us_updater_rejects_total_as_legacy_national_scope() -> None:
    with pytest.raises(ValueError, match="must use US RESIDENTS"):
        USWeeklyUpdater(reporting_area="TOTAL")


def test_us_updater_rejects_unknown_nndss_reporting_area(monkeypatch) -> None:
    monkeypatch.setattr(
        "src.data.processors.us.USNNDSSCrawler.fetch_raw_pages",
        lambda _self: (
            [
                {
                    "states": "UNKNOWN AGGREGATE",
                    "year": "2024",
                    "week": "1",
                    "label": "Anthrax",
                    "m1": "1",
                    "sort_order": "1",
                }
            ],
            "https://data.cdc.gov/nndss.csv",
        ),
    )

    with pytest.raises(ValueError, match="Unsupported NNDSS ReportingArea"):
        USWeeklyUpdater().fetch_latest(source="nndss_api")


async def test_import_gate_does_not_let_weekly_latest_date_block_annual_hiv() -> None:
    class FakeDB:
        def __init__(self) -> None:
            self.inserted: list[dict[str, object]] = []

        async def execute(self, _statement, parameters=None):
            if isinstance(parameters, list):
                self.inserted.extend(parameters)
            return None

    updater = USWeeklyUpdater()
    updater._get_country_id = AsyncMock(return_value=1)
    updater._load_mapping_dict = AsyncMock(
        return_value={
            "anthrax": 1,
            NHSS_HIV_LABEL.lower(): 162,
        }
    )
    updater._get_source_latest_dates = AsyncMock(
        return_value={
            "US CDC NNDSS": date(2026, 7, 25),
            "US CDC NHSS": date(2024, 12, 31),
        }
    )
    rows = [
        {
            "Date": "2020-01-04",
            "Diseases": "Anthrax",
            "RawDiseaseLabel": "Anthrax",
            "Cases": "1",
            "Source": "US CDC NNDSS",
            "Frequency": "weekly",
        },
        {
            "Date": "2026-07-18",
            "Diseases": "Anthrax",
            "RawDiseaseLabel": "Anthrax",
            "Cases": "2",
            "Source": "US CDC NNDSS",
            "Frequency": "weekly",
        },
        {
            "Date": "2024-12-31",
            "Diseases": NHSS_HIV_LABEL,
            "RawDiseaseLabel": NHSS_HIV_LABEL,
            "Cases": "38434",
            "Source": "US CDC NHSS",
            "Frequency": "annual",
            "Measure": "hiv_diagnoses",
        },
    ]
    db = FakeDB()

    result = await updater.import_rows(
        db,
        rows,
        db_latest_date=date(2026, 7, 25),
        source_latest_date=date(2026, 7, 25),
    )

    assert result.inserted_or_updated == 2
    assert {(row["disease_id"], row["data_source"]) for row in db.inserted} == {
        (1, "US CDC NNDSS"),
        (162, "US CDC NHSS"),
    }
    hiv_row = next(row for row in db.inserted if row["disease_id"] == 162)
    assert hiv_row["deaths"] is None


@pytest.mark.asyncio
async def test_import_skips_missing_cases_without_fabricating_zero() -> None:
    class FakeDB:
        def __init__(self) -> None:
            self.inserted: list[dict[str, object]] = []

        async def execute(self, _statement, parameters=None):
            if isinstance(parameters, list):
                self.inserted.extend(parameters)
            return None

    updater = USWeeklyUpdater()
    updater._get_country_id = AsyncMock(return_value=1)
    updater._load_mapping_dict = AsyncMock(return_value={"anthrax": 1})
    updater._get_source_latest_dates = AsyncMock(return_value={})
    rows = [
        {
            "Date": "2024-01-06",
            "RawDiseaseLabel": "Anthrax",
            "Cases": missing_value,
            "Source": "US CDC NNDSS",
        }
        for missing_value in ("", "-", "NaN")
    ]
    rows.append(
        {
            "Date": "2024-01-13",
            "RawDiseaseLabel": "Anthrax",
            "Cases": "0",
            "Source": "US CDC NNDSS",
        }
    )
    db = FakeDB()

    result = await updater.import_rows(
        db,
        rows,
        db_latest_date=None,
        source_latest_date=date(2024, 1, 13),
        force=True,
    )

    assert result.inserted_or_updated == 1
    assert len(db.inserted) == 1
    assert db.inserted[0]["cases"] == 0


def test_atlas_parser_keeps_only_unstratified_national_rows() -> None:
    buffer = StringIO()
    buffer.write("Title: historical extract\n\n")
    fieldnames = [
        "Indicator",
        "Year",
        "Geography",
        "Age Group",
        "Race/Ethnicity",
        "Sex",
        "Transmission Category",
        "Cases",
        "Rate per 100000",
        "Rate LCI",
        "Rate UCI",
        "Population",
    ]
    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
    writer.writeheader()
    common = {
        "Year": "2019",
        "Geography": "US",
        "Age Group": "Ages 13 years and older",
        "Race/Ethnicity": "All races/ethnicities",
        "Sex": "All gender identities",
        "Transmission Category": "All transmission categories",
    }
    writer.writerow({**common, "Indicator": "HIV diagnoses", "Cases": "36,350"})
    writer.writerow({**common, "Indicator": "AIDS classifications", "Cases": "16410"})
    writer.writerow(
        {
            **common,
            "Indicator": "HIV diagnoses",
            "Sex": "Man",
            "Cases": "29482",
        }
    )

    rows = parse_atlas_history(buffer.getvalue())

    assert [(row["RawDiseaseLabel"], row["Cases"]) for row in rows] == [
        (HIV_LABEL, "36350"),
        (AIDS_LABEL, "16410"),
    ]


def test_hiv_history_quality_check_accepts_separate_continuous_series(tmp_path) -> None:
    path = tmp_path / "history.csv"
    with path.open("w", encoding="utf-8", newline="") as handle:
        fieldnames = [
            "Date",
            "Diseases",
            "Cases",
            "Deaths",
            "Source",
            "RawDiseaseLabel",
            "__source_file",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for year in (2022, 2023, 2024):
            writer.writerow(
                {
                    "Date": f"{year}-12-31",
                    "Diseases": HIV_LABEL,
                    "Cases": "38000",
                    "Deaths": "",
                    "Source": "US CDC NHSS",
                    "RawDiseaseLabel": HIV_LABEL,
                    "__source_file": "https://www.cdc.gov/hiv.xlsx",
                }
            )
        for year in (2022, 2023):
            writer.writerow(
                {
                    "Date": f"{year}-12-31",
                    "Diseases": AIDS_LABEL,
                    "Cases": "16000",
                    "Deaths": "",
                    "Source": "US CDC NHSS",
                    "RawDiseaseLabel": AIDS_LABEL,
                    "__source_file": "https://www.cdc.gov/atlas.csv",
                }
            )

    summary = check_history(path, expected_latest_year=2024)

    assert summary["status"] == "pass"
    assert summary["nhss_rows"] == 5


def test_us_source_scopes_include_nhss() -> None:
    assert get_expected_scopes_for_country("US") == ["nndss_api", "nhss_hiv"]
    assert canonicalize_task_source("nhss", country_code="US") == "nhss_hiv"
    assert scope_from_data_source("US CDC NHSS") == "nhss_hiv"
