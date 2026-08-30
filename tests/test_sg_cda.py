from __future__ import annotations

import io
from datetime import date
from pathlib import Path

import openpyxl

from src.core.country_library import get_country_bootstrap_config, get_country_profile
from src.core.source_scopes import canonicalize_task_source, get_expected_scopes_for_country
from src.data.crawlers import sg
from src.data.crawlers.sg import (
    DEFAULT_SOURCE_NAME,
    DISEASE_ALIASES,
    HISTORICAL_SOURCE_NAME,
    OPEN_DATA_POLL_URL,
    SingaporeCDACrawler,
    parse_annual_workbook,
    parse_historical_csv,
    parse_weekly_pdf,
    singapore_week_start,
)
from src.data.processors.sg import SGWeeklyUpdater
from src.data.storage import SeriesObservationStore
from src.services.crawl_service import CrawlService


def test_singapore_week_calendar_keeps_source_week_53() -> None:
    assert singapore_week_start(2025, 1) == date(2024, 12, 29)
    assert singapore_week_start(2025, 53) == date(2025, 12, 28)


def test_historical_csv_preserves_zero_and_does_not_fill_absent_rows() -> None:
    payload = (
        "epi_week,disease,no._of_cases\n"
        "2016-W37,Cholera,0\n"
        "2016-W37,HFMD,99\n"
        "2022-W01,Monkeypox,0\n"
        "2022-W26,Monkeypox,1\n"
    ).encode()
    rows = parse_historical_csv(payload, source_url="https://example.test/history.csv", retrieved_at="2026-01-01T00:00:00+00:00")
    assert [(row["EpiWeek"], row["SourceDiseaseCode"], row["Cases"]) for row in rows] == [
        ("2016-W37", "cholera", "0"),
        ("2022-W26", "mpox", "1"),
    ]
    assert all(row["Source"] == HISTORICAL_SOURCE_NAME for row in rows)
    assert all(row["PublicReleaseEnabled"] == "true" for row in rows)
    assert all(row["LicenseReviewStatus"] == "singapore_open_data_licence" for row in rows)
    assert not any(row["SourceDiseaseCode"] == "measles" for row in rows)


def test_historical_fetch_uses_stable_dataset_url_but_archives_signed_url(monkeypatch) -> None:
    payload = b"epi_week,disease,no._of_cases\n2022-W01,Cholera,1\n"
    signed_url = "https://example.test/" + "signed" * 300
    archived: list[str] = []

    class Response:
        def __init__(self, *, body: bytes = b"", json_body: dict | None = None):
            self.content = body
            self._json_body = json_body

        def json(self):
            return self._json_body

    crawler = SingaporeCDACrawler()
    monkeypatch.setattr(
        crawler,
        "get",
        lambda url: Response(json_body={"data": {"url": signed_url}})
        if url == OPEN_DATA_POLL_URL
        else Response(body=payload),
    )
    monkeypatch.setattr(
        crawler,
        "_archive",
        lambda _name, _payload, source_url: archived.append(source_url) or "",
    )

    rows = crawler._historical()

    assert rows[0]["SourceURL"] == OPEN_DATA_POLL_URL
    assert archived == [signed_url]


def _workbook_bytes() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(1, 1, "Epidemiology Wk")
    sheet.cell(1, 4, "Total Number")
    current = [(code, label) for code, (label, _) in DISEASE_ALIASES.items() if code != "encephalitis"]
    for column, (_, label) in enumerate(current, 4):
        sheet.cell(2, column, label)
        sheet.cell(3, column, 0)
    attendance = 4 + len(current)
    sheet.cell(1, attendance, "Average Daily Number")
    sheet.cell(2, attendance, "Acute Upper Respiratory Infections")
    sheet.cell(3, attendance, 12345)
    sheet.cell(3, 1, 53)
    sheet.cell(3, 2, "28/12/2025 - 03/01/2026")
    sheet.cell(4, 1, 1)
    sheet.cell(4, 2, "29/12/2024 - 04/01/2025")
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_xlsx_imports_only_notification_columns_and_skips_future_blank_week() -> None:
    rows = parse_annual_workbook(_workbook_bytes(), year=2025, source_url="https://example.test/2025.xlsx")
    assert len(rows) == 38
    assert {row["Cases"] for row in rows} == {"0"}
    assert {row["EpiWeek"] for row in rows} == {"2025-W53"}
    assert not any("respiratory infections" in row["RawDiseaseLabel"].casefold() for row in rows)
    assert all(row["Source"] == DEFAULT_SOURCE_NAME for row in rows)
    assert all(row["PublicReleaseEnabled"] == "true" for row in rows)
    assert all(row["LicenseReviewStatus"] == "operator_authorized_public_release" for row in rows)


def test_pdf_parser_stops_before_polyclinic_attendances(monkeypatch) -> None:
    table = [
        ["FOOD/WATER-BORNE DISEASES", None],
        ["Acute Hepatitis A", "0"],
        ["Dengue Haemorrhagic Fever", "2"],
        ["Middle East Respiratory Syndrome^", "0"],
        ["POLYCLINIC ATTENDANCES - AVERAGE DAILY NUMBER", None],
        ["Hand, Foot And Mouth Disease", "999"],
    ]

    class Page:
        def extract_tables(self):
            return [table]

    class Document:
        pages = [Page()]
        def __enter__(self):
            return self
        def __exit__(self, *args):
            return None

    monkeypatch.setattr(sg.pdfplumber, "open", lambda _: Document())
    rows = parse_weekly_pdf(b"pdf", year=2023, week=30, source_url="https://example.test/w30.pdf")
    assert [(row["SourceDiseaseCode"], row["Cases"]) for row in rows] == [
        ("acute_viral_hepatitis_a", "0"),
        ("dengue_haemorrhagic_fever", "2"),
        ("middle_east_respiratory_syndrome", "0"),
    ]


def test_country_pipeline_scope_permission_and_registry_contract() -> None:
    profile = get_country_profile("SG")
    config = get_country_bootstrap_config("SG")
    assert profile.timezone == "Asia/Singapore"
    assert config["public_release_enabled"] is True
    assert config["crawler_config"]["reuse_status"] == "operator_authorized_public_release"
    assert config["crawler_config"]["current_source_terms_status"] == "cda_written_permission_required"
    assert config["crawler_config"]["historical_csv_end_year"] == 2022
    assert get_expected_scopes_for_country("SG") == ["cda_weekly_bulletin"]
    assert canonicalize_task_source("all", country_code="SG") == "cda_weekly_bulletin"
    assert "SG" in CrawlService.supported_country_codes()
    updater = SGWeeklyUpdater(output_csv=Path("unused.csv"))
    assert updater.public_release_enabled is True
    assert isinstance(updater.ontology_source_id, dict)

    rows = [
        {"Date": "2022-12-25", "Source": HISTORICAL_SOURCE_NAME, "RawDiseaseLabel": "Cholera", "SourceDiseaseCode": "cholera", "Cases": "0"},
        {"Date": "2023-01-01", "Source": DEFAULT_SOURCE_NAME, "RawDiseaseLabel": "Cholera", "SourceDiseaseCode": "cholera", "Cases": "0"},
    ]
    selected = SeriesObservationStore().select_registry_rows(rows, "SG", source_id=updater.ontology_source_id)
    assert len(selected.rows) == 2
    assert selected.skipped_unregistered == 0
