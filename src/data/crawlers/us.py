"""US CDC surveillance crawlers.

The NNDSS crawler fetches weekly notifiable-disease data from Socrata. HIV is
published through a separate channel, so :class:`USNHSSHIVCrawler` discovers
and parses the current National HIV Surveillance System (NHSS) workbook.
"""
from __future__ import annotations

import csv
import io
import re
from typing import Dict, List, Tuple
from urllib.parse import urljoin, urlparse

from openpyxl import load_workbook

from src.core import get_logger
from .base import BaseCrawler, CrawlerResult

logger = get_logger(__name__)

DEFAULT_CSV_API_URL = (
    "https://data.cdc.gov/resource/x9gk-5huc.csv"
    "?$select=states,year,week,label,m1,m1_flag,m2,m2_flag,m3,m3_flag,m4,m4_flag,"
    "location1,location2,sort_order,geocode"
    "&$where=upper(states) in ('TOTAL','US RESIDENTS','U.S. RESIDENTS')"
    "&$order=sort_order,states"
)
SOCRATA_PAGE_SIZE = 50_000

NHSS_RELEASE_PAGE_URL = (
    "https://www.cdc.gov/hiv-data/nhss/hiv-diagnoses-deaths-prevalence.html"
)
NHSS_HISTORIC_ATLAS_URL = (
    "https://www.cdc.gov/nchhstp/media/files/2025/05/"
    "AtlasPlus_historic-data-extract_20250430_1.csv"
)
NHSS_SOURCE_NAME = "US CDC NHSS"
NHSS_HIV_LABEL = "HIV diagnoses among persons aged 13 years and older"
NHSS_USER_AGENT = "GlobalIDBot/2.0 (+https://global.id)"


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\ufeff", "").split())


def _parse_year(value: object) -> int | None:
    match = re.search(r"\b(19|20)\d{2}\b", _normalize_text(value))
    return int(match.group(0)) if match else None


def _release_year_from_workbook_url(url: str) -> int | None:
    """Return the surveillance year encoded in an NHSS workbook filename."""

    filename = urlparse(url).path.rsplit("/", 1)[-1]
    match = re.search(r"_((?:19|20)\d{2})\.xlsx$", filename, re.IGNORECASE)
    return int(match.group(1)) if match else None


def _parse_count(value: object) -> int | None:
    text = _normalize_text(value).replace(",", "")
    if not text or text in {"—", "-", "Data not available", "NULL"}:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _parse_rate(value: object) -> str:
    text = _normalize_text(value).replace(",", "")
    if not text or text in {"—", "-", "Data not available", "NULL"}:
        return ""
    try:
        return str(float(text)).rstrip("0").rstrip(".")
    except (TypeError, ValueError):
        return ""


class USNNDSSCrawler(BaseCrawler):
    """Fetch US CDC NNDSS weekly data from the Socrata Open Data API (CSV endpoint).

    Usage::

        crawler = USNNDSSCrawler()
        raw_rows, source_url = crawler.fetch_raw_pages()
        # raw_rows: List[Dict[str, str]] — one dict per CSV row, unfiltered
    """

    def fetch_raw_pages(self) -> Tuple[List[Dict[str, object]], str]:
        """Fetch all paginated rows from the CDC Socrata CSV endpoint.

        Returns:
            (rows, source_url) — rows is the raw DictReader output, source_url is
            the base API URL for provenance tracking.

        Raises:
            RuntimeError: if the API returns no rows at all.
        """
        all_rows: List[Dict[str, object]] = []
        offset = 0

        while True:
            page_url = f"{DEFAULT_CSV_API_URL}&$limit={SOCRATA_PAGE_SIZE}&$offset={offset}"
            response = self.get(page_url)
            page_rows = list(csv.DictReader(io.StringIO(response.text)))
            if not page_rows:
                break
            all_rows.extend(page_rows)
            if len(page_rows) < SOCRATA_PAGE_SIZE:
                break
            offset += SOCRATA_PAGE_SIZE

        if not all_rows:
            raise RuntimeError("[US-NNDSS] Socrata API returned no rows")

        logger.info(f"[US-NNDSS] Fetch complete | rows={len(all_rows)} source={DEFAULT_CSV_API_URL}")
        return all_rows, DEFAULT_CSV_API_URL

    # ── BaseCrawler abstract method stubs ─────────────────────────────────────
    # US data flows through fetch_raw_pages() → processors/us.py, not the
    # generic crawl() / parse() interface used by CN.

    def crawl(self) -> List[CrawlerResult]:
        """Not used for US; call fetch_raw_pages() directly."""
        return []

    def parse(self, response) -> List[CrawlerResult]:
        """Not used for US."""
        return []


class USNHSSHIVCrawler(BaseCrawler):
    """Fetch the current CDC NHSS national HIV diagnosis workbook.

    The workbook URL is intentionally discovered from the stable CDC release
    page because CDC publishes each release under a dated media path.
    """

    _WORKBOOK_LINK_RE = re.compile(
        r"href=[\"']([^\"']*hiv_surveillance_data_release_tables_[^\"']*\.xlsx(?:\?[^\"']*)?)[\"']",
        re.IGNORECASE,
    )

    def __init__(self, *args, user_agent: str | None = None, **kwargs):
        # CDC's Akamai policy rejects the generic browser-compatible crawler
        # identity used by BaseCrawler.  Use a transparent, contactable bot
        # identity for this provider while retaining explicit caller overrides.
        super().__init__(
            *args,
            user_agent=user_agent or NHSS_USER_AGENT,
            **kwargs,
        )

    def discover_current_workbook_url(self) -> str:
        response = self.get(NHSS_RELEASE_PAGE_URL)
        matches = self._WORKBOOK_LINK_RE.findall(response.text)
        if not matches:
            raise RuntimeError(
                "[US-NHSS] Current HIV surveillance workbook link was not found "
                f"on {NHSS_RELEASE_PAGE_URL}"
            )
        workbook_urls = {
            urljoin(response.url or NHSS_RELEASE_PAGE_URL, match) for match in matches
        }
        # CDC release pages can retain links to previous releases. Prefer the
        # workbook with the newest surveillance year instead of relying on HTML
        # link order, which is not a stable publication contract.
        return max(
            workbook_urls,
            key=lambda url: (_release_year_from_workbook_url(url) or -1, url),
        )

    def fetch_current_workbook(self) -> Tuple[bytes, str]:
        workbook_url = self.discover_current_workbook_url()
        response = self.get(workbook_url)
        if not response.content:
            raise RuntimeError("[US-NHSS] Current HIV surveillance workbook is empty")
        return response.content, workbook_url

    @staticmethod
    def parse_current_workbook(
        payload: bytes,
        *,
        source_url: str,
    ) -> List[Dict[str, object]]:
        """Extract revised national annual HIV diagnoses from Table A1a.

        Table A1a is used because it is explicitly limited to the United States,
        represents persons aged 13 years and older, and provides the longest
        revised trend in the current release. The ``Total`` row avoids summing
        overlapping demographic categories.
        """

        try:
            workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        except Exception as exc:
            raise RuntimeError(f"[US-NHSS] Invalid HIV surveillance workbook: {exc}") from exc

        try:
            if "Table A1a" not in workbook.sheetnames:
                raise RuntimeError("[US-NHSS] Expected sheet 'Table A1a' is missing")

            sheet = workbook["Table A1a"]
            title = _normalize_text(sheet.cell(row=1, column=1).value)
            title_lower = title.lower()
            if "hiv diagnoses" not in title_lower or "united states" not in title_lower:
                raise RuntimeError(f"[US-NHSS] Unexpected Table A1a title: {title}")

            year_columns: List[Tuple[int, int]] = []
            for column in range(1, sheet.max_column + 1):
                year = _parse_year(sheet.cell(row=2, column=column).value)
                if year is not None:
                    year_columns.append((year, column))
            if not year_columns:
                raise RuntimeError("[US-NHSS] No surveillance years found in Table A1a")

            year_columns.sort()
            years = [year for year, _ in year_columns]
            if len(years) != len(set(years)):
                raise RuntimeError("[US-NHSS] Duplicate surveillance years in Table A1a")
            expected_years = list(range(years[0], years[-1] + 1))
            if years != expected_years:
                raise RuntimeError(
                    "[US-NHSS] Non-contiguous surveillance years in Table A1a: "
                    f"{years}"
                )

            latest_year = years[-1]
            release_year = _release_year_from_workbook_url(source_url)
            if release_year is not None and release_year != latest_year:
                raise RuntimeError(
                    "[US-NHSS] Workbook filename/table year mismatch: "
                    f"filename={release_year}, table={latest_year}"
                )

            total_row: int | None = None
            for row_number in range(1, sheet.max_row + 1):
                label = _normalize_text(
                    sheet.cell(row=row_number, column=1).value
                ).lower()
                if label.startswith("total"):
                    total_row = row_number
                    break
            if total_row is None:
                raise RuntimeError("[US-NHSS] National Total row is missing from Table A1a")

            rows: List[Dict[str, object]] = []
            for year, count_column in year_columns:
                cases = _parse_count(sheet.cell(row=total_row, column=count_column).value)
                if cases is None:
                    raise RuntimeError(f"[US-NHSS] Missing national HIV count for {year}")
                rate = _parse_rate(
                    sheet.cell(row=total_row, column=count_column + 1).value
                )
                rows.append(
                    {
                        "Date": f"{year}-12-31",
                        "Diseases": NHSS_HIV_LABEL,
                        "DiseasesCN": NHSS_HIV_LABEL,
                        "Cases": str(cases),
                        "Deaths": "",
                        "Incidence": rate,
                        "Source": NHSS_SOURCE_NAME,
                        "CountryCode": "US",
                        "ReportingArea": "TOTAL",
                        "SurveillanceYear": str(year),
                        "ReleaseYear": str(latest_year),
                        "RawDiseaseLabel": NHSS_HIV_LABEL,
                        "IsProvisional": "true" if year == latest_year else "false",
                        "UpdateMode": "current_release_xlsx",
                        "Frequency": "annual",
                        "Measure": "hiv_diagnoses",
                        "PopulationScope": "persons_age_13_plus",
                        "__source_file": source_url,
                    }
                )
        finally:
            workbook.close()

        logger.info(
            f"[US-NHSS] Parsed current HIV release | rows={len(rows)} "
            f"years={rows[0]['SurveillanceYear']}-{rows[-1]['SurveillanceYear']}"
        )
        return rows

    def fetch_national_annual_rows(self) -> Tuple[List[Dict[str, object]], str]:
        payload, source_url = self.fetch_current_workbook()
        return self.parse_current_workbook(payload, source_url=source_url), source_url

    def crawl(self) -> List[CrawlerResult]:
        """Not used for US; call fetch_national_annual_rows() directly."""
        return []

    def parse(self, response) -> List[CrawlerResult]:
        """Parse through the structured helper; the generic result API is unused."""
        return []
