"""Public Health Ontario IDTO monthly preliminary-data connector.

The public IDTO report is an embedded Power BI report.  The live path obtains a
short-lived embed token from PHO's public wrapper, discovers the current model,
page, visual, and query, then executes that visual's read-only query. Both the
legacy Power BI visual-container metadata and the newer Fabric PBIR definition
are supported. No token or response cookie is persisted; public report/model
identifiers are retained in raw provenance when archival is enabled.

For auditable replay and deployments where the undocumented Power BI data plane
is unavailable, the connector also accepts an official CSV/XLSX export through
an explicit ``input_file`` argument. ``CA_ON_IDTO_FILE`` is consulted only when
the caller explicitly opts into configured-file replay for that run; merely
leaving the environment variable set cannot change a live scheduled run.
"""

from __future__ import annotations

import base64
import copy
import csv
import hashlib
import io
import json
import os
import re
import time
import uuid
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional, Sequence
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

from src.core import get_logger
from src.core.country_library import get_country_bootstrap_config

from .base import BaseCrawler, CrawlerResult

logger = get_logger(__name__)

DEFAULT_SOURCE_NAME = "Public Health Ontario IDTO Monthly Preliminary Data"
DEFAULT_LANDING_URL = (
    "https://www.publichealthontario.ca/en/data-and-analysis/"
    "infectious-disease/reportable-disease-trends-annually"
)
DEFAULT_REPORT_ID = "14b5691a-c95d-46b2-84f1-9119080e083b"
DEFAULT_EMBED_URL = (
    "https://ws-rpt1.publichealthontario.ca/Home/EmbedReport/" + DEFAULT_REPORT_ID
)
DEFAULT_PAGE_DISPLAY_NAME = "Monthly Data Table"
DEFAULT_VISUAL_NAME = "8533f6960c0f199b51ae"
# Ontario is published as its own top-level jurisdiction in GlobalID, in the
# same way that Hong Kong has an independent country/region dataset.  The raw
# ISO subdivision identity remains explicit on every row via ``Geocode`` and
# ``ParentCountryCode``.
ONTARIO_GEOGRAPHY_KEY = "country:CA-ON:national"
ONTARIO_GEOCODE = "CA-ON"

MAX_EXPORT_BYTES = 25 * 1024 * 1024
MAX_EXPORT_ROWS = 10_000
MAX_EXPORT_COLUMNS = 100
MAX_EXPORT_CELLS = 500_000
MAX_XLSX_UNCOMPRESSED_BYTES = 150 * 1024 * 1024

DEFINITION_VERSION_BY_YEAR = {2026: "PHO_IDTO_2026_07"}

SPECIAL_TIME_BASES = {
    "acquired immunodeficiency syndrome (aids)": (
        "PHO AIDS Diagnosis Status Date"
    ),
    "hiv": "PHO HIV Encounter Date (Reported Date)",
    "tuberculosis": "PHO tuberculosis Diagnosis Date",
    "carbapenemase-producing enterobacteriaceae (cpe)": (
        "PHO earliest specimen collection date"
    ),
    "candida auris infection": "PHO earliest specimen collection date",
}

TOKEN_RE = re.compile(r"accessToken\s*=\s*[\"']([^\"']+)[\"']")
REPORT_ID_RE = re.compile(r"embedReportId\s*=\s*[\"']([^\"']+)[\"']")
YEAR_RE = re.compile(r"(?<!\d)(20\d{2})(?!\d)")
SUPPRESSION_RE = re.compile(r"^<\s*(\d+)\s*$")

MONTHS: tuple[tuple[str, int], ...] = (
    ("Jan", 1),
    ("Feb", 2),
    ("Mar", 3),
    ("Apr", 4),
    ("May", 5),
    ("Jun", 6),
    ("Jul", 7),
    ("Aug", 8),
    ("Sep", 9),
    ("Oct", 10),
    ("Nov", 11),
    ("Dec", 12),
)

MONTH_NAMES: dict[str, int] = {}
for _month_short, _month_number in MONTHS:
    MONTH_NAMES[_month_short.casefold()] = _month_number
    MONTH_NAMES[datetime(2000, _month_number, 1).strftime("%B").casefold()] = (
        _month_number
    )

OUTPUT_FIELDS = (
    "Date",
    "RawDiseaseLabel",
    "DiseaseCode",
    "Year",
    "Month",
    "Cases",
    "YearToDateCases",
    "YearToDateRatePer100000",
    "ReportingArea",
    "Geocode",
    "JurisdictionCode",
    "ParentCountryCode",
    "LocationType",
    "GeographyKey",
    "PopulationScope",
    "DatasetStatus",
    "IsProvisional",
    "AuthoritativeRevision",
    "AllowEqualQualityOverwrite",
    "TimeBasis",
    "DefinitionVersion",
    "DatasetTimestamp",
    "ModelRefreshTime",
    "RetrievedAt",
    "AcquisitionMode",
    "Source",
    "SourceURL",
)


@dataclass(frozen=True)
class EmbedContext:
    token: str = field(repr=False)
    cluster_url: str
    expires_at: int
    activity_id: str
    report_id: str


@dataclass(frozen=True)
class MonthlyVisual:
    model_id: int
    dataset_id: str
    query: dict[str, Any]
    page_name: str
    visual_name: str
    title: str
    model_refresh_time: str


@dataclass(frozen=True)
class CAOntarioFetchSummary:
    row_count: int
    disease_count: int
    latest_date: Optional[date]
    reporting_year: int
    source_url: str
    acquisition_mode: str
    content_sha256: str
    dataset_timestamp: str = ""
    model_refresh_time: str = ""
    populated_month_count: int = 0
    unpublished_month_slots: int = 0
    source_artifact_sha256: str = ""
    source_file_mtime: str = ""


def _norm_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\ufeff", "").split()).strip()


def _norm_header(value: object) -> str:
    text = _norm_text(value).casefold().replace("&", " and ")
    return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())


def _json_object(value: object, *, label: str) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        parsed = json.loads(value)
        if isinstance(parsed, dict):
            return parsed
    raise ValueError(f"{label} must be a JSON object")


def _decode_routing_payload(token: str) -> dict[str, Any]:
    parts = token.rsplit(".", 1)
    if len(parts) != 2:
        raise ValueError("Power BI embed token has no routing payload")
    encoded = parts[1] + "=" * (-len(parts[1]) % 4)
    try:
        payload = json.loads(base64.urlsafe_b64decode(encoded))
    except (ValueError, json.JSONDecodeError) as exc:
        raise ValueError("Power BI embed token routing payload is invalid") from exc
    if not isinstance(payload, dict):
        raise ValueError("Power BI embed token routing payload is not an object")
    return payload


def extract_embed_context(html: str, *, expected_report_id: str) -> EmbedContext:
    """Extract an in-memory Power BI token and validate its routing host."""

    token_match = TOKEN_RE.search(html)
    if token_match is None:
        raise RuntimeError("PHO IDTO Power BI embed token was not found")
    token = token_match.group(1)
    route = _decode_routing_payload(token)

    cluster_url = _norm_text(route.get("clusterUrl")).rstrip("/")
    parsed = urlparse(cluster_url)
    if (
        parsed.scheme != "https"
        or not parsed.hostname
        or not parsed.hostname.casefold().endswith(".analysis.windows.net")
        or parsed.username
        or parsed.password
    ):
        raise RuntimeError("PHO IDTO returned an unexpected Power BI cluster host")

    try:
        expires_at = int(route["exp"])
    except (KeyError, TypeError, ValueError) as exc:
        raise RuntimeError("PHO IDTO embed token has no valid expiry") from exc
    if expires_at <= int(time.time()):
        raise RuntimeError("PHO IDTO embed token is already expired")

    report_match = REPORT_ID_RE.search(html)
    report_id = report_match.group(1) if report_match else expected_report_id
    if report_id != expected_report_id:
        raise RuntimeError(
            "PHO IDTO embed report id changed; configuration review is required"
        )
    return EmbedContext(
        token=token,
        cluster_url=cluster_url,
        expires_at=expires_at,
        activity_id=str(uuid.uuid4()),
        report_id=report_id,
    )


def powerbi_headers(context: EmbedContext, *, post: bool = False) -> dict[str, str]:
    headers = {
        "Accept": "application/json",
        "ActivityId": context.activity_id,
        "RequestId": str(uuid.uuid4()),
        "Authorization": "EmbedToken " + context.token,
        "X-PowerBI-HostEnv": "Embed for Customers",
    }
    if post:
        headers["Content-Type"] = "application/json"
    return headers


def _visual_title(config: Mapping[str, Any]) -> str:
    try:
        value = config["singleVisual"]["vcObjects"]["title"][0]["properties"][
            "text"
        ]["expr"]["Literal"]["Value"]
    except (KeyError, IndexError, TypeError):
        return ""
    text = _norm_text(value)
    if len(text) >= 2 and text[0] == text[-1] == "'":
        text = text[1:-1]
    return text.replace("''", "'")


def _transform_query_tree(value: Any, transform: Any) -> Any:
    """Deep-copy a Power BI query fragment while transforming each object."""

    if isinstance(value, Mapping):
        transformed = {
            str(key): _transform_query_tree(item, transform)
            for key, item in value.items()
        }
        return transform(transformed)
    if isinstance(value, list):
        return [_transform_query_tree(item, transform) for item in value]
    return copy.deepcopy(value)


def _build_pbir_visual_query(
    *,
    report: Mapping[str, Any],
    page: Mapping[str, Any],
    visual: Mapping[str, Any],
) -> dict[str, Any]:
    """Compile a Fabric PBIR visual definition into a read-only query.

    Power BI's newer PBIR response no longer includes the legacy executable
    ``visualContainer.query``.  It retains the same source contract as
    declarative projections and report/page/visual filters, so build the
    equivalent SemanticQueryDataShape command without hard-coding PHO values.
    """

    visual_definition = visual.get("visual")
    if not isinstance(visual_definition, Mapping):
        raise ValueError("PHO IDTO PBIR table visual definition is invalid")
    query_state = (visual_definition.get("query") or {}).get("queryState")
    if not isinstance(query_state, Mapping):
        raise ValueError("PHO IDTO PBIR table query state is missing")
    values = query_state.get("Values")
    projections = values.get("projections") if isinstance(values, Mapping) else None
    if not isinstance(projections, list) or not projections:
        raise ValueError("PHO IDTO PBIR table projections are missing")

    entities: list[str] = []

    def alias_for(entity: object) -> str:
        name = _norm_text(entity)
        if not name:
            raise ValueError("PHO IDTO PBIR query contains a blank entity")
        if name not in entities:
            entities.append(name)
        return f"s{entities.index(name)}"

    def projection_transform(node: dict[str, Any]) -> dict[str, Any]:
        source_ref = node.get("SourceRef")
        if isinstance(source_ref, Mapping) and source_ref.get("Entity"):
            node["SourceRef"] = {"Source": alias_for(source_ref["Entity"])}
        return node

    selections: list[dict[str, Any]] = []
    semantic_names: set[str] = set()
    for projection in projections:
        if not isinstance(projection, Mapping):
            raise ValueError("PHO IDTO PBIR projection is not an object")
        field = projection.get("field")
        if not isinstance(field, Mapping) or len(field) != 1:
            raise ValueError("PHO IDTO PBIR projection field is invalid")
        field_kind = next(iter(field))
        if field_kind not in {"Column", "Measure"}:
            raise ValueError(
                f"PHO IDTO PBIR projection type is unsupported: {field_kind}"
            )
        query_ref = _norm_text(projection.get("queryRef"))
        if not query_ref or query_ref in semantic_names:
            raise ValueError("PHO IDTO PBIR projection names are invalid")
        semantic_names.add(query_ref)
        transformed = _transform_query_tree(field, projection_transform)
        selection = {
            field_kind: transformed[field_kind],
            "Name": query_ref,
            "NativeReferenceName": _norm_text(
                projection.get("nativeQueryRef")
                or projection.get("displayName")
                or query_ref
            ),
        }
        selections.append(selection)

    where: list[dict[str, Any]] = []
    filter_configs = (
        report.get("filterConfig"),
        page.get("filterConfig"),
        visual.get("filterConfig"),
    )
    for filter_config in filter_configs:
        filters = (
            filter_config.get("filters")
            if isinstance(filter_config, Mapping)
            else []
        )
        if not isinstance(filters, list):
            raise ValueError("PHO IDTO PBIR filter configuration is invalid")
        for item in filters:
            if not isinstance(item, Mapping):
                raise ValueError("PHO IDTO PBIR filter is not an object")
            query_filter = item.get("filter")
            # Power BI includes empty Advanced filter placeholders for visible
            # measures. They carry no condition and must not become queries.
            if query_filter is None:
                continue
            if not isinstance(query_filter, Mapping):
                raise ValueError("PHO IDTO PBIR filter query is invalid")
            filter_from = query_filter.get("From") or []
            filter_where = query_filter.get("Where") or []
            if not isinstance(filter_from, list) or not isinstance(
                filter_where, list
            ):
                raise ValueError("PHO IDTO PBIR filter query shape is invalid")
            local_aliases: dict[str, str] = {}
            for source in filter_from:
                if not isinstance(source, Mapping):
                    raise ValueError("PHO IDTO PBIR filter source is invalid")
                local_name = _norm_text(source.get("Name"))
                entity = _norm_text(source.get("Entity"))
                if not local_name or not entity or local_name in local_aliases:
                    raise ValueError("PHO IDTO PBIR filter source is ambiguous")
                local_aliases[local_name] = alias_for(entity)

            def filter_transform(node: dict[str, Any]) -> dict[str, Any]:
                source_ref = node.get("SourceRef")
                if not isinstance(source_ref, Mapping):
                    return node
                if source_ref.get("Entity"):
                    node["SourceRef"] = {
                        "Source": alias_for(source_ref["Entity"])
                    }
                    return node
                local_name = _norm_text(source_ref.get("Source"))
                if local_name:
                    if local_name not in local_aliases:
                        raise ValueError(
                            "PHO IDTO PBIR filter references an unknown source"
                        )
                    node["SourceRef"] = {"Source": local_aliases[local_name]}
                return node

            transformed_where = _transform_query_tree(
                filter_where, filter_transform
            )
            if not isinstance(transformed_where, list):
                raise ValueError("PHO IDTO PBIR transformed filter is invalid")
            where.extend(transformed_where)

    semantic_query: dict[str, Any] = {
        "Version": 2,
        "From": [
            {"Name": alias_for(entity), "Entity": entity, "Type": 0}
            for entity in entities
        ],
        "Select": selections,
    }
    if where:
        semantic_query["Where"] = where
    return {
        "Commands": [
            {
                "SemanticQueryDataShapeCommand": {
                    "Query": semantic_query,
                    "Binding": {
                        "Primary": {
                            "Groupings": [
                                {"Projections": list(range(len(selections)))}
                            ]
                        },
                        "DataReduction": {
                            "DataVolume": 3,
                            "Primary": {"Window": {"Count": 500}},
                        },
                        "Version": 1,
                    },
                    "ExecutionMetricsKind": 1,
                }
            }
        ]
    }


def _discover_pbir_monthly_visual(
    metadata: Mapping[str, Any],
    *,
    page_display_name: str,
    preferred_visual_name: str,
) -> tuple[dict[str, Any], str, str, str] | None:
    exploration = metadata.get("exploration")
    if not isinstance(exploration, Mapping):
        return None
    exploration_content = exploration.get("explorationContent")
    if not isinstance(exploration_content, Mapping):
        return None
    document_value = exploration_content.get("explorationDocument")
    if document_value is None:
        return None
    try:
        document = _json_object(document_value, label="PBIR exploration document")
    except (ValueError, json.JSONDecodeError) as exc:
        raise RuntimeError("PHO IDTO PBIR exploration document is invalid") from exc

    pages_container = document.get("pages")
    pages = pages_container.get("pages") if isinstance(pages_container, Mapping) else None
    if not isinstance(pages, list):
        raise RuntimeError("PHO IDTO PBIR page definitions are missing")
    page_matches = [
        item
        for item in pages
        if isinstance(item, Mapping)
        and isinstance(item.get("content"), Mapping)
        and _norm_text(item["content"].get("displayName")).casefold()
        == page_display_name.casefold()
    ]
    if len(page_matches) != 1:
        raise RuntimeError(
            f"PHO IDTO monthly page discovery returned {len(page_matches)} matches"
        )
    page_wrapper = page_matches[0]
    page = page_wrapper["content"]

    candidates: list[Mapping[str, Any]] = []
    for wrapper in page_wrapper.get("visualContainers") or []:
        if not isinstance(wrapper, Mapping):
            continue
        content = wrapper.get("content")
        definition = content.get("visual") if isinstance(content, Mapping) else None
        if (
            isinstance(definition, Mapping)
            and definition.get("visualType") == "tableEx"
            and isinstance(definition.get("query"), Mapping)
        ):
            candidates.append(content)
    preferred = [
        item
        for item in candidates
        if _norm_text(item.get("name")) == preferred_visual_name
    ]
    selected_pool = preferred or candidates
    if len(selected_pool) != 1:
        raise RuntimeError(
            f"PHO IDTO monthly table discovery returned {len(selected_pool)} matches"
        )
    visual = selected_pool[0]

    # The preliminary reporting year is a source contract, not the crawler's
    # wall-clock year. Recover it from the official current-date page filter and
    # fail closed if PHO changes that definition.
    reporting_years: set[int] = set()
    page_filter_config = page.get("filterConfig")
    page_filters = (
        page_filter_config.get("filters")
        if isinstance(page_filter_config, Mapping)
        else []
    )
    if not isinstance(page_filters, list):
        raise RuntimeError("PHO IDTO PBIR page filter configuration is invalid")
    for item in page_filters:
        if not isinstance(item, Mapping):
            continue
        field = item.get("field")
        column = field.get("Column") if isinstance(field, Mapping) else None
        if not isinstance(column, Mapping) or _norm_header(
            column.get("Property")
        ) != "date":
            continue
        filter_text = json.dumps(item.get("filter") or {}, ensure_ascii=False)
        reporting_years.update(int(value) for value in YEAR_RE.findall(filter_text))
    if len(reporting_years) != 1:
        raise RuntimeError(
            "PHO IDTO PBIR reporting-year filter is missing or ambiguous"
        )
    reporting_year = next(iter(reporting_years))

    report_wrapper = document.get("report")
    report = (
        report_wrapper.get("content")
        if isinstance(report_wrapper, Mapping)
        and isinstance(report_wrapper.get("content"), Mapping)
        else {}
    )
    try:
        query = _build_pbir_visual_query(
            report=report,
            page=page,
            visual=visual,
        )
    except ValueError as exc:
        raise RuntimeError(str(exc)) from exc
    return (
        query,
        _norm_text(page.get("displayName")),
        _norm_text(visual.get("name")),
        f"{_norm_text(page.get('displayName'))} {reporting_year}",
    )


def discover_monthly_visual(
    metadata: Mapping[str, Any],
    *,
    page_display_name: str = DEFAULT_PAGE_DISPLAY_NAME,
    preferred_visual_name: str = DEFAULT_VISUAL_NAME,
) -> MonthlyVisual:
    """Discover the monthly table by semantic page/type, not array position."""

    models = metadata.get("models")
    if not isinstance(models, list) or len(models) != 1:
        raise RuntimeError("PHO IDTO expected exactly one Power BI model")
    model = models[0]
    if not isinstance(model, dict):
        raise RuntimeError("PHO IDTO Power BI model metadata is invalid")

    page_name = ""
    visual_name = ""
    title = ""
    query: dict[str, Any] | None = None
    sections = (metadata.get("exploration") or {}).get("sections")
    if isinstance(sections, list):
        page_matches = [
            page
            for page in sections
            if _norm_text(page.get("displayName")).casefold()
            == page_display_name.casefold()
        ]
        if len(page_matches) == 1:
            page = page_matches[0]
            candidates: list[tuple[dict[str, Any], dict[str, Any], str]] = []
            for visual in page.get("visualContainers") or []:
                if not isinstance(visual, dict):
                    continue
                try:
                    config = _json_object(
                        visual.get("config") or {}, label="visual config"
                    )
                except (ValueError, json.JSONDecodeError):
                    continue
                single = config.get("singleVisual") or {}
                if single.get("visualType") != "tableEx" or not visual.get("query"):
                    continue
                candidates.append((visual, config, _visual_title(config)))

            preferred = [
                item
                for item in candidates
                if _norm_text(item[1].get("name")) == preferred_visual_name
            ]
            semantic = [
                item
                for item in candidates
                if "disease" in item[2].casefold()
                and "month" in item[2].casefold()
            ]
            selected_pool = preferred or semantic or candidates
            if len(selected_pool) == 1:
                visual, config, title = selected_pool[0]
                query = _json_object(
                    visual["query"], label="monthly visual query"
                )
                page_name = _norm_text(page.get("displayName"))
                visual_name = _norm_text(config.get("name"))

    if query is None:
        pbir = _discover_pbir_monthly_visual(
            metadata,
            page_display_name=page_display_name,
            preferred_visual_name=preferred_visual_name,
        )
        if pbir is None:
            if not isinstance(sections, list):
                raise RuntimeError("PHO IDTO Power BI page metadata is missing")
            page_count = len(
                [
                    page
                    for page in sections
                    if _norm_text(page.get("displayName")).casefold()
                    == page_display_name.casefold()
                ]
            )
            if page_count != 1:
                raise RuntimeError(
                    f"PHO IDTO monthly page discovery returned {page_count} matches"
                )
            raise RuntimeError("PHO IDTO monthly table discovery returned 0 matches")
        query, page_name, visual_name, title = pbir

    try:
        model_id = int(model["id"])
        dataset_id = _norm_text(model["dbName"])
    except (KeyError, TypeError, ValueError) as exc:
        raise RuntimeError("PHO IDTO model id/dataset id is missing") from exc
    if not dataset_id:
        raise RuntimeError("PHO IDTO dataset id is blank")

    return MonthlyVisual(
        model_id=model_id,
        dataset_id=dataset_id,
        query=query,
        page_name=page_name,
        visual_name=visual_name,
        title=title,
        model_refresh_time=_norm_text(
            model.get("LastRefreshTime") or model.get("lastRefreshTime")
        ),
    )


def build_query_payload(visual: MonthlyVisual, *, report_id: str) -> dict[str, Any]:
    return {
        "version": "1.0.0",
        "queries": [
            {
                "Query": visual.query,
                "ApplicationContext": {
                    "DatasetId": visual.dataset_id,
                    "Sources": [{"ReportId": report_id}],
                },
            }
        ],
        "cancelQueries": [],
        "modelId": visual.model_id,
    }


def decode_powerbi_dm0(response: Mapping[str, Any]) -> tuple[str, list[dict[str, Any]]]:
    """Decode the C/R/Ø compression used by the current IDTO table response."""

    try:
        results = response["results"]
        if not isinstance(results, list) or len(results) != 1:
            raise ValueError("PHO IDTO query must return exactly one result")
        data = results[0]["result"]["data"]
        datasets = data["dsr"]["DS"]
        if not isinstance(datasets, list) or len(datasets) != 1:
            raise ValueError("PHO IDTO query must return exactly one dataset")
        dataset = datasets[0]
        phase_groups = dataset["PH"]
        if not isinstance(phase_groups, list) or len(phase_groups) != 1:
            raise ValueError("PHO IDTO query must return exactly one phase group")
        phase_group = phase_groups[0]
        dm0 = phase_group["DM0"]
    except (KeyError, IndexError, TypeError) as exc:
        raise ValueError("PHO IDTO query response has no expected DSR/DM0 table") from exc
    if not isinstance(data, Mapping) or not isinstance(dataset, Mapping):
        raise ValueError("PHO IDTO query response has invalid DSR containers")
    if dataset.get("IC") is False:
        raise ValueError("PHO IDTO query returned an incomplete reduced dataset")
    unsupported_continuation_keys = {
        "continuation",
        "continuationtoken",
        "continuationtokens",
        "datareduction",
        "reductioncontinuation",
    }
    for container in (data, data.get("dsr", {}), dataset, phase_group):
        if not isinstance(container, Mapping):
            raise ValueError("PHO IDTO query response has invalid DSR containers")
        normalized_keys = {_norm_header(key).replace(" ", "") for key in container}
        if normalized_keys.intersection(unsupported_continuation_keys):
            raise ValueError("PHO IDTO query returned a paged or data-reduced table")
    alternate_matrices = {
        key
        for key in phase_group
        if re.fullmatch(r"DM\d+", str(key)) and key != "DM0"
    }
    if alternate_matrices:
        raise ValueError("PHO IDTO query returned multiple data matrices")
    if not isinstance(dm0, list) or not dm0:
        raise ValueError("PHO IDTO query returned an empty monthly table")
    if dataset.get("ValueDicts"):
        raise ValueError(
            "PHO IDTO introduced unsupported Power BI value-dictionary encoding"
        )

    schema_rows = [
        index
        for index, row in enumerate(dm0)
        if isinstance(row, dict) and "S" in row
    ]
    if schema_rows != [0]:
        raise ValueError("PHO IDTO DM0 schema must appear exactly once on row 0")
    schema_row = dm0[0]
    schema = schema_row.get("S")
    if not isinstance(schema, list) or not schema:
        raise ValueError("PHO IDTO DM0 schema is empty")
    physical = [_norm_text(item.get("N")) for item in schema]
    if any(not name for name in physical) or len(set(physical)) != len(physical):
        raise ValueError("PHO IDTO DM0 physical column names are invalid")

    previous: list[Any] = [None] * len(physical)
    decoded: list[dict[str, Any]] = []
    for row_number, encoded in enumerate(dm0):
        if not isinstance(encoded, dict):
            raise ValueError(f"PHO IDTO DM0 row {row_number} is not an object")
        try:
            repeat_mask = int(encoded.get("R", 0))
            null_mask = int(encoded.get("Ø", 0))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"PHO IDTO DM0 row {row_number} has an invalid mask") from exc
        if repeat_mask < 0 or null_mask < 0:
            raise ValueError(f"PHO IDTO DM0 row {row_number} has a negative mask")
        if (repeat_mask | null_mask) >> len(physical):
            raise ValueError(
                f"PHO IDTO DM0 row {row_number} mask exceeds the schema width"
            )
        if repeat_mask & null_mask:
            raise ValueError(f"PHO IDTO DM0 row {row_number} has overlapping masks")

        compact = encoded.get("C", [])
        if not isinstance(compact, list):
            raise ValueError(f"PHO IDTO DM0 row {row_number} has invalid compact data")
        position = 0
        values: list[Any] = []
        for column_number in range(len(physical)):
            bit = 1 << column_number
            if repeat_mask & bit:
                if row_number == 0:
                    raise ValueError("PHO IDTO first DM0 row cannot repeat values")
                value = previous[column_number]
            elif null_mask & bit:
                value = None
            else:
                if position >= len(compact):
                    raise ValueError(f"PHO IDTO DM0 row {row_number} compact data underflow")
                value = compact[position]
                position += 1
            values.append(value)
        if position != len(compact):
            raise ValueError(f"PHO IDTO DM0 row {row_number} compact data overflow")
        previous = values
        decoded.append(dict(zip(physical, values)))

    try:
        selections = data["descriptor"]["Select"]
    except (KeyError, TypeError) as exc:
        raise ValueError("PHO IDTO semantic column descriptor is missing") from exc
    if not isinstance(selections, list):
        raise ValueError("PHO IDTO semantic column descriptor is invalid")
    name_map: dict[str, str] = {}
    semantic_names: set[str] = set()
    for selection in selections:
        if not isinstance(selection, dict):
            continue
        value_name = _norm_text(selection.get("Value"))
        semantic_name = _norm_text(selection.get("Name"))
        if value_name and semantic_name:
            if value_name in name_map or semantic_name in semantic_names:
                raise ValueError("PHO IDTO semantic column descriptor is ambiguous")
            name_map[value_name] = semantic_name
            semantic_names.add(semantic_name)
        dynamic_name = _norm_text(
            ((selection.get("DynamicFormat") or {}).get("Format"))
        )
        if dynamic_name and semantic_name:
            dynamic_semantic = semantic_name + ".__dynamic_format__"
            if dynamic_name in name_map or dynamic_semantic in semantic_names:
                raise ValueError("PHO IDTO dynamic-format descriptor is ambiguous")
            name_map[dynamic_name] = dynamic_semantic
            semantic_names.add(dynamic_semantic)

    semantic_rows = [
        {name_map.get(column, column): value for column, value in row.items()}
        for row in decoded
    ]
    return _norm_text(data.get("timestamp")), semantic_rows


def _parse_case_value(value: object) -> tuple[str | None, bool]:
    """Return a canonical count string and whether it is suppressed."""

    if value is None:
        return None, False
    text = _norm_text(value)
    if not text or text.casefold() in {"n/a", "na", "null", "none", "—", "–"}:
        return None, False
    match = SUPPRESSION_RE.fullmatch(text)
    if match:
        return f"<{int(match.group(1))}", True
    if text.casefold() in {"suppressed", "suppression", "*"}:
        return "suppressed", True
    try:
        number = Decimal(text.replace(",", ""))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid PHO IDTO case value: {value!r}") from exc
    if not number.is_finite() or number < 0 or number != number.to_integral_value():
        raise ValueError(f"Invalid PHO IDTO case count: {value!r}")
    return str(int(number)), False


def _decimal_text(value: object) -> str:
    if value is None or not _norm_text(value):
        return ""
    try:
        number = Decimal(_norm_text(value).replace(",", ""))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid PHO IDTO decimal value: {value!r}") from exc
    if not number.is_finite() or number < 0:
        raise ValueError(f"Invalid PHO IDTO decimal value: {value!r}")
    return format(number, "f")


def _reporting_year(*values: object, fallback: int | None = None) -> int:
    for value in values:
        match = YEAR_RE.search(_norm_text(value))
        if match:
            return int(match.group(1))
    if fallback is not None and 2000 <= int(fallback) <= 2100:
        return int(fallback)
    raise ValueError("PHO IDTO reporting year could not be determined")


def _base_output_row(
    *,
    disease: str,
    reporting_year: int,
    month: int,
    cases: str,
    ytd_cases: str,
    ytd_rate: str,
    retrieved_at: str,
    dataset_timestamp: str,
    model_refresh_time: str,
    acquisition_mode: str,
    source_url: str,
    authoritative_revision: bool,
    allow_equal_quality_overwrite: bool,
) -> dict[str, str]:
    try:
        definition_version = DEFINITION_VERSION_BY_YEAR[reporting_year]
    except KeyError as exc:
        raise ValueError(
            f"PHO IDTO reporting year {reporting_year} has not passed "
            "definition-version review"
        ) from exc
    return {
        "Date": date(reporting_year, month, 1).isoformat(),
        "RawDiseaseLabel": disease,
        "DiseaseCode": "",
        "Year": str(reporting_year),
        "Month": str(month),
        "Cases": cases,
        "YearToDateCases": ytd_cases,
        "YearToDateRatePer100000": ytd_rate,
        "ReportingArea": "Ontario",
        "Geocode": ONTARIO_GEOCODE,
        "JurisdictionCode": ONTARIO_GEOCODE,
        "ParentCountryCode": "CA",
        "LocationType": "subdivision",
        "GeographyKey": ONTARIO_GEOGRAPHY_KEY,
        "PopulationScope": "Ontario residents at diagnosis",
        "DatasetStatus": "preliminary",
        "IsProvisional": "true",
        "AuthoritativeRevision": str(authoritative_revision).lower(),
        "AllowEqualQualityOverwrite": str(
            allow_equal_quality_overwrite
        ).lower(),
        "TimeBasis": SPECIAL_TIME_BASES.get(
            disease.casefold(), "PHO episode-date hierarchy"
        ),
        "DefinitionVersion": definition_version,
        "DatasetTimestamp": dataset_timestamp,
        "ModelRefreshTime": model_refresh_time,
        "RetrievedAt": retrieved_at,
        "AcquisitionMode": acquisition_mode,
        "Source": DEFAULT_SOURCE_NAME,
        "SourceURL": source_url,
    }


def normalize_powerbi_monthly_rows(
    response: Mapping[str, Any],
    *,
    visual: MonthlyVisual,
    retrieved_at: str,
    source_url: str = DEFAULT_LANDING_URL,
) -> tuple[str, list[dict[str, str]]]:
    dataset_timestamp, source_rows = decode_powerbi_dm0(response)
    reporting_year = _reporting_year(visual.title)
    disease_field = "Lookup Disease.Disease"
    ytd_field = "Monthly Data Table Measures.Cases YTD"
    rate_field = "Monthly ON + PHU Case.YTDRate"

    required = {
        disease_field,
        ytd_field,
        rate_field,
        *{
            f"Monthly Data Table Measures.{number:02d} {month_name}"
            for number, (month_name, _) in enumerate(MONTHS, 1)
        },
    }
    actual = {key for row in source_rows for key in row}
    missing = sorted(required - actual)
    if missing:
        raise ValueError("PHO IDTO monthly table schema changed: " + ", ".join(missing))

    output: list[dict[str, str]] = []
    seen_diseases: set[str] = set()
    for raw in source_rows:
        disease = _norm_text(raw.get(disease_field))
        if not disease or disease.casefold() in {"total", "all diseases"}:
            continue
        if disease in seen_diseases:
            raise ValueError(f"PHO IDTO returned duplicate disease row: {disease}")
        seen_diseases.add(disease)

        ytd_cases, ytd_suppressed = _parse_case_value(raw.get(ytd_field))
        ytd_rate = _decimal_text(raw.get(rate_field))
        month_sum = 0
        can_reconcile = not ytd_suppressed
        for index, (month_name, month_number) in enumerate(MONTHS, 1):
            field_name = f"Monthly Data Table Measures.{index:02d} {month_name}"
            cases, suppressed = _parse_case_value(raw.get(field_name))
            if cases is None:
                continue
            if suppressed:
                can_reconcile = False
            else:
                month_sum += int(cases)
            output.append(
                _base_output_row(
                    disease=disease,
                    reporting_year=reporting_year,
                    month=month_number,
                    cases=cases,
                    ytd_cases=ytd_cases or "",
                    ytd_rate=ytd_rate,
                    retrieved_at=retrieved_at,
                    dataset_timestamp=dataset_timestamp,
                    model_refresh_time=visual.model_refresh_time,
                    acquisition_mode="powerbi_read_only",
                    source_url=source_url,
                    # The persistence layer authorizes replacement only after
                    # comparing this refresh with the newest stored release.
                    authoritative_revision=False,
                    allow_equal_quality_overwrite=False,
                )
            )
        if can_reconcile and ytd_cases is not None and month_sum != int(ytd_cases):
            raise ValueError(
                f"PHO IDTO YTD mismatch for {disease}: months={month_sum}, ytd={ytd_cases}"
            )

    if not output:
        raise ValueError("PHO IDTO monthly table contained no usable observations")
    output.sort(key=lambda row: (row["Date"], row["RawDiseaseLabel"]))
    return dataset_timestamp, output


def _month_column(value: object) -> tuple[int, int | None] | None:
    if isinstance(value, datetime):
        return value.month, value.year
    if isinstance(value, date):
        return value.month, value.year

    raw = _norm_text(value).casefold()
    if not raw:
        return None

    numeric = re.fullmatch(
        r"(?:(20\d{2})[-/.](0[1-9]|1[0-2])|"
        r"(0[1-9]|1[0-2])[-/.](20\d{2})|"
        r"(0?[1-9]|1[0-2]))",
        raw,
    )
    if numeric:
        year_text = numeric.group(1) or numeric.group(4)
        month_text = numeric.group(2) or numeric.group(3) or numeric.group(5)
        assert month_text is not None
        return int(month_text), int(year_text) if year_text else None

    header = _norm_header(raw)
    tokens = header.split()
    year_tokens = [token for token in tokens if YEAR_RE.fullmatch(token)]
    if len(year_tokens) > 1:
        return None
    year = int(year_tokens[0]) if year_tokens else None
    remaining = [token for token in tokens if token not in year_tokens]
    month_words = [token for token in remaining if token in MONTH_NAMES]
    if len(month_words) != 1:
        return None
    month = MONTH_NAMES[month_words[0]]
    allowed_tokens = {month_words[0], f"{month:02d}"}
    if len(remaining) > 2 or any(token not in allowed_tokens for token in remaining):
        return None
    return month, year


def _find_header_row(matrix: Sequence[Sequence[object]]) -> int:
    for index, row in enumerate(matrix[:40]):
        headers = [_norm_header(value) for value in row]
        has_disease = any(
            value
            in {
                "disease",
                "disease name",
                "disease of public health significance",
                "dophs",
            }
            for value in headers
        )
        month_count = sum(_month_column(value) is not None for value in row)
        has_long = any(value in {"month", "reporting month"} for value in headers) and any(
            value in {"cases", "case count", "count", "monthly count"}
            for value in headers
        )
        if has_disease and (month_count >= 1 or has_long):
            return index
    raise ValueError("Official PHO export contains no recognizable monthly table header")


def _matrix_to_dicts(
    matrix: Sequence[Sequence[object]],
) -> tuple[list[str], list[dict[str, object]]]:
    header_index = _find_header_row(matrix)
    raw_headers = list(matrix[header_index])
    headers: list[str] = []
    for index, value in enumerate(raw_headers):
        text = _norm_text(value) or f"column_{index + 1}"
        if text in headers:
            text = f"{text}_{index + 1}"
        headers.append(text)
    rows = []
    for values in matrix[header_index + 1 :]:
        if len(values) > len(headers) and any(
            _norm_text(value) for value in values[len(headers) :]
        ):
            raise ValueError(
                "Official PHO export row contains values beyond the header width"
            )
        padded = list(values) + [None] * max(0, len(headers) - len(values))
        if not any(_norm_text(value) for value in padded[: len(headers)]):
            continue
        rows.append(dict(zip(headers, padded[: len(headers)])))
    title_cells = [
        _norm_text(value)
        for row in matrix[:header_index]
        for value in row
        if _norm_text(value)
    ]
    return title_cells, rows


def _bounded_matrix(rows: Iterable[Sequence[object]], *, label: str) -> list[list[object]]:
    matrix: list[list[object]] = []
    cell_count = 0
    for row_number, row in enumerate(rows, start=1):
        if row_number > MAX_EXPORT_ROWS:
            raise ValueError(f"{label} exceeds the {MAX_EXPORT_ROWS}-row limit")
        values = list(row)
        if len(values) > MAX_EXPORT_COLUMNS:
            raise ValueError(
                f"{label} exceeds the {MAX_EXPORT_COLUMNS}-column limit"
            )
        cell_count += len(values)
        if cell_count > MAX_EXPORT_CELLS:
            raise ValueError(f"{label} exceeds the export cell limit")
        matrix.append(values)
    return matrix


def _load_export_tables(
    path: Path,
    *,
    content: bytes | None = None,
) -> list[tuple[str, list[str], list[dict[str, object]]]]:
    source_bytes = path.read_bytes() if content is None else content
    if len(source_bytes) > MAX_EXPORT_BYTES:
        raise ValueError(
            f"PHO export exceeds the {MAX_EXPORT_BYTES}-byte compressed/file limit"
        )
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        text = ""
        for encoding in ("utf-8-sig", "utf-8", "cp1252"):
            try:
                text = source_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if not text:
            raise ValueError("PHO CSV export encoding is not supported")
        matrix = _bounded_matrix(
            csv.reader(io.StringIO(text)), label=f"PHO CSV {path.name}"
        )
        title_cells, rows = _matrix_to_dicts(matrix)
        return [(path.name, title_cells, rows)]
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("PHO export must be a CSV or XLSX file")

    try:
        with zipfile.ZipFile(io.BytesIO(source_bytes)) as archive:
            expanded = sum(info.file_size for info in archive.infolist())
    except zipfile.BadZipFile as exc:
        raise ValueError("PHO XLSX export is not a valid Office workbook") from exc
    if expanded > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise ValueError("PHO XLSX export exceeds the expanded-size limit")

    workbook = load_workbook(
        io.BytesIO(source_bytes), read_only=True, data_only=True
    )
    tables: list[tuple[str, list[str], list[dict[str, object]]]] = []
    errors: list[str] = []
    try:
        for sheet in workbook.worksheets:
            try:
                matrix = _bounded_matrix(
                    sheet.iter_rows(values_only=True),
                    label=f"PHO workbook sheet {sheet.title!r}",
                )
                title_cells, rows = _matrix_to_dicts(matrix)
                tables.append((sheet.title, title_cells, rows))
            except ValueError as exc:
                errors.append(f"{sheet.title}: {exc}")
    finally:
        workbook.close()
    if not tables:
        raise ValueError("No monthly table found in PHO workbook: " + "; ".join(errors))
    return tables


def _find_column(
    headers: Iterable[str],
    aliases: Sequence[str],
    *,
    label: str,
) -> str | None:
    matches = [
        header for header in headers if _norm_header(header) in set(aliases)
    ]
    if len(matches) > 1:
        raise ValueError(f"Official PHO export has ambiguous {label} columns")
    return matches[0] if matches else None


def _validate_export_scope(
    headers: Sequence[str], rows: Sequence[Mapping[str, object]]
) -> None:
    aggregate_geographies = {
        "ontario",
        "on",
        "ca on",
        "province of ontario",
        "all ontario",
        "all",
        "total",
    }
    aggregate_dimensions = {
        "all",
        "total",
        "all ages",
        "all sexes",
        "both sexes",
    }
    for header in headers:
        normalized = _norm_header(header)
        values = {
            _norm_header(row.get(header))
            for row in rows
            if _norm_text(row.get(header))
        }
        if not values:
            continue
        is_geography = normalized in {
            "geography",
            "region",
            "area",
            "reporting area",
            "province",
            "phu",
        } or "health unit" in normalized
        if is_geography and not values.issubset(aggregate_geographies):
            raise ValueError(
                "Official PHO export contains sub-provincial geography rows"
            )
        is_dimension = normalized in {
            "age",
            "age group",
            "sex",
            "gender",
            "case status",
            "case classification",
            "classification",
        } or "age group" in normalized or normalized.startswith("sex ")
        if is_dimension and not values.issubset(aggregate_dimensions):
            raise ValueError(
                f"Official PHO export contains non-aggregate {header!r} rows"
            )
        known_exact = {
            "disease",
            "disease name",
            "disease code",
            "code",
            "disease of public health significance",
            "dophs",
            "year",
            "reporting year",
            "month",
            "reporting month",
            "cases",
            "case count",
            "monthly cases",
            "monthly count",
            "count",
            "ytd total",
            "cases ytd",
            "year to date total",
            "year to date cases",
            "ytd rate per 100000 population",
            "ytd rate",
            "year to date rate",
        }
        if (
            normalized not in known_exact
            and _month_column(header) is None
            and not is_geography
            and not is_dimension
        ):
            raise ValueError(
                f"Official PHO export contains an unrecognized populated column: {header!r}"
            )


def _unique_export_year(
    *,
    title_cells: Sequence[str],
    rows: Sequence[Mapping[str, object]],
    year_column: str | None,
    reporting_year: int | None,
) -> int:
    years: set[int] = set()
    for value in title_cells:
        years.update(int(match) for match in YEAR_RE.findall(_norm_text(value)))
    if year_column:
        for row in rows:
            text = _norm_text(row.get(year_column))
            if text:
                years.add(_reporting_year(text))
    if reporting_year is not None:
        years.add(_reporting_year("", fallback=reporting_year))
    if len(years) != 1:
        raise ValueError(
            "Official PHO export must identify exactly one reporting year"
        )
    return next(iter(years))


def normalize_export_table(
    title_cells: Sequence[str],
    rows: Sequence[Mapping[str, object]],
    *,
    reporting_year: int | None,
    retrieved_at: str,
    source_url: str = DEFAULT_LANDING_URL,
    allow_authoritative_revision: bool = False,
) -> list[dict[str, str]]:
    """Normalize either wide official exports or a long replay CSV."""

    if not rows:
        raise ValueError("Official PHO export table is empty")
    headers = list(rows[0])
    disease_column = _find_column(
        headers,
        (
            "disease",
            "disease name",
            "disease of public health significance",
            "dophs",
        ),
        label="disease",
    )
    if disease_column is None:
        raise ValueError("Official PHO export has no disease column")
    year_column = _find_column(
        headers, ("year", "reporting year"), label="year"
    )
    month_column = _find_column(
        headers, ("month", "reporting month"), label="month"
    )
    cases_column = _find_column(
        headers,
        ("cases", "case count", "monthly cases", "monthly count", "count"),
        label="case count",
    )
    ytd_column = _find_column(
        headers,
        ("ytd total", "cases ytd", "year to date total", "year to date cases"),
        label="year-to-date count",
    )
    rate_column = _find_column(
        headers,
        ("ytd rate per 100000 population", "ytd rate", "year to date rate"),
        label="year-to-date rate",
    )
    _validate_export_scope(headers, rows)
    title_year = _unique_export_year(
        title_cells=title_cells,
        rows=rows,
        year_column=year_column,
        reporting_year=reporting_year,
    )
    non_month_headers = {
        item
        for item in (
            disease_column,
            year_column,
            month_column,
            cases_column,
            ytd_column,
            rate_column,
        )
        if item is not None
    }
    month_columns = {
        header: parsed
        for header in headers
        if header not in non_month_headers
        if (parsed := _month_column(header)) is not None
    }
    if month_columns and (month_column or cases_column):
        raise ValueError(
            "Official PHO export mixes wide and long monthly schemas"
        )
    if bool(month_column) != bool(cases_column):
        raise ValueError(
            "Official PHO long export requires both month and case-count columns"
        )
    resolved_month_headers: dict[tuple[int, int], str] = {}
    for header, (month_number, header_year) in month_columns.items():
        identity = (header_year or title_year, month_number)
        previous_header = resolved_month_headers.get(identity)
        if previous_header is not None:
            raise ValueError(
                "Official PHO export has duplicate month columns: "
                f"{previous_header!r} and {header!r}"
            )
        resolved_month_headers[identity] = header
    if not month_columns and not (month_column and cases_column):
        raise ValueError("Official PHO export has neither wide nor long monthly values")

    output: list[dict[str, str]] = []
    identities: dict[tuple[str, int, int], tuple[str, str, str]] = {}
    wide_diseases: set[str] = set()
    for raw in rows:
        disease = _norm_text(raw.get(disease_column))
        if not disease or disease.casefold() in {"total", "all diseases"}:
            continue
        if month_columns:
            disease_key = disease.casefold()
            if disease_key in wide_diseases:
                raise ValueError(
                    f"Official PHO wide export repeats disease row: {disease}"
                )
            wide_diseases.add(disease_key)
        row_year = title_year
        if year_column and _norm_text(raw.get(year_column)):
            row_year = _reporting_year(raw.get(year_column), fallback=title_year)
        if row_year != title_year:
            raise ValueError("Official PHO export unexpectedly spans multiple years")
        ytd_cases, ytd_suppressed = (
            _parse_case_value(raw.get(ytd_column))
            if ytd_column
            else (None, False)
        )
        ytd_rate = _decimal_text(raw.get(rate_column)) if rate_column else ""

        values: list[tuple[int, int, object]] = []
        if month_columns:
            for header, (month_number, header_year) in month_columns.items():
                values.append((header_year or row_year, month_number, raw.get(header)))
        else:
            assert month_column is not None and cases_column is not None
            parsed_month = _month_column(raw.get(month_column))
            if parsed_month is None:
                raise ValueError(
                    f"Official PHO export has invalid month {raw.get(month_column)!r}"
                )
            month_number, month_year = parsed_month
            values.append((month_year or row_year, month_number, raw.get(cases_column)))

        row_month_sum = 0
        can_reconcile = not ytd_suppressed
        for value_year, month_number, raw_cases in values:
            if value_year != title_year:
                raise ValueError(
                    "Official PHO export month headers span multiple years"
                )
            cases, suppressed = _parse_case_value(raw_cases)
            if cases is None:
                continue
            if suppressed:
                can_reconcile = False
            else:
                row_month_sum += int(cases)
            identity = (disease.casefold(), value_year, month_number)
            provenance = (cases, ytd_cases or "", ytd_rate)
            previous = identities.get(identity)
            if previous is not None and previous != provenance:
                raise ValueError(
                    "Official PHO export contains conflicting duplicate observation: "
                    f"{disease} {value_year}-{month_number:02d}"
                )
            if previous is not None:
                continue
            identities[identity] = provenance
            output.append(
                _base_output_row(
                    disease=disease,
                    reporting_year=value_year,
                    month=month_number,
                    cases=cases,
                    ytd_cases=ytd_cases or "",
                    ytd_rate=ytd_rate,
                    retrieved_at=retrieved_at,
                    dataset_timestamp="",
                    model_refresh_time="",
                    acquisition_mode="official_export_file",
                    source_url=source_url,
                    authoritative_revision=allow_authoritative_revision,
                    allow_equal_quality_overwrite=allow_authoritative_revision,
                )
            )
        if (
            month_columns
            and can_reconcile
            and ytd_cases is not None
            and row_month_sum != int(ytd_cases)
        ):
            raise ValueError(
                f"Official PHO export YTD mismatch for {disease}: "
                f"months={row_month_sum}, ytd={ytd_cases}"
            )
    if not month_columns and ytd_column:
        grouped: dict[tuple[str, str], list[dict[str, str]]] = {}
        for row in output:
            grouped.setdefault(
                (row["RawDiseaseLabel"].casefold(), row["Year"]), []
            ).append(row)
        for disease_rows in grouped.values():
            ytd_values = {
                row["YearToDateCases"]
                for row in disease_rows
                if row["YearToDateCases"]
            }
            if len(ytd_values) > 1:
                raise ValueError(
                    "Official PHO long export contains inconsistent YTD values"
                )
            if not ytd_values:
                continue
            parsed_values = [_parse_case_value(row["Cases"]) for row in disease_rows]
            ytd_value, ytd_suppressed = _parse_case_value(next(iter(ytd_values)))
            if ytd_suppressed or any(suppressed for _value, suppressed in parsed_values):
                continue
            assert ytd_value is not None
            month_total = sum(int(value) for value, _suppressed in parsed_values if value)
            if month_total != int(ytd_value):
                raise ValueError(
                    "Official PHO long export YTD does not reconcile with monthly values"
                )
    if not output:
        raise ValueError("Official PHO export contained no usable monthly observations")
    output.sort(key=lambda row: (row["Date"], row["RawDiseaseLabel"]))
    return output


def normalize_export_file(
    path: Path,
    *,
    reporting_year: int | None = None,
    retrieved_at: str | None = None,
    source_url: str = DEFAULT_LANDING_URL,
    allow_authoritative_revision: bool = False,
    content: bytes | None = None,
) -> list[dict[str, str]]:
    actual_retrieved_at = retrieved_at or datetime.now(timezone.utc).isoformat()
    candidates: list[tuple[str, list[dict[str, str]]]] = []
    errors: list[str] = []
    for table_name, title_cells, rows in _load_export_tables(path, content=content):
        try:
            candidates.append(
                (
                    table_name,
                    normalize_export_table(
                        # A dated official filename is valid year provenance.
                        # It is not used for any other semantic inference.
                        [*title_cells, path.name],
                        rows,
                        reporting_year=reporting_year,
                        retrieved_at=actual_retrieved_at,
                        source_url=source_url,
                        allow_authoritative_revision=allow_authoritative_revision,
                    ),
                )
            )
        except ValueError as exc:
            errors.append(f"{table_name}: {exc}")
    if not candidates:
        raise ValueError("Unable to normalize PHO export: " + "; ".join(errors))
    if len(candidates) != 1:
        raise ValueError(
            "PHO export contains multiple eligible monthly tables: "
            + ", ".join(name for name, _rows in candidates)
        )
    return candidates[0][1]


def _content_hash(rows: Sequence[Mapping[str, str]]) -> str:
    # Retrieval time is operational provenance, not source content. Excluding
    # it makes an unchanged release hash identically across idempotent replays.
    operational_fields = {
        "RetrievedAt",
        "DatasetTimestamp",
        "ModelRefreshTime",
        "AcquisitionMode",
        "AuthoritativeRevision",
        "AllowEqualQualityOverwrite",
    }
    stable_rows = [
        {key: value for key, value in row.items() if key not in operational_fields}
        for row in rows
    ]
    canonical = json.dumps(
        stable_rows,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _bytes_hash(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _write_output_csv(path: Path, rows: Sequence[Mapping[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=OUTPUT_FIELDS)
            writer.writeheader()
            writer.writerows(rows)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _assert_archive_safe(value: object, *, forbidden_values: Sequence[str]) -> None:
    if isinstance(value, Mapping):
        for key, item in value.items():
            normalized_key = _norm_header(key)
            if normalized_key in {
                "access token",
                "accesstoken",
                "authorization",
                "embed token",
            }:
                raise ValueError("Refusing to archive a payload with credential fields")
            _assert_archive_safe(item, forbidden_values=forbidden_values)
        return
    if isinstance(value, (list, tuple)):
        for item in value:
            _assert_archive_safe(item, forbidden_values=forbidden_values)
        return
    if isinstance(value, str) and any(
        secret and secret in value for secret in forbidden_values
    ):
        raise ValueError("Refusing to archive a payload containing an embed token")


class CanadaOntarioPHOCrawler(BaseCrawler):
    """Fetch and normalize PHO IDTO current-year Ontario monthly case counts."""

    def __init__(
        self,
        *,
        save_raw: bool = False,
        raw_dir: Path | None = None,
    ) -> None:
        super().__init__(
            user_agent="Mozilla/5.0 (compatible; GlobalID/2.0; CA-ON-PHO-IDTO)",
            timeout=120,
            max_retries=3,
            delay=0.2,
        )
        config = get_country_bootstrap_config("CA-ON")
        crawler = config.get("crawler_config", {}) if isinstance(config, dict) else {}
        self.landing_url = str(crawler.get("landing_url") or DEFAULT_LANDING_URL)
        self.embed_url = str(crawler.get("embed_url") or DEFAULT_EMBED_URL)
        self.report_id = str(crawler.get("report_id") or DEFAULT_REPORT_ID)
        self.page_display_name = str(
            crawler.get("page_display_name") or DEFAULT_PAGE_DISPLAY_NAME
        )
        self.visual_name = str(crawler.get("visual_name") or DEFAULT_VISUAL_NAME)
        self.file_env = str(crawler.get("file_env") or "CA_ON_IDTO_FILE")
        self.save_raw = save_raw
        self.raw_dir = Path(raw_dir) if raw_dir else Path("data/raw/ca/on_idto")

    def _fetch_live(
        self,
    ) -> tuple[MonthlyVisual, str, dict[str, Any], dict[str, Any], str]:
        wrapper = self.get(self.embed_url).text
        context = extract_embed_context(wrapper, expected_report_id=self.report_id)
        metadata_url = (
            f"{context.cluster_url}/explore/reports/{context.report_id}/modelsAndExploration"
            "?preferReadOnlySession=true&skipQueryData=true"
        )
        metadata_response = self.get(
            metadata_url,
            headers=powerbi_headers(context),
            allow_redirects=False,
        )
        if metadata_response.is_redirect:
            raise RuntimeError("Power BI metadata endpoint returned an unexpected redirect")
        metadata = metadata_response.json()
        visual = discover_monthly_visual(
            metadata,
            page_display_name=self.page_display_name,
            preferred_visual_name=self.visual_name,
        )
        query_response = self.post(
            f"{context.cluster_url}/explore/querydata?synchronous=true",
            headers=powerbi_headers(context, post=True),
            json=build_query_payload(visual, report_id=context.report_id),
            allow_redirects=False,
        )
        if query_response.is_redirect:
            raise RuntimeError("Power BI query endpoint returned an unexpected redirect")
        return (
            visual,
            context.cluster_url,
            metadata,
            query_response.json(),
            context.token,
        )

    def _archive_live_response(
        self,
        *,
        metadata: Mapping[str, Any],
        query_response: Mapping[str, Any],
        retrieved_at: datetime,
        cluster_url: str,
        visual: MonthlyVisual,
        embed_token: str,
    ) -> None:
        _assert_archive_safe(metadata, forbidden_values=(embed_token,))
        _assert_archive_safe(query_response, forbidden_values=(embed_token,))
        stamp = retrieved_at.strftime("%Y%m%dT%H%M%SZ")
        target = self.raw_dir / retrieved_at.strftime("%Y") / stamp
        target.mkdir(parents=True, exist_ok=True)
        metadata_bytes = json.dumps(
            metadata, ensure_ascii=False, sort_keys=True
        ).encode("utf-8")
        query_bytes = json.dumps(
            query_response, ensure_ascii=False, sort_keys=True
        ).encode("utf-8")
        (target / "models_and_exploration.json").write_bytes(metadata_bytes)
        (target / "monthly_query_response.json").write_bytes(query_bytes)
        manifest = {
            "retrieved_at": retrieved_at.isoformat(),
            "landing_url": self.landing_url,
            "embed_url": self.embed_url,
            "cluster_host": urlparse(cluster_url).hostname,
            "report_id": self.report_id,
            "page_name": visual.page_name,
            "visual_name": visual.visual_name,
            "model_refresh_time": visual.model_refresh_time,
            "artifacts": {
                "models_and_exploration.json": {
                    "sha256": _bytes_hash(metadata_bytes),
                    "bytes": len(metadata_bytes),
                },
                "monthly_query_response.json": {
                    "sha256": _bytes_hash(query_bytes),
                    "bytes": len(query_bytes),
                },
            },
            "tokens_persisted": False,
        }
        (target / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, sort_keys=True),
            encoding="utf-8",
        )

    def crawl_monthly_ontario(
        self,
        output_csv: Path,
        *,
        months: Iterable[tuple[int, int]] | None = None,
        input_file: Path | None = None,
        reporting_year: int | None = None,
        allow_file_revisions: bool = False,
        use_configured_file: bool = False,
    ) -> CAOntarioFetchSummary:
        retrieved = datetime.now(timezone.utc)
        actual_input = input_file
        if actual_input is None and use_configured_file:
            configured_file = _norm_text(os.getenv(self.file_env))
            if not configured_file:
                raise ValueError(
                    f"Configured-file replay requested but {self.file_env} is unset"
                )
            actual_input = Path(configured_file)

        dataset_timestamp = ""
        model_refresh_time = ""
        source_artifact_sha256 = ""
        source_file_mtime = ""
        source_bytes: bytes | None = None
        live_archive: tuple[
            MonthlyVisual, str, dict[str, Any], dict[str, Any], str
        ] | None = None
        if actual_input is not None:
            if not actual_input.is_file():
                raise FileNotFoundError(f"Configured PHO IDTO export not found: {actual_input}")
            if actual_input.stat().st_size > MAX_EXPORT_BYTES:
                raise ValueError("Configured PHO IDTO export exceeds the file-size limit")
            source_file_mtime = datetime.fromtimestamp(
                actual_input.stat().st_mtime, tz=timezone.utc
            ).isoformat()
            source_bytes = actual_input.read_bytes()
            source_artifact_sha256 = _bytes_hash(source_bytes)
            rows = normalize_export_file(
                actual_input,
                reporting_year=reporting_year,
                retrieved_at=retrieved.isoformat(),
                source_url=self.landing_url,
                allow_authoritative_revision=allow_file_revisions,
                content=source_bytes,
            )
            acquisition_mode = "official_export_file"
        else:
            visual, cluster_url, metadata, response, embed_token = self._fetch_live()
            dataset_timestamp, rows = normalize_powerbi_monthly_rows(
                response,
                visual=visual,
                retrieved_at=retrieved.isoformat(),
                source_url=self.landing_url,
            )
            model_refresh_time = visual.model_refresh_time
            acquisition_mode = "powerbi_read_only"
            live_archive = (
                visual,
                cluster_url,
                metadata,
                response,
                embed_token,
            )

        source_disease_count = len({row["RawDiseaseLabel"] for row in rows})
        source_populated_months = {
            (int(row["Year"]), int(row["Month"])) for row in rows
        }
        unpublished_month_slots = max(
            0, source_disease_count * 12 - len(rows)
        )

        requested = (
            {(int(year), int(month)) for year, month in months}
            if months is not None
            else None
        )
        if requested is not None:
            rows = [
                row
                for row in rows
                if (int(row["Year"]), int(row["Month"])) in requested
            ]
        if not rows:
            raise RuntimeError("PHO IDTO returned no rows for the requested months")

        years = {int(row["Year"]) for row in rows}
        if len(years) != 1:
            raise ValueError("PHO IDTO preliminary snapshot unexpectedly spans multiple years")
        if self.save_raw and actual_input is not None:
            assert source_bytes is not None
            target = (
                self.raw_dir
                / retrieved.strftime("%Y")
                / retrieved.strftime("%Y%m%dT%H%M%SZ")
            )
            target.mkdir(parents=True, exist_ok=True)
            artifact_name = "official_export" + actual_input.suffix.casefold()
            artifact = target / artifact_name
            artifact.write_bytes(source_bytes)
            manifest = {
                "retrieved_at": retrieved.isoformat(),
                "landing_url": self.landing_url,
                "acquisition_mode": "official_export_file",
                "original_filename": actual_input.name,
                "original_mtime_utc": source_file_mtime,
                "artifacts": {
                    artifact_name: {
                        "sha256": source_artifact_sha256,
                        "bytes": len(source_bytes),
                    }
                },
                "tokens_persisted": False,
            }
            (target / "manifest.json").write_text(
                json.dumps(manifest, ensure_ascii=False, sort_keys=True),
                encoding="utf-8",
            )
        elif self.save_raw and live_archive is not None:
            visual, cluster_url, metadata, response, embed_token = live_archive
            self._archive_live_response(
                metadata=metadata,
                query_response=response,
                retrieved_at=retrieved,
                cluster_url=cluster_url,
                visual=visual,
                embed_token=embed_token,
            )

        _write_output_csv(output_csv, rows)
        latest = max(date.fromisoformat(row["Date"]) for row in rows)
        return CAOntarioFetchSummary(
            row_count=len(rows),
            disease_count=len({row["RawDiseaseLabel"] for row in rows}),
            latest_date=latest,
            reporting_year=next(iter(years)),
            source_url=self.landing_url,
            acquisition_mode=acquisition_mode,
            content_sha256=_content_hash(rows),
            dataset_timestamp=dataset_timestamp,
            model_refresh_time=model_refresh_time,
            populated_month_count=len(source_populated_months),
            unpublished_month_slots=unpublished_month_slots,
            source_artifact_sha256=source_artifact_sha256,
            source_file_mtime=source_file_mtime,
        )

    async def crawl(self, **kwargs: Any) -> list[CrawlerResult]:
        output_csv = Path(
            kwargs.get("output_csv") or "data/current/ca/ontario_idto_monthly.csv"
        )
        summary = self.crawl_monthly_ontario(
            output_csv,
            months=kwargs.get("months"),
            input_file=Path(kwargs["input_file"]) if kwargs.get("input_file") else None,
            reporting_year=kwargs.get("reporting_year"),
            use_configured_file=bool(kwargs.get("use_configured_file", False)),
        )
        return [
            CrawlerResult(
                title="Public Health Ontario IDTO monthly preliminary data",
                url=summary.source_url,
                date=(
                    datetime.combine(summary.latest_date, datetime.min.time(), timezone.utc)
                    if summary.latest_date
                    else None
                ),
                metadata={
                    "rows": summary.row_count,
                    "diseases": summary.disease_count,
                    "reporting_year": summary.reporting_year,
                    "acquisition_mode": summary.acquisition_mode,
                    "content_sha256": summary.content_sha256,
                    "dataset_timestamp": summary.dataset_timestamp,
                    "model_refresh_time": summary.model_refresh_time,
                    "geography_key": ONTARIO_GEOGRAPHY_KEY,
                },
            )
        ]

    def parse(self, response: requests.Response) -> list[CrawlerResult]:
        """The IDTO parser operates on Power BI JSON or official export files."""

        return []


__all__ = [
    "CAOntarioFetchSummary",
    "CanadaOntarioPHOCrawler",
    "DEFAULT_SOURCE_NAME",
    "ONTARIO_GEOGRAPHY_KEY",
    "build_query_payload",
    "decode_powerbi_dm0",
    "discover_monthly_visual",
    "extract_embed_context",
    "normalize_export_file",
    "normalize_export_table",
    "normalize_powerbi_monthly_rows",
    "powerbi_headers",
]
