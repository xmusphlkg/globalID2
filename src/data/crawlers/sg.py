"""Singapore CDA national weekly infectious-disease notifications.

The adapter joins the official data.gov.sg 2012--2022 CSV history to CDA's
publications.  CDA annual workbooks are preferred from 2023 onward (the 2023
link uses an ``.xls`` suffix but contains readable OOXML); the official weekly
PDFs remain a failover for that legacy year.

The CDA web-site terms require written permission for reproduction. Public
release is enabled only under explicit operator authorization, while the
source-terms status remains preserved in every provenance record.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence
from urllib.parse import urljoin

import openpyxl
import pdfplumber
from bs4 import BeautifulSoup

from src.core import get_logger

from .base import BaseCrawler, CrawlerResult

logger = get_logger(__name__)

DEFAULT_SCOPE = "cda_weekly_bulletin"
DEFAULT_SOURCE_NAME = "Singapore CDA Weekly Infectious Diseases Bulletin"
HISTORICAL_SOURCE_NAME = "Singapore data.gov.sg Weekly Infectious Diseases Bulletin (2012-2022)"
ONTOLOGY_SOURCE_ID = "SRC_SG_CDA_WIDB"
HISTORICAL_ONTOLOGY_SOURCE_ID = "SRC_SG_DATA_GOV_WIDB"
HISTORY_START_YEAR = 2012
OPEN_DATASET_ID = "d_ca168b2cb763640d72c4600a68f9909e"
OPEN_DATA_POLL_URL = f"https://api-open.data.gov.sg/v1/public/api/datasets/{OPEN_DATASET_ID}/poll-download"
YEAR_PAGE = "https://www.cda.gov.sg/resources/weekly-infectious-diseases-bulletin-{year}/"
CONTRACT_VERSION = "sg-cda-widb-v1-observed-2026-08"


class SGContractError(ValueError):
    """Raised when an official source no longer matches the observed schema."""


@dataclass(frozen=True)
class SGFetchSummary:
    row_count: int
    years_fetched: tuple[int, ...]
    latest_date: Optional[date]
    artifacts_fetched: int


def _text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\ufeff", "").replace("\xa0", " ").split())


def stable_disease_code(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", _text(value))
    code = re.sub(r"[^a-z0-9]+", "_", normalized.encode("ascii", "ignore").decode().casefold()).strip("_")
    if not code:
        raise SGContractError(f"Cannot derive disease code from {value!r}")
    return code


# A source-local identity prevents spelling changes across the historical CSV,
# XLSX and PDF generations from fragmenting one surveillance series.
DISEASE_ALIASES: Dict[str, tuple[str, tuple[str, ...]]] = {
    "cholera": ("Cholera", ()),
    "paratyphoid": ("Paratyphoid", ()),
    "typhoid": ("Typhoid", ()),
    "acute_viral_hepatitis_a": ("Acute Viral Hepatitis A", ("Viral Hepatitis A", "Acute Hepatitis A")),
    "acute_viral_hepatitis_e": ("Acute Viral Hepatitis E", ("Viral Hepatitis E", "Acute Hepatitis E")),
    "poliomyelitis": ("Poliomyelitis", ()),
    "plague": ("Plague", ()),
    "yellow_fever": ("Yellow Fever", ()),
    "dengue": ("Dengue", ("Dengue Fever",)),
    "dengue_haemorrhagic_fever": ("DHF", ("Dengue Haemorrhagic Fever", "Dengue Hemorrhagic Fever")),
    "malaria": ("Malaria", ()),
    "chikungunya": ("Chikungunya", ("Chikungunya Fever",)),
    "diphtheria": ("Diphtheria", ()),
    "measles": ("Measles", ()),
    "mumps": ("Mumps", ()),
    "rubella": ("Rubella", ()),
    "sars": ("SARS", ("Severe Acute Respiratory Syndrome",)),
    "nipah": ("Nipah", ("Nipah virus infection",)),
    "acute_viral_hepatitis_b": ("Acute Viral Hepatitis B", ("Viral Hepatitis B", "Acute Hepatitis B")),
    "legionellosis": ("Legionellosis", ("Legionnaires' Disease",)),
    "campylobacter_enteritis": ("Campylobacter enteritis", ("Campylobacterenterosis", "Campylobacteriosis")),
    "acute_viral_hepatitis_c": ("Acute Viral Hepatitis C", ("Viral Hepatitis C", "Acute Hepatitis C")),
    "leptospirosis": ("Leptospirosis", ()),
    "melioidosis": ("Melioidosis", ()),
    "meningococcal_infection": ("Meningococcal Infection", ("Meningococcal Disease",)),
    "pertussis": ("Pertussis", ("Whooping Cough",)),
    "invasive_pneumococcal_disease": ("Pneumococcal Disease (invasive)", ("Invasive Pneumococcal Disease", "Pneumococcal Disease (Invasive)")),
    "haemophilus_influenzae_type_b": ("Haemophilus influenzae type b", ("Haemophilus Influenzae Type B",)),
    "salmonellosis_non_enteric_fevers": ("Salmonellosis (non-enteric fevers)", ("Salmonellosis(non-enteric fevers)", "Salmonellosis")),
    "avian_influenza": ("Avian Influenza", ()),
    "zika": ("Zika", ("Zika Virus Infection",)),
    "ebola_virus_disease": ("Ebola Virus Disease", ()),
    "japanese_encephalitis": ("Japanese Encephalitis", ()),
    "tetanus": ("Tetanus", ()),
    "botulism": ("Botulism", ()),
    "murine_typhus": ("Murine Typhus", ()),
    "mpox": ("Mpox", ("Monkeypox",)),
    "middle_east_respiratory_syndrome": ("Middle East Respiratory Syndrome", ("MERS-CoV", "MERS")),
    "encephalitis": ("Encephalitis", ()),
}

_ALIAS_TO_ID = {
    _text(label).casefold(): (code, canonical)
    for code, (canonical, aliases) in DISEASE_ALIASES.items()
    for label in (canonical, *aliases)
}


def canonical_disease(value: object) -> tuple[str, str]:
    label = re.sub(r"\s*[*#^]+\s*$", "", _text(value))
    hit = _ALIAS_TO_ID.get(label.casefold())
    if hit is None:
        raise SGContractError(f"Unregistered CDA disease category: {value!r}")
    return hit


def singapore_week_start(year: int, week: int) -> date:
    """Return the Sunday starting a Singapore epidemiological week."""
    if not 1 <= int(week) <= 53:
        raise SGContractError(f"Invalid Singapore epidemiological week: {year}-W{week}")
    jan4 = date(int(year), 1, 4)
    first_sunday = jan4 - timedelta(days=(jan4.weekday() + 1) % 7)
    start = first_sunday + timedelta(days=7 * (int(week) - 1))
    next_jan4 = date(int(year) + 1, 1, 4)
    next_first = next_jan4 - timedelta(days=(next_jan4.weekday() + 1) % 7)
    if start >= next_first:
        raise SGContractError(f"Week {week} does not exist in Singapore epidemiological year {year}")
    return start


def _parse_week_label(value: object) -> tuple[int, int]:
    match = re.fullmatch(r"\s*(\d{4})-W(\d{1,2})\s*", str(value or ""), re.I)
    if not match:
        raise SGContractError(f"Invalid epi_week value: {value!r}")
    year, week = int(match.group(1)), int(match.group(2))
    singapore_week_start(year, week)
    return year, week


def _parse_date_range(value: object, *, year: int, week: int) -> date:
    match = re.fullmatch(r"\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*-\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*", str(value or ""))
    if not match:
        raise SGContractError(f"Invalid CDA week date range: {value!r}")
    start = date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    end = date(int(match.group(6)), int(match.group(5)), int(match.group(4)))
    expected = singapore_week_start(year, week)
    if start != expected or end != start + timedelta(days=6):
        raise SGContractError(f"CDA week/date mismatch for {year}-W{week:02d}: {value!r}")
    return start


def _parse_count(value: object) -> Optional[int]:
    if value is None or _text(value) in {"", "-", "NA", "N/A"}:
        return None
    if isinstance(value, bool):
        raise SGContractError("Boolean case count is invalid")
    try:
        numeric = float(str(value).replace(",", ""))
    except (TypeError, ValueError) as exc:
        raise SGContractError(f"Invalid case count: {value!r}") from exc
    if numeric < 0 or not numeric.is_integer():
        raise SGContractError(f"Case count must be a non-negative integer: {value!r}")
    return int(numeric)


def _row(*, year: int, week: int, disease: object, cases: int, source_url: str,
         retrieved_at: str, raw_sha256: str, artifact: str, status: str,
         source_name: str = DEFAULT_SOURCE_NAME) -> Dict[str, str]:
    code, canonical = canonical_disease(disease)
    start = singapore_week_start(year, week)
    return {
        "Date": start.isoformat(), "RawDiseaseLabel": canonical,
        "SourceDiseaseCode": code, "DiseaseCode": "__source_native__",
        "Year": str(year), "Week": str(week), "EpiWeek": f"{year}-W{week:02d}",
        "Cases": str(cases), "Deaths": "", "ValueStatus": "reported",
        "ReportingArea": "Singapore", "GeographyKey": "country:SG:national",
        "Frequency": "weekly", "Measure": "case_notifications", "Unit": "count",
        "ReportingBasis": "national_notifiable_disease_surveillance",
        "TimeBasis": "Singapore epidemiological week (Sunday-Saturday)",
        "DatasetStatus": status, "IsProvisional": "true" if status == "provisional_revisable" else "false",
        "AuthoritativeRevision": "true", "MissingValuePolicy": "missing_is_unknown",
        "Source": source_name, "SourceScope": DEFAULT_SCOPE,
        "SourceURL": source_url, "RetrievedAt": retrieved_at,
        "RawArtifact": artifact, "RawSHA256": raw_sha256,
        "SourceContract": CONTRACT_VERSION, "PublicReleaseEnabled": "true",
        "LicenseReviewStatus": (
            "singapore_open_data_licence" if source_name == HISTORICAL_SOURCE_NAME
            else "operator_authorized_public_release"
        ),
    }


def parse_historical_csv(payload: bytes, *, source_url: str,
                         retrieved_at: Optional[str] = None) -> List[Dict[str, str]]:
    """Parse the data.gov.sg historical CSV without inventing absent rows."""
    digest = hashlib.sha256(payload).hexdigest()
    timestamp = retrieved_at or datetime.now(timezone.utc).isoformat()
    reader = csv.DictReader(io.StringIO(payload.decode("utf-8-sig")))
    expected = {"epi_week", "disease", "no._of_cases"}
    if set(reader.fieldnames or ()) != expected:
        raise SGContractError(f"Unexpected data.gov.sg columns: {reader.fieldnames!r}")
    rows: List[Dict[str, str]] = []
    seen: set[tuple[int, int, str]] = set()
    for raw in reader:
        raw_label = _text(raw["disease"])
        if raw_label.casefold() in {"hfmd", "hand, foot mouth disease", "hand foot mouth disease"}:
            continue  # average daily polyclinic attendance, not case notifications
        year, week = _parse_week_label(raw["epi_week"])
        code, _ = canonical_disease(raw_label)
        if code == "mpox" and singapore_week_start(year, week) < date(2022, 6, 26):
            continue
        cases = _parse_count(raw["no._of_cases"])
        if cases is None:
            continue
        key = (year, week, code)
        if key in seen:
            raise SGContractError(f"Duplicate historical disease/week row: {key}")
        seen.add(key)
        rows.append(_row(
            year=year, week=week, disease=raw_label, cases=cases,
            source_url=source_url, retrieved_at=timestamp, raw_sha256=digest,
            artifact="data.gov.sg CSV 2012-2022", status="final",
            source_name=HISTORICAL_SOURCE_NAME,
        ))
    return sorted(rows, key=lambda item: (item["Date"], item["SourceDiseaseCode"]))


def parse_annual_workbook(payload: bytes, *, year: int, source_url: str,
                          retrieved_at: Optional[str] = None) -> List[Dict[str, str]]:
    digest = hashlib.sha256(payload).hexdigest()
    timestamp = retrieved_at or datetime.now(timezone.utc).isoformat()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    except Exception as exc:
        raise SGContractError("CDA annual artifact is not a readable XLSX workbook") from exc
    sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    if _text(sheet.cell(1, 1).value) != "Epidemiology Wk" or _text(sheet.cell(1, 4).value) != "Total Number":
        raise SGContractError("CDA workbook notification table header changed")
    attendance_col = next((column for column in range(4, sheet.max_column + 1)
                           if _text(sheet.cell(1, column).value) == "Average Daily Number"), None)
    if attendance_col is None:
        raise SGContractError("CDA workbook has no Average Daily Number boundary")
    disease_columns: list[tuple[int, str]] = []
    for column in range(4, attendance_col):
        label = _text(sheet.cell(2, column).value)
        if label:
            canonical_disease(label)
            disease_columns.append((column, label))
    expected_columns = len(DISEASE_ALIASES) - 1  # historical-only Encephalitis is absent from CDA workbooks
    if len(disease_columns) != expected_columns:
        raise SGContractError(
            f"CDA workbook exposed {len(disease_columns)} notification columns; expected {expected_columns}"
        )
    rows: List[Dict[str, str]] = []
    seen: set[tuple[int, str]] = set()
    for row_number in range(3, sheet.max_row + 1):
        raw_week = sheet.cell(row_number, 1).value
        if raw_week in (None, ""):
            continue
        try:
            week = int(raw_week)
        except (TypeError, ValueError) as exc:
            raise SGContractError(f"Invalid CDA workbook week at row {row_number}: {raw_week!r}") from exc
        values = [(column, label, _parse_count(sheet.cell(row_number, column).value)) for column, label in disease_columns]
        if all(value is None for _, _, value in values):
            continue  # Future week scaffolding is not a zero observation.
        _parse_date_range(sheet.cell(row_number, 2).value, year=year, week=week)
        for _, label, cases in values:
            if cases is None:
                continue
            code, _ = canonical_disease(label)
            key = (week, code)
            if key in seen:
                raise SGContractError(f"Duplicate CDA workbook disease/week row: {key}")
            seen.add(key)
            rows.append(_row(year=year, week=week, disease=label, cases=cases,
                             source_url=source_url, retrieved_at=timestamp, raw_sha256=digest,
                             artifact=f"CDA annual XLSX {year}",
                             status="provisional_revisable" if year == datetime.now(timezone.utc).year else "closed_revisable"))
    if not rows:
        raise SGContractError(f"CDA workbook {year} contains no populated notification rows")
    return sorted(rows, key=lambda item: (item["Date"], item["SourceDiseaseCode"]))


def parse_weekly_pdf(payload: bytes, *, year: int, week: int, source_url: str,
                     retrieved_at: Optional[str] = None) -> List[Dict[str, str]]:
    """Parse the notification section/current-week column of one CDA PDF."""
    digest = hashlib.sha256(payload).hexdigest()
    timestamp = retrieved_at or datetime.now(timezone.utc).isoformat()
    try:
        with pdfplumber.open(io.BytesIO(payload)) as document:
            tables = document.pages[0].extract_tables()
    except Exception as exc:
        raise SGContractError(f"Unreadable CDA weekly PDF for {year}-W{week:02d}") from exc
    if not tables:
        raise SGContractError(f"CDA weekly PDF has no first-page table for {year}-W{week:02d}")
    rows: List[Dict[str, str]] = []
    seen: set[str] = set()
    in_notifications = True
    for cells in tables[0]:
        label = _text(cells[0] if cells else "")
        upper = label.upper()
        if "CASE NOTIFICATIONS" in upper or "INFECTIOUS DISEASES" in upper:
            in_notifications = True
            continue
        if any(marker in upper for marker in ("POLYCLINIC ATTENDANCES", "HIV/STI/TB NOTIFICATIONS")):
            break
        if not in_notifications or not label or len(cells) < 2:
            continue
        # Before E-week 27, MERS rows represented tests rather than confirmed cases.
        if year == 2023 and week < 27 and ("MERS" in upper or "MIDDLE EAST RESPIRATORY" in upper):
            continue
        try:
            code, _ = canonical_disease(label)
        except SGContractError:
            try:
                unknown_count = _parse_count(cells[1])
            except SGContractError:
                continue  # table headers
            if unknown_count is not None:
                raise SGContractError(
                    f"Unregistered numeric CDA PDF row for {year}-W{week:02d}: {label!r}"
                )
            continue  # section headings and footnotes
        cases = _parse_count(cells[1])
        if cases is None:
            continue
        if code in seen:
            raise SGContractError(f"Duplicate CDA PDF disease row: {year}-W{week:02d}/{code}")
        seen.add(code)
        rows.append(_row(year=year, week=week, disease=label, cases=cases,
                         source_url=source_url, retrieved_at=timestamp, raw_sha256=digest,
                         artifact=f"CDA weekly PDF {year}-W{week:02d}", status="closed_revisable"))
    if not rows:
        raise SGContractError(f"CDA weekly PDF produced no notification rows for {year}-W{week:02d}")
    return rows


class SingaporeCDACrawler(BaseCrawler):
    def __init__(self, *, save_raw: bool = False, raw_dir: Optional[Path] = None, **kwargs) -> None:
        kwargs.setdefault("user_agent", "Mozilla/5.0 (compatible; GlobalID/2.0; SG-CDA-WIDB)")
        super().__init__(**kwargs)
        self.save_raw = save_raw
        self.raw_dir = Path(raw_dir or "data/raw/sg")

    def _archive(self, name: str, payload: bytes, source_url: str) -> str:
        if not self.save_raw:
            return ""
        digest = hashlib.sha256(payload).hexdigest()
        folder = self.raw_dir / datetime.now(timezone.utc).strftime("%Y/%m/%d")
        folder.mkdir(parents=True, exist_ok=True)
        path = folder / f"{name}_{digest[:12]}"
        path.write_bytes(payload)
        path.with_suffix(path.suffix + ".json").write_text(json.dumps({
            "source_url": source_url, "retrieved_at": datetime.now(timezone.utc).isoformat(),
            "sha256": digest, "public_release_enabled": True,
            "license_review_status": "operator_authorized_public_release",
            "source_terms_status": "cda_written_permission_required",
        }, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        return str(path)

    def _historical(self) -> List[Dict[str, str]]:
        poll = self.get(OPEN_DATA_POLL_URL).json()
        url = ((poll.get("data") or {}).get("url") if isinstance(poll, dict) else None)
        if not url:
            raise SGContractError("data.gov.sg poll-download response has no data.url")
        response = self.get(url)
        self._archive("historical_2012_2022.csv", response.content, url)
        # The download URL is a long-lived-in-minutes signed S3 URL and can
        # exceed registry URI limits. Preserve it in raw archive metadata, but
        # attach the stable official dataset API URL to normalized facts.
        return parse_historical_csv(
            response.content,
            source_url=OPEN_DATA_POLL_URL,
        )

    def discover_year_assets(self, year: int) -> tuple[Optional[str], list[tuple[int, str]], bytes]:
        page_url = YEAR_PAGE.format(year=year)
        response = self.get(page_url)
        soup = BeautifulSoup(response.content, "html.parser")
        workbook_url: Optional[str] = None
        pdfs: dict[int, str] = {}
        for anchor in soup.select("a[href]"):
            href = urljoin(page_url, str(anchor.get("href")))
            label = _text(anchor.get_text(" "))
            if re.search(r"\.xlsx?(?:\?|$)", href, re.I):
                workbook_url = href
            if re.search(r"\.pdf(?:\?|$)", href, re.I):
                match = re.search(r"(?:week|ew|wk)[_\s-]*(\d{1,2})", f"{label} {href}", re.I)
                if match:
                    pdfs[int(match.group(1))] = href
        return workbook_url, sorted(pdfs.items()), response.content

    def _year(self, year: int) -> tuple[List[Dict[str, str]], int]:
        workbook_url, pdfs, page = self.discover_year_assets(year)
        self._archive(f"cda_page_{year}.html", page, YEAR_PAGE.format(year=year))
        if workbook_url:
            try:
                response = self.get(workbook_url)
                self._archive(f"cda_{year}.xlsx", response.content, workbook_url)
                return parse_annual_workbook(response.content, year=year, source_url=workbook_url), 2
            except Exception as exc:
                if year != 2023:
                    raise
                logger.warning("CDA 2023 workbook unavailable; falling back to weekly PDFs: {}", exc)
        if year != 2023 or not pdfs:
            raise SGContractError(f"No usable CDA annual source for {year}")
        rows: List[Dict[str, str]] = []
        for week, url in pdfs:
            response = self.get(url)
            self._archive(f"cda_{year}_w{week:02d}.pdf", response.content, url)
            rows.extend(parse_weekly_pdf(response.content, year=year, week=week, source_url=url))
        return rows, 1 + len(pdfs)

    @staticmethod
    def _write_merged(output_csv: Path, fresh_rows: Iterable[Mapping[str, str]]) -> List[Dict[str, str]]:
        merged: Dict[tuple[str, str], Dict[str, str]] = {}
        if output_csv.exists():
            with output_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                for raw in csv.DictReader(handle):
                    row = {str(key): _text(value) for key, value in raw.items()}
                    if row.get("Date") and row.get("SourceDiseaseCode"):
                        merged[(row["Date"], row["SourceDiseaseCode"])] = row
        for raw in fresh_rows:
            row = dict(raw)
            merged[(row["Date"], row["SourceDiseaseCode"])] = row
        rows = sorted(merged.values(), key=lambda item: (item["Date"], item["SourceDiseaseCode"]))
        fields = list(rows[0]) if rows else []
        for row in rows:
            for field in row:
                if field not in fields:
                    fields.append(field)
        output_csv.parent.mkdir(parents=True, exist_ok=True)
        temporary = output_csv.with_suffix(output_csv.suffix + ".tmp")
        with temporary.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader(); writer.writerows(rows)
        os.replace(temporary, output_csv)
        return rows

    def crawl_weekly_national(self, output_csv: Path, *, years: Sequence[int], include_history: bool = False) -> SGFetchSummary:
        fresh: List[Dict[str, str]] = []
        artifacts = 0
        requested = set(int(value) for value in years if int(value) >= 2023)
        if include_history:
            fresh.extend(self._historical()); artifacts += 1
            requested.update(range(2023, datetime.now(timezone.utc).year + 1))
        used_years = sorted(requested)
        for year in used_years:
            rows, count = self._year(year)
            fresh.extend(rows); artifacts += count
        all_rows = self._write_merged(Path(output_csv), fresh)
        latest = max((date.fromisoformat(row["Date"]) for row in all_rows), default=None)
        return SGFetchSummary(len(all_rows), tuple(used_years), latest, artifacts)

    async def crawl(self, **kwargs) -> List[CrawlerResult]:
        del kwargs
        return []

    def parse(self, response) -> List[CrawlerResult]:
        del response
        return []


__all__ = [
    "CONTRACT_VERSION", "DEFAULT_SCOPE", "DEFAULT_SOURCE_NAME", "DISEASE_ALIASES",
    "HISTORICAL_ONTOLOGY_SOURCE_ID", "HISTORICAL_SOURCE_NAME", "HISTORY_START_YEAR",
    "ONTOLOGY_SOURCE_ID", "SGContractError", "SGFetchSummary",
    "SingaporeCDACrawler", "canonical_disease", "parse_annual_workbook",
    "parse_historical_csv", "parse_weekly_pdf", "singapore_week_start", "stable_disease_code",
]
