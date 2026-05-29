"""
GlobalID V2 New Zealand PHF Science Crawler

Fetches monthly notifiable disease surveillance data from the PHF Science
(formerly ESR) Digital Library.

Data source: https://www.phfscience.nz/digital-library/
Format: ZIP files containing Excel workbooks (National, District, Rolling)

Public interface
----------------
  crawl(months, force)                  -> List[CrawlerResult]
  crawl_monthly_national(output_csv)    -> NZFetchSummary
  parse(response)                       -> []  (BaseCrawler contract)
"""
from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from src.core import get_logger
from .base import BaseCrawler, CrawlerResult

logger = get_logger(__name__)

# ── Constants ────────────────────────────────────────────────────────────────

_LIBRARY_BASE_URL = "https://www.phfscience.nz/digital-library/"
_LIBRARY_SEARCH_URL = (
    "https://www.phfscience.nz/digital-library/"
    "?q=monthly%20notifiable&page={page}"
    "&expertise=Public%20health"
    "&focusArea%5B0%5D=Public%20health%20%3E%20Health%20intelligence%20and%20disease%20surveillance"
    "&topic%5B0%5D=Public%20health%20%3E%20Health%20intelligence%20and%20disease%20surveillance"
    "%20%3E%20Notifiable%20diseases"
    "&researchType%5B0%5D=dashboardItem&researchType%5B1%5D=reportItem"
)

# Month name mapping
_MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12,
}

# Regex to extract year/month from report titles or filenames
# e.g. "Monthly notifiable disease surveillance data March 2026"
_TITLE_PATTERN = re.compile(
    r"(?:monthly\s+notifiable.*?)\s+(\w+)\s+(\d{4})",
    re.IGNORECASE,
)
# e.g. "202603MarchNational.xlsx" or "202603_march26.zip"
_FILENAME_PATTERN = re.compile(
    r"(\d{4})(\d{2})[\s_]?(\w+?)(?:National|District|Rolling|_)",
    re.IGNORECASE,
)

# Unicode superscript footnote markers to strip from disease names
_FOOTNOTE_PATTERN = re.compile(r"[¹²³⁴⁵⁶⁷⁸⁹⁰]+$")

_NON_NATIONAL_PDF_HINTS = ("rolling", "dhb", "district")


@dataclass
class NZFetchSummary:
    """Summary of a fetch operation."""
    row_count: int
    latest_date: Optional[date]
    months_fetched: int
    source_url: str


def _clean_disease_name(raw: str) -> str:
    """Strip footnote markers and normalize whitespace."""
    cleaned = _FOOTNOTE_PATTERN.sub("", raw).strip()
    return " ".join(cleaned.split())


def _parse_month_name(name: str) -> Optional[int]:
    """Convert month name to number."""
    return _MONTH_NAMES.get(name.lower().strip())


def _is_national_filename(name: str) -> bool:
    """Return true when a NZ surveillance archive member is the national file."""
    filename = Path(name).name.lower()
    stem = filename.rsplit(".", 1)[0]
    return "national" in stem or stem.endswith("nat")


def _select_national_pdf(pdf_files: List[str]) -> Optional[str]:
    """Pick the national PDF from a monthly archive, avoiding rolling/DHB tables."""
    if not pdf_files:
        return None

    national_pdfs = [name for name in pdf_files if _is_national_filename(name)]
    if national_pdfs:
        return min(national_pdfs, key=lambda value: (len(Path(value).name), value.lower()))

    non_rolling_pdfs = [
        name
        for name in pdf_files
        if not any(hint in Path(name).name.lower() for hint in _NON_NATIONAL_PDF_HINTS)
    ]
    if non_rolling_pdfs:
        return non_rolling_pdfs[0]

    return pdf_files[0]


class NewZealandPHFCrawler(BaseCrawler):
    """
    Crawler for New Zealand PHF Science monthly notifiable disease data.

    The data is published as ZIP files on the PHF Science Digital Library.
    Each ZIP contains Excel workbooks with national, district, and rolling
    12-month data.
    """

    SOURCE_URL = _LIBRARY_BASE_URL

    def __init__(
        self,
        *,
        save_raw: bool = False,
        raw_dir: Optional[Path] = None,
    ):
        super().__init__(timeout=60, max_retries=3, delay=2.0)
        self.save_raw = save_raw
        self.raw_dir = Path(raw_dir) if raw_dir else Path("data/raw/nz")

    # ── Phase 1: Discover available reports ──────────────────────────────────

    def fetch_report_index(self, max_pages: int = 10) -> List[Dict[str, Any]]:
        """
        Discover all monthly notifiable disease surveillance reports via sitemap
        and known URL patterns.

        Returns a list of dicts with keys: title, url, download_url, year, month, date
        """
        reports: List[Dict[str, Any]] = []
        seen_keys: set = set()

        # Strategy 1: Use sitemap.xml to find all report pages
        try:
            sitemap_reports = self._fetch_from_sitemap()
            for r in sitemap_reports:
                if r["key"] not in seen_keys:
                    seen_keys.add(r["key"])
                    reports.append(r)
        except Exception as e:
            logger.warning(f"[NZ] Sitemap fetch failed: {e}")

        reports.sort(key=lambda r: (r.get("year", 0), r.get("month", 0)))
        logger.info(f"[NZ] Index complete | total_reports={len(reports)}")
        return reports

    def _fetch_from_sitemap(self) -> List[Dict[str, Any]]:
        """Parse sitemap.xml to find monthly notifiable disease report pages."""
        import re as _re

        response = self.get("https://www.phfscience.nz/sitemap.xml")
        urls = _re.findall(
            r"<loc>(https://www\.phfscience\.nz/digital-library/monthly-notifiable[^<]+)</loc>",
            response.text,
        )

        reports: List[Dict[str, Any]] = []

        for url in urls:
            slug = url.rstrip("/").split("/")[-1]

            # Individual monthly pages: monthly-notifiable-disease-surveillance-{report|data}-{month}-{year}
            match = _re.search(
                r"monthly-notifiable-disease-surveillance-(?:report|data)-(\w+)-(\d{4})",
                slug,
            )
            if match:
                month_name, year_str = match.group(1), match.group(2)
                month = _parse_month_name(month_name)
                if month is not None:
                    year = int(year_str)
                    reports.append({
                        "title": f"Monthly notifiable disease surveillance {month_name} {year}",
                        "url": url,
                        "download_url": None,
                        "year": year,
                        "month": month,
                        "date": date(year, month, 1),
                        "key": f"{year:04d}-{month:02d}",
                        "type": "individual",
                    })
                continue

            # Yearly archive pages: monthly-notifiable-disease-surveillance-reports-for-{year}
            match = _re.search(r"reports-for-(\d{4})", slug)
            if match:
                year = int(match.group(1))
                reports.append({
                    "title": f"Monthly notifiable disease surveillance reports for {year}",
                    "url": url,
                    "download_url": None,
                    "year": year,
                    "month": 0,  # Represents full year
                    "date": date(year, 1, 1),
                    "key": f"{year:04d}-yearly",
                    "type": "yearly_archive",
                })

        return reports

    # ── Phase 2: Resolve download URLs ───────────────────────────────────────

    def resolve_download_url(self, report: Dict[str, Any]) -> Optional[str]:
        """
        Visit a report detail page and extract the ZIP download URL.
        """
        from bs4 import BeautifulSoup

        if report.get("download_url"):
            return report["download_url"]

        page_url = report.get("url")
        if not page_url:
            return None

        try:
            response = self.get(page_url)
            soup = BeautifulSoup(response.text, "html.parser")

            # Look for download link (ZIP file)
            for link in soup.find_all("a", href=True):
                href = link.get("href", "")
                if href.endswith(".zip"):
                    full_url = href if href.startswith("http") else f"https://www.phfscience.nz{href}"
                    return full_url

            logger.warning(f"[NZ] No ZIP download found on {page_url}")
            return None
        except Exception as e:
            logger.error(f"[NZ] Failed to resolve download URL from {page_url}: {e}")
            return None

    def resolve_yearly_archive(self, report: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        For a yearly archive page, download the master ZIP and return info
        about each monthly sub-ZIP or PDF inside.

        Returns list of dicts with: year, month, zip_bytes/pdf_bytes, type
        """
        download_url = self.resolve_download_url(report)
        if not download_url:
            return []

        try:
            response = self.session.get(download_url, timeout=self.timeout * 2)
            response.raise_for_status()
        except Exception as e:
            logger.error(f"[NZ] Failed to download yearly archive: {e}")
            return []

        monthly_items: List[Dict[str, Any]] = []
        try:
            with zipfile.ZipFile(io.BytesIO(response.content)) as outer_zip:
                for name in outer_zip.namelist():
                    match = re.search(r"(\d{4})-?(\d{2})", name)
                    if not match:
                        continue
                    year = int(match.group(1))
                    month = int(match.group(2))

                    if name.endswith(".zip"):
                        with outer_zip.open(name) as f:
                            monthly_items.append({
                                "year": year,
                                "month": month,
                                "zip_bytes": io.BytesIO(f.read()),
                                "filename": name,
                                "type": "zip",
                            })
                    elif name.endswith(".pdf"):
                        with outer_zip.open(name) as f:
                            monthly_items.append({
                                "year": year,
                                "month": month,
                                "pdf_bytes": f.read(),
                                "filename": name,
                                "type": "pdf",
                            })
        except Exception as e:
            logger.error(f"[NZ] Failed to parse yearly archive ZIP: {e}")

        logger.info(f"[NZ] Yearly archive {report['year']}: found {len(monthly_items)} monthly items")
        return monthly_items

    # ── Phase 3: Download and parse Excel data ───────────────────────────────

    def download_and_parse_zip(
        self,
        download_url: str,
        year: int,
        month: int,
    ) -> List[Dict[str, str]]:
        """
        Download a ZIP file and parse the National Excel workbook inside.

        Returns a list of row dicts with keys:
            Date, RawDiseaseLabel, Cases, Rate, CumulativeTotal,
            PrevYearCases, PrevYearCumulative, PrevYearRate
        """
        try:
            response = self.session.get(download_url, timeout=self.timeout)
            response.raise_for_status()
        except Exception as e:
            logger.error(f"[NZ] Download failed: {download_url} | {e}")
            return []

        zip_bytes = io.BytesIO(response.content)

        # Save raw if configured
        if self.save_raw:
            self._save_raw_zip(zip_bytes.getvalue(), year, month)

        return self._parse_zip_content(zip_bytes, year, month)

    def _parse_zip_content(
        self,
        zip_bytes: io.BytesIO,
        year: int,
        month: int,
    ) -> List[Dict[str, str]]:
        """Parse the National Excel/XLS/PDF file from a ZIP archive (handles nested ZIPs)."""
        import openpyxl

        try:
            with zipfile.ZipFile(zip_bytes) as zf:
                # Categorize files
                xlsx_national = None
                xls_national = None
                pdf_files: List[str] = []
                nested_zip = None

                for name in zf.namelist():
                    name_lower = name.lower()
                    if name_lower.endswith(".xlsx") and _is_national_filename(name):
                        xlsx_national = name
                    elif name_lower.endswith(".xls") and _is_national_filename(name):
                        xls_national = name
                    elif name_lower.endswith(".pdf"):
                        pdf_files.append(name)
                    elif name_lower.endswith(".zip"):
                        nested_zip = name

                # Priority: xlsx > xls > pdf > nested zip
                if xlsx_national:
                    with zf.open(xlsx_national) as f:
                        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
                        ws = wb.active
                        return self._parse_national_sheet(ws, year, month)

                if xls_national:
                    from src.data.parsers.nz_pdf_parser import parse_nz_xls
                    with zf.open(xls_national) as f:
                        return parse_nz_xls(f.read(), year, month)

                pdf_file = _select_national_pdf(pdf_files)
                if pdf_file:
                    from src.data.parsers.nz_pdf_parser import parse_nz_pdf
                    with zf.open(pdf_file) as f:
                        result = parse_nz_pdf(f.read(), year, month)
                        return result.rows

                # Try nested ZIP
                if nested_zip:
                    with zf.open(nested_zip) as f:
                        inner_bytes = io.BytesIO(f.read())
                        return self._parse_zip_content(inner_bytes, year, month)

                logger.warning(f"[NZ] No parseable file found in ZIP for {year}-{month:02d}")
                return []
        except Exception as e:
            logger.error(f"[NZ] Failed to parse ZIP for {year}-{month:02d}: {e}")
            return []

    def _parse_national_sheet(self, ws, year: int, month: int) -> List[Dict[str, str]]:
        """
        Parse the National surveillance data sheet.

        Expected structure:
          Row 1: Title header
          Row 2: Year headers (Current Year - YYYY / Previous Year - YYYY)
          Row 3: Column headers (Disease, Month Cases, Cumulative, Rate, ...)
          Row 4+: Disease data rows
        """
        rows: List[Dict[str, str]] = []
        report_date = date(year, month, 1)

        # Find the data start row (first row with a disease name in column A
        # that is not a header/title)
        data_start = None
        for row_idx in range(1, min(ws.max_row + 1, 10)):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and "disease" in str(cell_val).lower():
                data_start = row_idx + 1
                break

        if data_start is None:
            # Fallback: assume data starts at row 4
            data_start = 4

        value_col_offset = 0
        for row_idx in range(data_start, min(ws.max_row + 1, data_start + 8)):
            disease_raw = ws.cell(row=row_idx, column=1).value
            if not disease_raw or not str(disease_raw).strip():
                continue
            col_b_cases = self._safe_int(ws.cell(row=row_idx, column=2).value)
            col_c_cases = self._safe_int(ws.cell(row=row_idx, column=3).value)
            if col_b_cases is None and col_c_cases is not None:
                value_col_offset = 1
            break

        for row_idx in range(data_start, ws.max_row + 1):
            disease_raw = ws.cell(row=row_idx, column=1).value
            if not disease_raw or not str(disease_raw).strip():
                continue

            disease_str = str(disease_raw).strip()

            # Skip footnotes and "Other notifiable" summary lines
            if disease_str.startswith(("¹", "²", "³", "⁴", "⁵", "Other notifiable")):
                continue

            # Skip if it looks like a footnote (starts with a digit followed by text)
            if re.match(r"^[¹²³⁴⁵⁶⁷⁸⁹]", disease_str):
                continue

            disease_name = _clean_disease_name(disease_str)
            if not disease_name:
                continue

            # Column B: Current month cases
            cases_val = ws.cell(row=row_idx, column=2 + value_col_offset).value
            cases = self._safe_int(cases_val)

            # Column C: Cumulative total since 1 Jan
            cumulative_val = ws.cell(row=row_idx, column=3 + value_col_offset).value
            cumulative = self._safe_int(cumulative_val)

            # Column D: Current 12-month rate (per 100,000)
            rate_val = ws.cell(row=row_idx, column=4 + value_col_offset).value
            rate = self._safe_float(rate_val)

            # Column E: Previous year same month cases
            prev_cases_val = ws.cell(row=row_idx, column=5 + value_col_offset).value
            prev_cases = self._safe_int(prev_cases_val)

            # Column F: Previous year cumulative
            prev_cumulative_val = ws.cell(row=row_idx, column=6 + value_col_offset).value
            prev_cumulative = self._safe_int(prev_cumulative_val)

            # Column G: Previous year 12-month rate
            prev_rate_val = ws.cell(row=row_idx, column=7 + value_col_offset).value
            prev_rate = self._safe_float(prev_rate_val)

            if cases is None and cumulative is None:
                continue

            rows.append({
                "Date": report_date.isoformat(),
                "RawDiseaseLabel": disease_name,
                "Cases": str(cases if cases is not None else 0),
                "CumulativeTotal": str(cumulative) if cumulative is not None else "",
                "Rate": str(rate) if rate is not None else "",
                "PrevYearCases": str(prev_cases) if prev_cases is not None else "",
                "PrevYearCumulative": str(prev_cumulative) if prev_cumulative is not None else "",
                "PrevYearRate": str(prev_rate) if prev_rate is not None else "",
                "Year": str(year),
                "Month": str(month),
                "Source": "NZ PHF Science Monthly Notifiable Disease Surveillance",
            })

        logger.info(f"[NZ] Parsed {len(rows)} disease rows for {year}-{month:02d}")
        return rows

    def parse_rolling_sheet(
        self,
        zip_bytes: io.BytesIO,
        year: int,
        month: int,
    ) -> List[Dict[str, str]]:
        """
        Parse the Rolling 12-month Excel file from a ZIP archive.
        This provides monthly case counts for the past 12 months.

        Returns rows with Date, RawDiseaseLabel, Cases for each of the 12 months.
        """
        import openpyxl

        try:
            with zipfile.ZipFile(zip_bytes) as zf:
                rolling_file = None
                for name in zf.namelist():
                    if "rolling" in name.lower() and name.endswith(".xlsx"):
                        rolling_file = name
                        break

                if not rolling_file:
                    return []

                with zf.open(rolling_file) as f:
                    wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
                    ws = wb.active
                    return self._parse_rolling_data(ws, year, month)
        except Exception as e:
            logger.error(f"[NZ] Failed to parse Rolling sheet for {year}-{month:02d}: {e}")
            return []

    def _parse_rolling_data(self, ws, year: int, month: int) -> List[Dict[str, str]]:
        """
        Parse the rolling 12-month data sheet.

        Structure:
          Row 1: Title
          Row 2: Year headers (2026, 2025, etc.)
          Row 3: Month names (Mar, Feb, Jan, Dec, Nov, ...)
          Row 4+: Disease data (disease name in col A, monthly counts in cols B-M)
        """
        rows: List[Dict[str, str]] = []

        # Build the month/year mapping from columns B onwards
        # Row 2 has year values, Row 3 has month abbreviations
        month_columns: List[Tuple[int, int, int]] = []  # (col_idx, year, month)

        # Read year row (row 2) - years may span multiple columns
        year_values: Dict[int, int] = {}
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if val and isinstance(val, (int, float)):
                year_values[col] = int(val)

        # Forward-fill years across columns
        current_year = None
        for col in range(2, ws.max_column + 1):
            if col in year_values:
                current_year = year_values[col]
            if current_year:
                year_values[col] = current_year

        # Read month row (row 3)
        for col in range(2, min(ws.max_column + 1, 14)):  # Max 12 months
            month_name = ws.cell(row=3, column=col).value
            if not month_name:
                continue
            m = _parse_month_name(str(month_name).strip())
            y = year_values.get(col)
            if m is not None and y is not None:
                month_columns.append((col, y, m))

        if not month_columns:
            return []

        # Parse disease rows starting from row 4
        for row_idx in range(4, ws.max_row + 1):
            disease_raw = ws.cell(row=row_idx, column=1).value
            if not disease_raw or not str(disease_raw).strip():
                continue

            disease_str = str(disease_raw).strip()
            if disease_str.startswith(("¹", "²", "³", "⁴", "⁵", "Other")):
                continue

            disease_name = _clean_disease_name(disease_str)
            if not disease_name:
                continue

            for col_idx, col_year, col_month in month_columns:
                cases_val = ws.cell(row=row_idx, column=col_idx).value
                cases = self._safe_int(cases_val)
                if cases is None:
                    cases = 0

                report_date = date(col_year, col_month, 1)
                rows.append({
                    "Date": report_date.isoformat(),
                    "RawDiseaseLabel": disease_name,
                    "Cases": str(cases),
                    "Year": str(col_year),
                    "Month": str(col_month),
                    "Source": "NZ PHF Science Monthly Notifiable Disease Surveillance",
                })

        return rows

    def download_and_parse_rolling(
        self,
        download_url: str,
        year: int,
        month: int,
    ) -> List[Dict[str, str]]:
        """
        Download a ZIP and parse the Rolling 12-month workbook.
        Returns rows for all 12 months in the rolling window.
        """
        try:
            response = self.session.get(download_url, timeout=self.timeout)
            response.raise_for_status()
        except Exception as e:
            logger.error(f"[NZ] Download failed for rolling: {download_url} | {e}")
            return []

        zip_bytes = io.BytesIO(response.content)
        return self.parse_rolling_sheet(zip_bytes, year, month)

    # ── Full crawl pipeline ──────────────────────────────────────────────────

    def crawl_monthly_national(
        self,
        output_csv: Path,
        *,
        months: Optional[List[Tuple[int, int]]] = None,
        max_pages: int = 10,
    ) -> NZFetchSummary:
        """
        Crawl NZ monthly data and write to a CSV file.

        If months is None, fetches the most recent 3 months.
        Uses the Rolling workbook to get historical monthly breakdowns when
        available, and falls back to individual report pages or yearly archives.
        """
        import csv as csv_mod

        # Discover available reports
        index = self.fetch_report_index(max_pages=max_pages)
        if not index:
            raise RuntimeError("[NZ] No reports found in Digital Library index")

        # Determine which months to fetch
        if months is None:
            now = datetime.now()
            target_months = set()
            for delta in range(3):
                m = now.month - delta
                y = now.year
                if m <= 0:
                    m += 12
                    y -= 1
                target_months.add((y, m))
        else:
            target_months = set(months)

        all_rows: List[Dict[str, str]] = []
        fetched_months: set = set()

        # Strategy 1: Use individual monthly report pages (most recent data)
        individual_reports = sorted(
            [r for r in index if r.get("type") == "individual"],
            key=lambda r: (r["year"], r["month"]),
            reverse=True,
        )

        for report in individual_reports:
            if not target_months - fetched_months:
                break

            report_ym = (report["year"], report["month"])

            # Skip if this month isn't needed
            if report_ym not in (target_months - fetched_months):
                # But check if the Rolling sheet covers months we need
                remaining = target_months - fetched_months
                # Rolling covers 12 months back from the report month
                rolling_coverage = set()
                for delta in range(12):
                    m = report["month"] - delta
                    y = report["year"]
                    if m <= 0:
                        m += 12
                        y -= 1
                    rolling_coverage.add((y, m))
                if not (remaining & rolling_coverage):
                    continue

            download_url = self.resolve_download_url(report)
            if not download_url:
                continue

            # Get National sheet data for this month
            if report_ym in (target_months - fetched_months):
                national_rows = self.download_and_parse_zip(
                    download_url, report["year"], report["month"]
                )
                if national_rows:
                    all_rows.extend(national_rows)
                    fetched_months.add(report_ym)

            # Use Rolling sheet to fill in other months
            remaining = target_months - fetched_months
            if remaining:
                rolling_rows = self.download_and_parse_rolling(
                    download_url, report["year"], report["month"]
                )
                for row in rolling_rows:
                    row_ym = (int(row["Year"]), int(row["Month"]))
                    if row_ym in remaining:
                        all_rows.append(row)
                # Mark months as fetched based on rolling data
                rolling_months_found = set()
                for row in rolling_rows:
                    rolling_months_found.add((int(row["Year"]), int(row["Month"])))
                fetched_months.update(rolling_months_found & remaining)

        # Strategy 2: Use yearly archives for older months still missing
        remaining = target_months - fetched_months
        if remaining:
            yearly_reports = sorted(
                [r for r in index if r.get("type") == "yearly_archive"],
                key=lambda r: r["year"],
                reverse=True,
            )

            needed_years = {ym[0] for ym in remaining}
            for report in yearly_reports:
                if not remaining:
                    break
                if report["year"] not in needed_years:
                    continue

                logger.info(f"[NZ] Fetching yearly archive for {report['year']}")
                monthly_zips = self.resolve_yearly_archive(report)

                for mz in monthly_zips:
                    mz_ym = (mz["year"], mz["month"])
                    if mz_ym not in remaining:
                        continue

                    if mz.get("type") == "pdf":
                        from src.data.parsers.nz_pdf_parser import parse_nz_pdf
                        result = parse_nz_pdf(mz["pdf_bytes"], mz["year"], mz["month"])
                        rows = result.rows
                    else:
                        rows = self._parse_zip_content(mz["zip_bytes"], mz["year"], mz["month"])

                    if rows:
                        all_rows.extend(rows)
                        fetched_months.add(mz_ym)
                        remaining = target_months - fetched_months

        if not all_rows:
            raise RuntimeError("[NZ] No data rows parsed from any report")

        # Deduplicate: keep the row with most metadata for each (date, disease)
        seen: Dict[Tuple[str, str], Dict[str, str]] = {}
        for row in all_rows:
            key = (row["Date"], row["RawDiseaseLabel"])
            if key not in seen or "CumulativeTotal" in row:
                seen[key] = row

        deduped_rows = sorted(seen.values(), key=lambda r: (r["Date"], r["RawDiseaseLabel"]))

        # Write output CSV
        output_csv.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = [
            "", "Disease", "Year", "Month", "Date", "Cases",
            "CumulativeTotal", "Rate", "Source",
        ]
        with output_csv.open("w", encoding="utf-8", newline="") as handle:
            writer = csv_mod.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for idx, row in enumerate(deduped_rows, start=1):
                writer.writerow({
                    "": str(idx),
                    "Disease": row["RawDiseaseLabel"],
                    "Year": row.get("Year", ""),
                    "Month": row.get("Month", ""),
                    "Date": row["Date"],
                    "Cases": row["Cases"],
                    "CumulativeTotal": row.get("CumulativeTotal", ""),
                    "Rate": row.get("Rate", ""),
                    "Source": row.get("Source", ""),
                })

        latest_date = max(
            (datetime.strptime(r["Date"], "%Y-%m-%d").date() for r in deduped_rows),
            default=None,
        )

        return NZFetchSummary(
            row_count=len(deduped_rows),
            latest_date=latest_date,
            months_fetched=len(fetched_months),
            source_url=_LIBRARY_BASE_URL,
        )

    # ── BaseCrawler contract ─────────────────────────────────────────────────

    async def crawl(self, **kwargs) -> List[CrawlerResult]:
        """Execute the full crawl pipeline."""
        months = kwargs.get("months")
        force = kwargs.get("force", False)
        max_pages = kwargs.get("max_pages", 10)

        index = self.fetch_report_index(max_pages=max_pages)
        results: List[CrawlerResult] = []

        for report in index:
            download_url = self.resolve_download_url(report)
            if not download_url:
                continue

            national_rows = self.download_and_parse_zip(
                download_url, report["year"], report["month"]
            )
            if national_rows:
                results.append(CrawlerResult(
                    title=report.get("title", f"NZ Monthly {report['year']}-{report['month']:02d}"),
                    url=report.get("url"),
                    content=None,
                    date=datetime(report["year"], report["month"], 1, tzinfo=timezone.utc),
                    year_month=f"{report['year']} {self._month_name(report['month'])}",
                    metadata={
                        "language": "en",
                        "source": "nz_phf_monthly",
                        "country_code": "NZ",
                        "download_url": download_url,
                    },
                    raw_data={"rows": national_rows},
                ))

        return results

    def parse(self, response) -> List[CrawlerResult]:
        """BaseCrawler contract — not used for NZ (parsing is integrated)."""
        return []

    # ── Helpers ──────────────────────────────────────────────────────────────

    def _save_raw_zip(self, content: bytes, year: int, month: int) -> None:
        """Archive raw ZIP to disk."""
        archive_dir = self.raw_dir / str(year) / f"{month:02d}"
        archive_dir.mkdir(parents=True, exist_ok=True)
        archive_path = archive_dir / f"nz_{year}{month:02d}_raw.zip"
        archive_path.write_bytes(content)
        logger.debug(f"[NZ] Saved raw ZIP: {archive_path}")

    @staticmethod
    def _safe_int(value: Any) -> Optional[int]:
        if value is None:
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _safe_float(value: Any) -> Optional[float]:
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _month_name(month: int) -> str:
        names = [
            "", "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December",
        ]
        return names[month] if 1 <= month <= 12 else ""
