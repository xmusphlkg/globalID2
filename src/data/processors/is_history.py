"""Parse Iceland's heterogeneous historical Excel surveillance releases.

This module deliberately preserves three source-series families instead of
pretending that they are one continuous metric:

``registry_annual``
    Annual case notifications from the historical national registry table.
``registry_disease_monthly``
    Disease-specific monthly case-notification workbooks (and published
    incidence where present).
``legacy_icd_monthly``
    Monthly *registered diagnoses* from primary-care ICD tables.  These are a
    different reporting basis and are never projected into legacy
    ``disease_records`` rows.

Only explicitly reviewed disease identities are candidates, and candidates
are emitted to ``series_rows`` only when the checked-in ontology contains the
exact source-series code, source, local code, label, concept, frequency, and
measure.  Everything else is counted and retained in quarantine metadata.
"""

from __future__ import annotations

import ast
import csv
import hashlib
import json
import math
import os
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import openpyxl
import xlrd
from openpyxl.utils.cell import column_index_from_string
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from src.data.crawlers.is_history import (
    IcelandHistoryRawFile,
    IcelandHistoryWorkbookSpec,
    OFFICIAL_WORKBOOKS,
)
from src.services.disease_series_policy import (
    SOURCE_OBSERVATIONS_ONLY_POLICY,
    is_case_count_series,
    select_series_projection,
)


ROOT = Path(__file__).resolve().parents[3]
DEFAULT_ONTOLOGY_PATH = ROOT / "configs" / "disease_ontology.json"
DEFAULT_OUTPUT_DIR = ROOT / "data" / "current" / "is" / "history"
PARSED_MANIFEST_SCHEMA = "globalid.iceland-history-parsed-manifest.v1"
PARSER_VERSION = "1.0.0"

HISTORY_SOURCE_ID = "SRC_IS_DOH_HISTORY"
LEGACY_SOURCE_ID = "SRC_IS_DOH_LEGACY_ICD"
HISTORY_SOURCE_NAME = "Iceland Directorate of Health Historical Registry"
LEGACY_SOURCE_NAME = "Iceland Directorate of Health Legacy ICD Monthly"


class IcelandWorkbookLayoutError(ValueError):
    """Raised when a workbook no longer matches a reviewed layout."""


@dataclass(frozen=True)
class ReviewedSeries:
    """One manually reviewed source identity and its canonical target."""

    key: str
    token: str
    concept_id: str
    label_is: str
    label_en: str
    aliases: tuple[str, ...] = ()

    def source_series_code(self, source_kind: str) -> str:
        if source_kind == "registry_annual":
            return f"SER_IS_HISTORY_{self.token}_ANNUAL"
        if source_kind == "registry_disease_monthly":
            return f"SER_IS_HISTORY_{self.token}_MONTHLY"
        raise ValueError(f"Unsupported history source kind: {source_kind}")

    def local_code(self, source_kind: str) -> str:
        if source_kind == "registry_annual":
            return f"IS_HISTORY_ANNUAL_{self.token}"
        if source_kind == "registry_disease_monthly":
            return f"IS_HISTORY_MONTHLY_{self.token}"
        raise ValueError(f"Unsupported history source kind: {source_kind}")


@dataclass(frozen=True)
class ReviewedLegacySeries:
    """A legacy diagnosis series identified by ICD-10 text and local label."""

    token: str
    concept_id: str
    icd10: str
    label_is: str
    label_en: str

    @property
    def source_series_code(self) -> str:
        return f"SER_IS_LEGACY_ICD_{self.token}_MONTHLY"

    @property
    def local_code(self) -> str:
        return f"IS_LEGACY_ICD_{self.token}"


@dataclass
class IcelandHistoryPreparedResult:
    """Normalized projections, lossless series rows, and audit artifacts."""

    rows: list[dict[str, Any]]
    series_rows: list[dict[str, Any]]
    quarantine: list[dict[str, Any]]
    manifest: dict[str, Any]
    raw_hashes: dict[str, str]


@dataclass(frozen=True)
class IcelandHistoryImportResult:
    inserted_or_updated: int
    skipped_unmapped: int
    source_latest_date: date | None
    skipped_current_precedence: int = 0


@dataclass
class _ParseContext:
    stats: Counter[str] = field(default_factory=Counter)
    quarantine: list[dict[str, Any]] = field(default_factory=list)

    def reject(self, reason: str, **details: Any) -> None:
        self.stats[f"quarantine.{reason}"] += 1
        self.quarantine.append({"reason": reason, **details})


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    normalized = unicodedata.normalize("NFKC", str(value).replace("\ufeff", ""))
    return " ".join(normalized.split()).strip()


def _label_key(value: object) -> str:
    value_text = _text(value)
    value_text = re.sub(r"\s*\*+\s*$", "", value_text)
    value_text = value_text.replace("–", "-").replace("—", "-")
    return value_text.casefold().strip(" .:;")


def _explanation_key(value: object) -> str:
    key = _label_key(value)
    key = re.sub(r"\s*\([^)]*\)\s*$", "", key)
    return key.strip()


def _normalize_icd(value: object) -> str:
    code = _text(value).upper().replace("–", "-").replace("—", "-")
    code = re.sub(r"\s*,\s*", ", ", code)
    code = re.sub(r"\s*-\s*", "-", code)
    return " ".join(code.split())


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _as_year(value: object) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and float(value).is_integer():
        year = int(value)
        return year if 1900 <= year <= 2100 else None
    text_value = _text(value)
    if re.fullmatch(r"(?:19|20)\d{2}", text_value):
        return int(text_value)
    return None


MONTHS = {
    "jan": 1,
    "january": 1,
    "janúar": 1,
    "feb": 2,
    "february": 2,
    "febrúar": 2,
    "mar": 3,
    "march": 3,
    "mars": 3,
    "apr": 4,
    "april": 4,
    "apríl": 4,
    "may": 5,
    "maí": 5,
    "jun": 6,
    "jún": 6,
    "june": 6,
    "júní": 6,
    "jul": 7,
    "júl": 7,
    "july": 7,
    "júlí": 7,
    "aug": 8,
    "ág": 8,
    "ágú": 8,
    "august": 8,
    "ágúst": 8,
    "sep": 9,
    "september": 9,
    "oct": 10,
    "okt": 10,
    "october": 10,
    "október": 10,
    "nov": 11,
    "nóv": 11,
    "november": 11,
    "nóvember": 11,
    "dec": 12,
    "des": 12,
    "december": 12,
    "desember": 12,
}


def _month_number(value: object) -> int | None:
    token = _label_key(value).split(" ", 1)[0].strip(".:")
    return MONTHS.get(token)


# The annual list contains only mappings whose source wording and target
# concept have been manually checked.  Five published categories remain
# intentionally absent: Anisakiasis, acute toxin/radiological symptoms,
# influenza-like illness, hepatitis non-A--E, and unexpected health events.
ANNUAL_SERIES: tuple[ReviewedSeries, ...] = (
    ReviewedSeries("aids", "AIDS", "D005", "Alnæmi", "AIDS"),
    ReviewedSeries("cysticercosis", "CYSTICERCOSIS", "D088", "Bandormslirfusýki", "Cysticercosis"),
    ReviewedSeries("diphtheria", "DIPHTHERIA", "D029", "Barnaveiki", "Diphtheria"),
    ReviewedSeries("dengue", "DENGUE", "D021", "Beinbrunasótt", "Dengue"),
    ReviewedSeries("tuberculosis", "TUBERCULOSIS", "D025", "Berklar", "Tuberculosis"),
    ReviewedSeries("viral_hemorrhagic_fever", "VIRAL_HEMORRHAGIC_FEVER", "D127", "Blæðandi veiruhitasóttir", "Viral hemorrhagic fevers", ("Hemorrhagic viral fever",)),
    ReviewedSeries("smallpox", "SMALLPOX", "D064", "Bólusótt", "Smallpox"),
    ReviewedSeries("botulism", "BOTULISM", "D091", "Bótúlismi", "Botulism"),
    ReviewedSeries("esbl_ampc", "ESBL_AMPC", "D237", "Breiðvirkir betalaktamasamyndandi sýklar (ESBL/AmpC)", "ESBL/AmpC-producing organism surveillance", ("Breiðvirkir betalaktamasamyndandi sýklar (ESBL)", "Extended Spectrum Beta Lactamase (ESBL)", "Extended Spectrum Beta Lactamase (ESBL)*")),
    ReviewedSeries("covid_19", "COVID_19", "D004", "COVID-19", "COVID-19"),
    ReviewedSeries("sars", "SARS", "D003", "Córónaveirulungnabólga", "SARS"),
    ReviewedSeries("cjd", "CJD", "D144", "Creutzfeldt Jakobs veiki/afbrigði", "Creutzfeldt-Jakob disease", ("New variant Creutzfeldt Jakobs Disease (CJD)",)),
    ReviewedSeries("stec", "STEC", "D115", "Enterohaemorrhagisk E. coli sýking", "Shiga toxin-producing Escherichia coli infection", ("Enterohemorrhagisk E. coli sýking", "Enterohaemorrhagic E. coli infection")),
    ReviewedSeries("giardiasis", "GIARDIASIS", "D099", "Giardiasis", "Giardiasis"),
    ReviewedSeries("yellow_fever", "YELLOW_FEVER", "D059", "Gulusótt", "Yellow fever"),
    ReviewedSeries("hib", "HIB", "D100", "Haemofilus influenzae sýking b", "Haemophilus influenzae type b", ("Hemofilus influenzae sýking b", "Haemophilus influenzae infection type b", "Hemophilus influenzae infection type b")),
    ReviewedSeries("mumps", "MUMPS", "D039", "Hettusótt", "Mumps"),
    ReviewedSeries("tularemia", "TULAREMIA", "D123", "Hérasótt", "Tularemia"),
    ReviewedSeries("hiv", "HIV", "D162", "HIV sýking (human immunod. virus)", "HIV infection"),
    ReviewedSeries("leprosy", "LEPROSY", "D042", "Holdsveiki", "Leprosy", ("Lepra",)),
    ReviewedSeries("q_fever", "Q_FEVER", "D113", "Huldusótt", "Q fever", ("Q-fever",)),
    ReviewedSeries("rabies", "RABIES", "D019", "Hundaæði", "Rabies"),
    ReviewedSeries("pandemic_influenza_h1n1", "PANDEMIC_INFLUENZA_A_H1N1_2009", "D038", "Inflúensa A (H1N1) 2009", "Influenza A(H1N1)pdm09", ("Pandemic influenzae A(H1N1) 2009",)),
    ReviewedSeries("influenza_a_h3", "INFLUENZA_A_H3", "D038", "Inflúensa A H3", "Influenza A(H3)", ("Influenzae A(H3)",)),
    ReviewedSeries("invasive_h_influenzae", "INVASIVE_HAEMOPHILUS_INFLUENZAE", "D100", "Ífarandi Haemophilus influenzae sýking", "Invasive Haemophilus influenzae", ("Ífarandi Hemófílus inflúensusýking", "Invasive Hemophilus influenzae")),
    ReviewedSeries("invasive_pneumococcal", "INVASIVE_PNEUMOCOCCAL", "D106", "Ífarandi pneumókokkasýkingar", "Invasive pneumococcal disease", ("Invasive pneumococcal infections",)),
    ReviewedSeries("yersiniosis", "YERSINIOSIS", "D155", "Jersiníusýking", "Yersiniosis", ("Yersinia enterocolitica, Yersinia pseudotuberculosis",)),
    ReviewedSeries("campylobacteriosis", "CAMPYLOBACTERIOSIS", "D092", "Kampýlóbaktersýking", "Campylobacteriosis"),
    ReviewedSeries("pertussis", "PERTUSSIS", "D028", "Kikhósti", "Pertussis"),
    ReviewedSeries("chlamydia", "CHLAMYDIA", "D094", "Klamydíusýking", "Chlamydia trachomatis infection", ("Chlamydia trachomatis",)),
    ReviewedSeries("cholera", "CHOLERA", "D002", "Kólera og kólerulíkar sýkingar", "Cholera"),
    ReviewedSeries("cryptosporidiosis", "CRYPTOSPORIDIOSIS", "D096", "Launsporasýking (cryptósporidium sýking)", "Cryptosporidiosis"),
    ReviewedSeries("legionellosis", "LEGIONELLOSIS", "D107", "Legíónellusýking", "Legionellosis"),
    ReviewedSeries("gonorrhea", "GONORRHEA", "D033", "Lekandi", "Gonorrhea", ("Gonorrhoea",)),
    ReviewedSeries("leptospirosis", "LEPTOSPIROSIS", "D035", "Leptóspirusýking", "Leptospirosis"),
    ReviewedSeries("hepatitis_a", "HEPATITIS_A", "D007", "Lifrarbólga A", "Hepatitis A"),
    ReviewedSeries("hepatitis_b_combined", "HEPATITIS_B_COMBINED", "D008", "Lifrarbólga B (bráð, viðvarandi)", "Hepatitis B (acute and chronic combined)", ("Hepatitis B (acute, chronic)",)),
    ReviewedSeries("hepatitis_c", "HEPATITIS_C", "D009", "Lifrarbólga C", "Hepatitis C"),
    ReviewedSeries("hepatitis_e", "HEPATITIS_E", "D011", "Lifrarbólga E", "Hepatitis E"),
    ReviewedSeries("listeriosis", "LISTERIOSIS", "D108", "Listeríusýking", "Listeriosis"),
    ReviewedSeries("poliomyelitis", "POLIOMYELITIS", "D013", "Lömunarveiki", "Poliomyelitis"),
    ReviewedSeries("malaria", "MALARIA", "D037", "Malaría", "Malaria"),
    ReviewedSeries("meningococcal", "MENINGOCOCCAL", "D110", "Meningókokkasjúkdómur", "Meningococcal disease"),
    ReviewedSeries("mrsa", "MRSA", "D236", "Methicillin ónæmur stafýlokokkus aureus, MÓSA", "MRSA surveillance", ("Methicillin resistant Staphylococcus aureus (MRSA)",)),
    ReviewedSeries("anthrax", "ANTHRAX", "D023", "Miltisbrandur", "Anthrax"),
    ReviewedSeries("measles", "MEASLES", "D017", "Mislingar", "Measles"),
    ReviewedSeries("rubella", "RUBELLA", "D040", "Rauðir hundar", "Rubella"),
    ReviewedSeries("salmonellosis", "SALMONELLOSIS", "D114", "Salmonellusýking", "Salmonellosis"),
    ReviewedSeries("syphilis", "SYPHILIS", "D034", "Sárasótt", "Syphilis"),
    ReviewedSeries("shigellosis", "SHIGELLOSIS", "D105", "Sígellusýking", "Shigellosis"),
    ReviewedSeries("tetanus", "TETANUS", "D120", "Stífkrampi", "Tetanus"),
    ReviewedSeries("echinococcosis", "ECHINOCOCCOSIS", "D045", "Sullaveiki", "Echinococcosis"),
    ReviewedSeries("plague", "PLAGUE", "D001", "Svarti dauði", "Plague"),
    ReviewedSeries("toxoplasmosis", "TOXOPLASMOSIS", "D085", "Toxóplasmasýking", "Toxoplasmosis"),
    ReviewedSeries("typhoid_paratyphoid", "TYPHOID_PARATYPHOID", "D026", "Taugaveiki/taugaveikibróðir", "Typhoid and paratyphoid fever", ("Typhoid/paratyphoid fever",)),
    ReviewedSeries("trichinellosis", "TRICHINELLOSIS", "D122", "Tríkínusýking", "Trichinellosis", ("Trichinosis/Trichinellosis",)),
    ReviewedSeries("vre", "VRE", "D125", "Vankomýcín ónæmur enterókokkur", "Vancomycin-resistant Enterococcus (VRE)", ("Vancomycin resistant Enterococcus (VRE)",)),
    ReviewedSeries("west_nile", "WEST_NILE", "D128", "Vesturnílarveirusótt", "West Nile virus infection", ("West Nile Virus Infection",)),
    ReviewedSeries("brucellosis", "BRUCELLOSIS", "D032", "Öldusótt", "Brucellosis"),
)


MONTHLY_SERIES: dict[str, ReviewedSeries] = {
    item.key: item
    for item in (
        ReviewedSeries("esbl_ampc", "ESBL_AMPC", "D237", "Breiðvirkir beta-laktamasar (ESBL)", "ESBL/AmpC-producing organism surveillance"),
        ReviewedSeries("giardiasis", "GIARDIASIS", "D099", "Giardíusýking", "Giardiasis"),
        ReviewedSeries("hiv", "HIV", "D162", "HIV sýking", "HIV infection"),
        ReviewedSeries("campylobacteriosis", "CAMPYLOBACTERIOSIS", "D092", "Kampýlóbaktersýking", "Campylobacteriosis"),
        ReviewedSeries("pertussis", "PERTUSSIS", "D028", "Kikhósti", "Pertussis"),
        ReviewedSeries("chlamydia", "CHLAMYDIA", "D094", "Klamydíusýking", "Chlamydia trachomatis infection"),
        ReviewedSeries("gonorrhea", "GONORRHEA", "D033", "Lekandi", "Gonorrhea"),
        ReviewedSeries("hepatitis_a", "HEPATITIS_A", "D007", "Lifrarbólga A", "Hepatitis A"),
        ReviewedSeries("hepatitis_b_combined", "HEPATITIS_B_COMBINED", "D008", "Lifrarbólga B", "Hepatitis B (acute and chronic combined)"),
        ReviewedSeries("hepatitis_c", "HEPATITIS_C", "D009", "Lifrarbólga C", "Hepatitis C"),
        ReviewedSeries("invasive_pneumococcal", "INVASIVE_PNEUMOCOCCAL", "D106", "Ífarandi pneumókokkasýking", "Invasive pneumococcal disease"),
        ReviewedSeries("salmonellosis", "SALMONELLOSIS", "D114", "Salmonellusýking", "Salmonellosis"),
        ReviewedSeries("syphilis", "SYPHILIS", "D034", "Sárasótt", "Syphilis"),
        ReviewedSeries("vre", "VRE", "D125", "Vankómýsínónæmir enterókokkar (VRE)", "Vancomycin-resistant Enterococcus (VRE)"),
    )
}


LEGACY_SERIES: tuple[ReviewedLegacySeries, ...] = (
    ReviewedLegacySeries("B01_HLAUPABOLA", "D054", "B01", "Hlaupabóla", "Varicella"),
    ReviewedLegacySeries("B01_B01_9_HLAUPABOLA", "D054", "B01-B01.9", "Hlaupabóla", "Varicella"),
    ReviewedLegacySeries("J10_J11_INFLUENSA", "D038", "J10-J11", "Inflúensa", "Influenza"),
    ReviewedLegacySeries("J09_J10_J10_8_U05_9_STADFEST_INFLUENSA", "D038", "J09, J10-J10.8, U05.9", "Staðfest inflúensa", "Laboratory-confirmed influenza"),
    ReviewedLegacySeries("B02_RISTILL", "D063", "B02", "Ristill", "Zoster"),
    ReviewedLegacySeries("B02_B02_9_RISTILL", "D063", "B02-B02.9", "Ristill", "Zoster"),
    ReviewedLegacySeries("A38_SKARLATSOTT", "D031", "A38", "Skarlatsótt", "Scarlet fever"),
    ReviewedLegacySeries("A38_SKARLATSSOTT", "D031", "A38", "Skarlatssótt", "Scarlet fever"),
    ReviewedLegacySeries("J02_0_J03_0_STREP_PHARYNGITIS_LIST", "D224", "J02.0, J03.0", "Streptókokka-hálsbólga", "Group A streptococcal pharyngitis"),
    ReviewedLegacySeries("J02_0_J03_0_STREP_PHARYNGITIS_RANGE", "D224", "J02.0-J03.0", "Streptókokka-hálsbólga", "Group A streptococcal pharyngitis"),
    ReviewedLegacySeries("A09_BRADUR_NIDURGANGUR", "D047", "A09", "Bráður niðurgangur", "Acute infectious diarrhea"),
    ReviewedLegacySeries("A09_NIDURGANGUR", "D047", "A09", "Niðurgangur", "Infectious diarrhea"),
    ReviewedLegacySeries("A37_KIKHOSTI", "D028", "A37", "Kikhósti", "Pertussis"),
    ReviewedLegacySeries("A36_BARNAVEIKI", "D029", "A36", "Barnaveiki", "Diphtheria"),
    ReviewedLegacySeries("B26_HETTUSOTT", "D039", "B26", "Hettusótt", "Mumps"),
    ReviewedLegacySeries("B05_MISLINGAR", "D017", "B05", "Mislingar", "Measles"),
    ReviewedLegacySeries("B06_RAUDIR_HUNDAR", "D040", "B06", "Rauðir hundar", "Rubella"),
    ReviewedLegacySeries("A35_STIFKRAMPI", "D120", "A35", "Stífkrampi", "Tetanus"),
    ReviewedLegacySeries("B58_TOXOPLASMASYKING", "D085", "B58", "Toxóplasmasýking", "Toxoplasmosis"),
    ReviewedLegacySeries("B54_MYRARKALDA", "D037", "B54", "Mýrarkalda", "Malaria"),
    ReviewedLegacySeries("B54_MYRARKALDA_MALARIA", "D037", "B54", "Mýrarkalda (malaría)", "Malaria"),
    ReviewedLegacySeries("A70_PSITTACOSIS", "D112", "A70", "Psittacosis", "Psittacosis"),
    ReviewedLegacySeries("A70_FYLASOTT", "D112", "A70", "Fýlasótt (psittacosis)", "Psittacosis"),
    ReviewedLegacySeries("A69_2_LYME", "D109", "A69.2", "Lyme sjúkdómur", "Lyme disease"),
    ReviewedLegacySeries("A69_2_LYME_BORRELIOSIS", "D109", "A69.2", "Lyme sjúkdómur (borreliosis)", "Lyme disease"),
    ReviewedLegacySeries("G00_BACTERIAL_MENINGITIS", "D135", "G00", "Heilahimnubólga af völdum sýkla", "Bacterial meningitis"),
    ReviewedLegacySeries("G00_RANGE_BACTERIAL_MENINGITIS", "D135", "G00, G00.2, G00.3-G00.9", "Heilahimnubólga af völdum sýkla", "Bacterial meningitis"),
    ReviewedLegacySeries("G00_RANGE_BACTERIAL_MENINGITIDES", "D135", "G00, G00.2, G00.3-G00.9", "Heilahimnubólgur af völdum baktería", "Bacterial meningitis"),
    ReviewedLegacySeries("A08_0_ROTAVIRUS", "D199", "A08.0", "Rótaveirusýking", "Rotavirus infection"),
    ReviewedLegacySeries("J12_1_J20_5_J21_0_RSV", "D142", "J12.1, J20.5, J21.0", "RS veirusýking", "Respiratory syncytial virus infection"),
)


def _build_annual_aliases() -> dict[str, ReviewedSeries]:
    lookup: dict[str, ReviewedSeries] = {}
    for definition in ANNUAL_SERIES:
        for alias in (definition.label_is, definition.label_en, *definition.aliases):
            key = _label_key(alias)
            prior = lookup.get(key)
            if prior is not None and prior != definition:
                raise RuntimeError(f"Conflicting annual Iceland alias: {alias!r}")
            lookup[key] = definition
    return lookup


ANNUAL_ALIAS_LOOKUP = _build_annual_aliases()
LEGACY_PAIR_LOOKUP = {
    (_normalize_icd(item.icd10), _label_key(item.label_is)): item
    for item in LEGACY_SERIES
}


class _Workbook:
    """Uniform 1-based cell access for xls and xlsx files."""

    def __init__(self, path: Path) -> None:
        self.path = path
        self.kind = path.suffix.casefold()
        if self.kind == ".xlsx":
            self.formula_book = openpyxl.load_workbook(path, data_only=False)
            self.cached_book = openpyxl.load_workbook(path, data_only=True)
            self.sheet_names = list(self.formula_book.sheetnames)
        elif self.kind == ".xls":
            self.formula_book = xlrd.open_workbook(path, on_demand=True)
            self.cached_book = self.formula_book
            self.sheet_names = list(self.formula_book.sheet_names())
        else:
            raise ValueError(f"Unsupported Iceland workbook extension: {path}")

    def dimensions(self, sheet: str) -> tuple[int, int]:
        if self.kind == ".xlsx":
            ws = self.formula_book[sheet]
            return ws.max_row, ws.max_column
        ws = self.formula_book.sheet_by_name(sheet)
        return ws.nrows, ws.ncols

    def raw(self, sheet: str, row: int, column: int) -> object:
        if self.kind == ".xlsx":
            return self.formula_book[sheet].cell(row, column).value
        ws = self.formula_book.sheet_by_name(sheet)
        if row < 1 or column < 1 or row > ws.nrows or column > ws.ncols:
            return None
        value = ws.cell_value(row - 1, column - 1)
        return None if value == "" else value

    def cached(self, sheet: str, row: int, column: int) -> object:
        if self.kind == ".xlsx":
            return self.cached_book[sheet].cell(row, column).value
        return self.raw(sheet, row, column)

    def resolved(self, sheet: str, row: int, column: int) -> tuple[object, str]:
        raw = self.raw(sheet, row, column)
        if not (isinstance(raw, str) and raw.startswith("=")):
            return raw, "reported"
        cached = self.cached(sheet, row, column)
        if cached is not None and _text(cached) != "":
            return cached, "formula_cached"
        try:
            return _evaluate_formula(self, sheet, raw), "formula_evaluated"
        except (ValueError, ZeroDivisionError, TypeError):
            return None, "formula_unresolved"

    def close(self) -> None:
        if self.kind == ".xlsx":
            self.formula_book.close()
            self.cached_book.close()
        else:
            self.formula_book.release_resources()


_CELL_RE = re.compile(r"(?<![A-Za-z0-9_])\$?([A-Z]{1,3})\$?(\d+)", re.I)
_RANGE_RE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)", re.I)


def _numeric_cell(book: _Workbook, sheet: str, column: int, row: int) -> float:
    value, _ = book.resolved(sheet, row, column)
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError("Formula references a non-numeric cell")
    return float(value)


def _safe_arithmetic(expression: str) -> float:
    tree = ast.parse(expression, mode="eval")
    allowed_binary = (ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Pow)
    allowed_unary = (ast.UAdd, ast.USub)
    for node in ast.walk(tree):
        if isinstance(node, (ast.Expression, ast.Load, ast.Constant)):
            continue
        if isinstance(node, ast.BinOp) and isinstance(node.op, allowed_binary):
            continue
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, allowed_unary):
            continue
        if isinstance(node, (*allowed_binary, *allowed_unary)):
            continue
        raise ValueError("Unsupported Excel formula expression")
    value = eval(compile(tree, "<iceland-excel-formula>", "eval"), {"__builtins__": {}}, {})
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(value):
        raise ValueError("Excel formula did not evaluate to a finite number")
    return float(value)


def _evaluate_formula(book: _Workbook, sheet: str, formula: str) -> float:
    expression = formula.lstrip("=").strip()

    def sum_range(match: re.Match[str]) -> str:
        c1, r1, c2, r2 = match.groups()
        first_col, last_col = column_index_from_string(c1), column_index_from_string(c2)
        first_row, last_row = int(r1), int(r2)
        total = 0.0
        for row in range(min(first_row, last_row), max(first_row, last_row) + 1):
            for column in range(min(first_col, last_col), max(first_col, last_col) + 1):
                total += _numeric_cell(book, sheet, column, row)
        return repr(total)

    # SUM is the only aggregate used by the official files.  IFERROR and ROUND
    # are accepted for synthetic/source revisions without implementing a full
    # spreadsheet engine.
    while re.search(r"SUM\s*\(", expression, re.I):
        updated = re.sub(
            r"SUM\s*\(\s*([A-Z]{1,3}\$?\d+\s*:\s*\$?[A-Z]{1,3}\$?\d+)\s*\)",
            lambda match: sum_range(_RANGE_RE.search(match.group(1).replace(" ", ""))),
            expression,
            flags=re.I,
        )
        if updated == expression:
            raise ValueError("Unsupported SUM formula")
        expression = updated

    if re.fullmatch(r"IFERROR\s*\(.+\)", expression, re.I):
        inner = expression[expression.find("(") + 1 : -1]
        parts = inner.rsplit(",", 1)
        if len(parts) != 2:
            raise ValueError("Unsupported IFERROR formula")
        try:
            return _evaluate_formula(book, sheet, "=" + parts[0])
        except (ValueError, ZeroDivisionError, TypeError):
            return _safe_arithmetic(parts[1])

    round_match = re.fullmatch(r"ROUND\s*\((.+),\s*(\d+)\s*\)", expression, re.I)
    if round_match:
        return float(round(_evaluate_formula(book, sheet, "=" + round_match.group(1)), int(round_match.group(2))))

    def cell_value(match: re.Match[str]) -> str:
        column, row = match.groups()
        return repr(_numeric_cell(book, sheet, column_index_from_string(column), int(row)))

    expression = _CELL_RE.sub(cell_value, expression)
    return _safe_arithmetic(expression)


def _count_value(
    value: object,
    status: str,
    *,
    dash_is_zero: bool,
) -> tuple[int | None, str]:
    text_value = _text(value)
    if not text_value:
        return None, "missing"
    if text_value in {"-", "–", "—"}:
        return (0, "dash_zero") if dash_is_zero else (None, "not_applicable")
    if isinstance(value, bool):
        return None, "invalid"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None, "invalid"
    if not math.isfinite(number) or number < 0 or not math.isclose(number, round(number), abs_tol=1e-8):
        return None, "invalid"
    integer = int(round(number))
    if status.startswith("formula_"):
        return integer, status
    return integer, "reported_zero" if integer == 0 else "reported"


def _rate_value(value: object) -> float | None:
    text_value = _text(value)
    if not text_value or text_value in {"-", "–", "—"} or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) and number >= 0 else None


def _base_row(
    definition: ReviewedSeries | ReviewedLegacySeries,
    *,
    source_kind: str,
    report_date: date,
    cases: int,
    incidence: float | None,
    value_status: str,
    raw_label_source: str,
    source_url: str,
    source_file: str,
    source_sha256: str,
    workbook_sheet: str,
    icd10: str = "",
    icd9: str = "",
) -> dict[str, Any]:
    legacy = isinstance(definition, ReviewedLegacySeries)
    if legacy:
        source_id = LEGACY_SOURCE_ID
        source_name = LEGACY_SOURCE_NAME
        source_series_code = definition.source_series_code
        local_code = definition.local_code
        frequency = "monthly"
        measure = "registered_diagnoses"
    else:
        source_id = HISTORY_SOURCE_ID
        source_name = HISTORY_SOURCE_NAME
        source_series_code = definition.source_series_code(source_kind)
        local_code = definition.local_code(source_kind)
        frequency = "annual" if source_kind == "registry_annual" else "monthly"
        measure = "case_notifications"
    return {
        "Date": report_date.isoformat(),
        "Year": str(report_date.year),
        "Month": str(report_date.month),
        "Diseases": definition.label_en,
        "RawDiseaseLabel": definition.label_is,
        "SourceRawDiseaseLabel": raw_label_source,
        "DiseaseCode": local_code,
        "DiseaseFull": definition.concept_id,
        "Cases": str(cases),
        "Incidence": "" if incidence is None else format(incidence, ".12g"),
        "Deaths": "",
        "CountryCode": "IS",
        "Frequency": frequency,
        "Measure": measure,
        "Unit": "count",
        "ValueStatus": value_status,
        "Source": source_name,
        "SourceId": source_id,
        "SourceKind": source_kind,
        "SourceSeriesCode": source_series_code,
        "SourceURL": source_url,
        "GeographyKey": "country:IS:national",
        "DatasetStatus": "historical_published_workbook",
        "AuthoritativeRevision": "true",
        "ICD10": icd10,
        "ICD9": icd9,
        "WorkbookSheet": workbook_sheet,
        "__source_file": source_file,
        "__source_sha256": source_sha256,
    }


class _OntologySeriesGate:
    """Require exact checked-in Registry coverage before rows can escape."""

    def __init__(self, ontology_path: Path) -> None:
        document = json.loads(ontology_path.read_text(encoding="utf-8"))
        self.series = {str(item.get("id")): item for item in document.get("source_series", [])}

    def validate(self, row: Mapping[str, Any]) -> str | None:
        code = _text(row.get("SourceSeriesCode"))
        definition = self.series.get(code)
        if definition is None:
            return "source_series_not_registered"
        expected = {
            "source_id": _text(row.get("SourceId")),
            "concept_id": _text(row.get("DiseaseFull")),
            "frequency": _text(row.get("Frequency")),
            "measure": _text(row.get("Measure")),
        }
        for field_name, expected_value in expected.items():
            if _text(definition.get(field_name)) != expected_value:
                return f"source_series_{field_name}_mismatch"
        if _text(definition.get("status")).casefold() in {"deprecated", "inactive"}:
            return "source_series_inactive"
        local_codes = {_text(value) for value in definition.get("local_codes", [])}
        if _text(row.get("DiseaseCode")) not in local_codes:
            return "source_series_local_code_mismatch"
        local_labels = {_label_key(value) for value in definition.get("local_labels", [])}
        if _label_key(row.get("RawDiseaseLabel")) not in local_labels:
            return "source_series_local_label_mismatch"
        return None


def _find_year_header(book: _Workbook, sheet: str, *, require_count_metric: bool) -> tuple[int, dict[int, int]]:
    max_row, max_col = book.dimensions(sheet)
    for row in range(1, min(max_row, 30) + 1):
        years = {
            column: year
            for column in range(1, max_col + 1)
            if (year := _as_year(book.raw(sheet, row, column))) is not None
        }
        if len(years) < 2:
            continue
        if require_count_metric:
            count_markers = 0
            for column in years:
                below = " ".join(
                    _label_key(book.raw(sheet, candidate, column))
                    for candidate in range(row + 1, min(row + 3, max_row) + 1)
                )
                if "fjöldi" in below or re.search(r"\bnumber\b", below):
                    count_markers += 1
            if count_markers < max(1, len(years) // 2):
                continue
        return row, years
    raise IcelandWorkbookLayoutError(f"No reviewed year header found in {book.path.name}:{sheet}")


def _parse_annual(
    book: _Workbook,
    raw_file: IcelandHistoryRawFile,
    context: _ParseContext,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet in book.sheet_names:
        max_row, _ = book.dimensions(sheet)
        header_row, year_columns = _find_year_header(book, sheet, require_count_metric=True)
        blank_run = 0
        sheet_data_started = False
        for row_number in range(header_row + 1, max_row + 1):
            label_is = _text(book.raw(sheet, row_number, 1))
            label_en = _text(book.raw(sheet, row_number, 2))
            if not label_is and not label_en:
                blank_run += 1
                if blank_run >= 3 and sheet_data_started:
                    break
                continue
            blank_run = 0
            if label_is.startswith("*") or "number of cases may" in _label_key(label_en):
                continue
            sheet_data_started = True
            definition = ANNUAL_ALIAS_LOOKUP.get(_label_key(label_is)) or ANNUAL_ALIAS_LOOKUP.get(_label_key(label_en))
            for column, year in sorted(year_columns.items(), key=lambda item: item[1]):
                value, formula_status = book.resolved(sheet, row_number, column)
                cases, value_status = _count_value(value, formula_status, dash_is_zero=False)
                if value_status == "missing":
                    context.stats["annual.blank_unknown_cells"] += 1
                    continue
                if value_status == "not_applicable":
                    context.stats["annual.not_applicable_dash_cells"] += 1
                    continue
                if cases is None:
                    context.reject(
                        "invalid_annual_value",
                        source_file=raw_file.filename,
                        sheet=sheet,
                        row=row_number,
                        column=column,
                        raw_value=_text(value),
                        raw_label=label_is or label_en,
                    )
                    continue
                if definition is None:
                    context.reject(
                        "unreviewed_annual_disease",
                        source_file=raw_file.filename,
                        sheet=sheet,
                        row=row_number,
                        year=year,
                        raw_label_is=label_is,
                        raw_label_en=label_en,
                        cases=cases,
                    )
                    continue
                rows.append(
                    _base_row(
                        definition,
                        source_kind="registry_annual",
                        report_date=date(year, 1, 1),
                        cases=cases,
                        incidence=None,
                        value_status=value_status,
                        raw_label_source=label_is or label_en,
                        source_url=raw_file.source_url,
                        source_file=raw_file.filename,
                        source_sha256=raw_file.sha256,
                        workbook_sheet=sheet,
                    )
                )
                context.stats["annual.rows_parsed"] += 1
    return rows


def _find_rate_column(
    book: _Workbook,
    sheet: str,
    header_row: int,
    count_column: int,
    next_year_column: int,
) -> int | None:
    for column in range(count_column + 1, next_year_column):
        markers = " ".join(
            _label_key(book.raw(sheet, row, column))
            for row in range(header_row + 1, header_row + 3)
        )
        if "100.000" in markers or "100,000" in markers or "/100" in markers:
            return column
    return None


def _parse_disease_monthly(
    book: _Workbook,
    raw_file: IcelandHistoryRawFile,
    context: _ParseContext,
) -> list[dict[str, Any]]:
    definition = MONTHLY_SERIES.get(raw_file.disease_key)
    if definition is None:
        raise IcelandWorkbookLayoutError(
            f"No reviewed disease identity for {raw_file.key}:{raw_file.disease_key}"
        )
    candidate_sheets = [
        sheet for sheet in book.sheet_names if "greining" in _label_key(sheet)
    ] or list(book.sheet_names)
    selected: tuple[str, int, dict[int, int]] | None = None
    for sheet in candidate_sheets:
        try:
            header_row, year_columns = _find_year_header(book, sheet, require_count_metric=True)
        except IcelandWorkbookLayoutError:
            continue
        selected = sheet, header_row, year_columns
        break
    if selected is None:
        raise IcelandWorkbookLayoutError(f"No monthly diagnosis sheet found in {book.path.name}")
    sheet, header_row, year_columns = selected
    max_row, max_col = book.dimensions(sheet)
    ordered_years = sorted(year_columns.items(), key=lambda item: item[0])
    rate_columns: dict[int, int | None] = {}
    for index, (count_column, _) in enumerate(ordered_years):
        next_column = ordered_years[index + 1][0] if index + 1 < len(ordered_years) else max_col + 1
        rate_columns[count_column] = _find_rate_column(
            book, sheet, header_row, count_column, next_column
        )

    month_rows: dict[int, int] = {}
    for row_number in range(header_row + 1, max_row + 1):
        month = _month_number(book.raw(sheet, row_number, 1))
        if month is not None and month not in month_rows:
            month_rows[month] = row_number
    if set(month_rows) != set(range(1, 13)):
        raise IcelandWorkbookLayoutError(
            f"Expected 12 month rows in {book.path.name}:{sheet}; found {sorted(month_rows)}"
        )

    rows: list[dict[str, Any]] = []
    for count_column, year in ordered_years:
        rate_column = rate_columns[count_column]
        for month, row_number in sorted(month_rows.items()):
            raw_value = book.raw(sheet, row_number, count_column)
            value, formula_status = book.resolved(sheet, row_number, count_column)
            cases, value_status = _count_value(value, formula_status, dash_is_zero=True)
            if value_status == "missing":
                context.stats["disease_monthly.blank_unknown_cells"] += 1
                continue
            if cases is None:
                context.reject(
                    "invalid_disease_monthly_value",
                    source_file=raw_file.filename,
                    sheet=sheet,
                    row=row_number,
                    column=count_column,
                    raw_value=_text(raw_value),
                    disease_key=raw_file.disease_key,
                )
                continue
            incidence = None
            if rate_column is not None:
                rate_raw, _ = book.resolved(sheet, row_number, rate_column)
                incidence = _rate_value(rate_raw)
            rows.append(
                _base_row(
                    definition,
                    source_kind="registry_disease_monthly",
                    report_date=date(year, month, 1),
                    cases=cases,
                    incidence=incidence,
                    value_status=value_status,
                    raw_label_source=definition.label_is,
                    source_url=raw_file.source_url,
                    source_file=raw_file.filename,
                    source_sha256=raw_file.sha256,
                    workbook_sheet=sheet,
                )
            )
            context.stats["disease_monthly.rows_parsed"] += 1
            if value_status == "dash_zero":
                context.stats["disease_monthly.dash_zero_cells"] += 1
            if value_status.startswith("formula_"):
                context.stats[f"disease_monthly.{value_status}"] += 1
    return rows


def _month_header(book: _Workbook, sheet: str) -> tuple[int, dict[int, int]]:
    max_row, max_col = book.dimensions(sheet)
    for row in range(1, min(max_row, 20) + 1):
        columns: dict[int, int] = {}
        for column in range(1, max_col + 1):
            month = _month_number(book.raw(sheet, row, column))
            if month is not None:
                columns[column] = month
        if len(columns) >= 10:
            return row, columns
    raise IcelandWorkbookLayoutError(f"No month header found in {book.path.name}:{sheet}")


def _explanation_codes(book: _Workbook) -> dict[str, str]:
    codes: dict[str, str] = {}
    for sheet in book.sheet_names:
        if not _label_key(sheet).startswith("ský"):
            continue
        max_row, max_col = book.dimensions(sheet)
        if max_col < 2:
            continue
        for row in range(1, max_row + 1):
            label = _text(book.raw(sheet, row, 1))
            code = _normalize_icd(book.raw(sheet, row, 2))
            if label and code and code != "ICD-10":
                codes[_label_key(label)] = code
                codes.setdefault(_explanation_key(label), code)
    return codes


def _sheet_year(book: _Workbook, sheet: str) -> int | None:
    match = re.search(r"(?:19|20)\d{2}", sheet)
    if match:
        return int(match.group(0))
    max_row, max_col = book.dimensions(sheet)
    for row in range(1, min(max_row, 5) + 1):
        for column in range(1, min(max_col, 5) + 1):
            if (year := _as_year(book.raw(sheet, row, column))) is not None:
                return year
    return None


def _parse_legacy_icd(
    book: _Workbook,
    raw_file: IcelandHistoryRawFile,
    context: _ParseContext,
) -> list[dict[str, Any]]:
    explanation = _explanation_codes(book)
    rows: list[dict[str, Any]] = []
    parsed_year_sheets = 0
    for sheet in book.sheet_names:
        year = _sheet_year(book, sheet)
        if year is None:
            continue
        try:
            header_row, month_columns = _month_header(book, sheet)
        except IcelandWorkbookLayoutError:
            continue
        parsed_year_sheets += 1
        max_row, _ = book.dimensions(sheet)
        header_values = {
            column: _label_key(book.raw(sheet, header_row, column))
            for column in range(1, min(month_columns) + 1)
        }
        icd10_column = next((column for column, value in header_values.items() if value == "icd-10"), None)
        icd9_column = next((column for column, value in header_values.items() if value == "icd-9"), None)
        first_month_column = min(month_columns)
        if icd9_column is not None:
            label_column = max(1, icd9_column - 1)
            english_column = None
        elif icd10_column is not None and icd10_column + 1 < first_month_column:
            label_column = icd10_column + 1
            english_column = None
        else:
            label_column = 1
            english_column = 2 if first_month_column >= 3 else None

        for row_number in range(header_row + 1, max_row + 1):
            label_is = _text(book.raw(sheet, row_number, label_column))
            label_en = _text(book.raw(sheet, row_number, english_column)) if english_column else ""
            if not label_is:
                continue
            explicit_icd10 = _normalize_icd(book.raw(sheet, row_number, icd10_column)) if icd10_column else ""
            icd10 = explicit_icd10 or explanation.get(_label_key(label_is), "") or explanation.get(_explanation_key(label_is), "")
            icd9 = _normalize_icd(book.raw(sheet, row_number, icd9_column)) if icd9_column else ""
            definition = LEGACY_PAIR_LOOKUP.get((_normalize_icd(icd10), _label_key(label_is)))
            any_observation = False
            for column, month in sorted(month_columns.items(), key=lambda item: item[1]):
                raw_value = book.raw(sheet, row_number, column)
                value, formula_status = book.resolved(sheet, row_number, column)
                cases, value_status = _count_value(value, formula_status, dash_is_zero=True)
                if value_status == "missing":
                    context.stats["legacy_icd.blank_unknown_cells"] += 1
                    continue
                any_observation = True
                if cases is None:
                    context.reject(
                        "invalid_legacy_icd_value",
                        source_file=raw_file.filename,
                        sheet=sheet,
                        row=row_number,
                        column=column,
                        raw_value=_text(raw_value),
                        raw_label_is=label_is,
                        icd10=icd10,
                    )
                    continue
                if definition is None:
                    context.reject(
                        "unreviewed_legacy_icd_series",
                        source_file=raw_file.filename,
                        sheet=sheet,
                        row=row_number,
                        year=year,
                        month=month,
                        raw_label_is=label_is,
                        raw_label_en=label_en,
                        icd10=icd10,
                        icd9=icd9,
                        cases=cases,
                    )
                    continue
                rows.append(
                    _base_row(
                        definition,
                        source_kind="legacy_icd_monthly",
                        report_date=date(year, month, 1),
                        cases=cases,
                        incidence=None,
                        value_status=value_status,
                        raw_label_source=label_is,
                        source_url=raw_file.source_url,
                        source_file=raw_file.filename,
                        source_sha256=raw_file.sha256,
                        workbook_sheet=sheet,
                        icd10=icd10,
                        icd9=icd9,
                    )
                )
                context.stats["legacy_icd.rows_parsed"] += 1
                if value_status == "dash_zero":
                    context.stats["legacy_icd.dash_zero_cells"] += 1
            if any_observation and not icd10:
                context.stats["legacy_icd.rows_without_icd10"] += 1
    if parsed_year_sheets == 0:
        raise IcelandWorkbookLayoutError(f"No legacy ICD year sheets parsed in {book.path.name}")
    return rows


def _raw_file_from_manifest(
    item: Mapping[str, Any], manifest_path: Path
) -> IcelandHistoryRawFile:
    raw_path = Path(_text(item.get("path")))
    filename = _text(item.get("filename")) or raw_path.name
    # v1 manifests originally recorded machine-specific absolute paths.  A
    # relocated archive must still be replayable, so a missing old path falls
    # back to the workbook beside the manifest.  New manifests use relative
    # paths and take the same deterministic resolution route.
    candidates: list[Path] = []
    if raw_path.is_absolute():
        candidates.append(raw_path)
    elif _text(item.get("path")):
        candidates.append(manifest_path.parent / raw_path)
    if filename:
        candidates.append(manifest_path.parent / filename)
    if not candidates:
        candidates.append(manifest_path.parent / filename)
    path = next(
        (candidate for candidate in candidates if candidate.exists()), candidates[0]
    )
    size_bytes = int(
        item.get("size_bytes") or (path.stat().st_size if path.exists() else 0)
    )
    return IcelandHistoryRawFile(
        key=_text(item.get("key")),
        source_kind=_text(item.get("source_kind")),
        filename=filename or path.name,
        path=str(path.resolve()),
        source_url=_text(item.get("source_url")),
        sha256=_text(item.get("sha256")),
        size_bytes=size_bytes,
        media_type=_text(item.get("media_type")),
        disease_key=_text(item.get("disease_key")),
        validation_only=bool(item.get("validation_only", False)),
    )


def _quarantine_summary(quarantine: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    by_reason = Counter(_text(item.get("reason")) for item in quarantine)
    identities: dict[tuple[str, str, str, str], int] = Counter()
    for item in quarantine:
        identity = (
            _text(item.get("reason")),
            _text(item.get("raw_label_is") or item.get("RawDiseaseLabel")),
            _text(item.get("raw_label_en")),
            _text(item.get("icd10") or item.get("ICD10")),
        )
        identities[identity] += 1
    unmapped = [
        {
            "reason": reason,
            "raw_label_is": label_is,
            "raw_label_en": label_en,
            "icd10": icd10,
            "row_count": count,
        }
        for (reason, label_is, label_en, icd10), count in sorted(identities.items())
    ]
    return {
        "row_count": len(quarantine),
        "by_reason": dict(sorted(by_reason.items())),
        "identities": unmapped,
    }


class IcelandHistoryProcessor:
    """Prepare auditable Iceland history rows and import the safe projection."""

    country_code = "IS"
    series_registered_rows_only = True

    def __init__(
        self,
        *,
        ontology_path: Path | str = DEFAULT_ONTOLOGY_PATH,
        require_registered_series: bool = True,
    ) -> None:
        self.ontology_path = Path(ontology_path)
        self.require_registered_series = require_registered_series

    def prepare_manifest(self, manifest_path: Path | str) -> IcelandHistoryPreparedResult:
        path = Path(manifest_path)
        document = json.loads(path.read_text(encoding="utf-8"))
        files = [_raw_file_from_manifest(item, path) for item in document.get("files", [])]
        if not files:
            raise ValueError(f"Iceland raw manifest contains no files: {path}")
        return self.prepare_files(files, raw_manifest=document)

    def prepare_files(
        self,
        raw_files: Iterable[IcelandHistoryRawFile],
        *,
        raw_manifest: Mapping[str, Any] | None = None,
    ) -> IcelandHistoryPreparedResult:
        files = list(raw_files)
        context = _ParseContext()
        candidates: list[dict[str, Any]] = []
        raw_hashes: dict[str, str] = {}
        file_manifest: list[dict[str, Any]] = []

        for raw_file in files:
            path = Path(raw_file.path)
            if not path.exists():
                raise FileNotFoundError(f"Iceland raw workbook not found: {path}")
            digest = _sha256_path(path)
            if raw_file.sha256 and raw_file.sha256 != digest:
                raise ValueError(
                    f"Iceland raw hash mismatch for {raw_file.filename}: "
                    f"manifest={raw_file.sha256} actual={digest}"
                )
            raw_hashes[raw_file.filename] = digest
            file_manifest.append(
                {
                    **asdict(raw_file),
                    "path": str(path.resolve()),
                    "sha256": digest,
                }
            )
            if raw_file.validation_only or raw_file.source_kind == "validation_annual":
                context.stats["files.validation_only_skipped"] += 1
                continue
            book = _Workbook(path)
            try:
                if raw_file.source_kind == "registry_annual":
                    candidates.extend(_parse_annual(book, raw_file, context))
                elif raw_file.source_kind == "registry_disease_monthly":
                    candidates.extend(_parse_disease_monthly(book, raw_file, context))
                elif raw_file.source_kind == "legacy_icd_monthly":
                    candidates.extend(_parse_legacy_icd(book, raw_file, context))
                else:
                    raise ValueError(
                        f"Unsupported Iceland source kind {raw_file.source_kind!r}"
                    )
            finally:
                book.close()
            context.stats[f"files.{raw_file.source_kind}"] += 1

        gate = _OntologySeriesGate(self.ontology_path) if self.require_registered_series else None
        series_rows: list[dict[str, Any]] = []
        validation_cache: dict[tuple[str, str, str], str | None] = {}
        for row in candidates:
            cache_key = (
                _text(row.get("SourceSeriesCode")),
                _text(row.get("DiseaseCode")),
                _text(row.get("RawDiseaseLabel")),
            )
            reason = validation_cache.get(cache_key)
            if cache_key not in validation_cache:
                reason = gate.validate(row) if gate is not None else None
                validation_cache[cache_key] = reason
            if reason:
                context.reject(reason, **row)
                continue
            series_rows.append(row)

        # The legacy disease_records projection cannot represent frequency or
        # reporting basis.  Prefer disease-specific monthly notification rows;
        # retain annual rows only for disease-years absent from those workbooks.
        monthly_years = {
            (_text(row.get("DiseaseFull")), int(row["Year"]))
            for row in series_rows
            if row.get("SourceKind") == "registry_disease_monthly"
        }
        projection_candidates = [
            row
            for row in series_rows
            if row.get("SourceKind") == "registry_disease_monthly"
            or (
                row.get("SourceKind") == "registry_annual"
                and (_text(row.get("DiseaseFull")), int(row["Year"])) not in monthly_years
            )
        ]
        context.stats["projection.annual_rows_superseded_by_monthly"] = sum(
            1
            for row in series_rows
            if row.get("SourceKind") == "registry_annual"
            and (_text(row.get("DiseaseFull")), int(row["Year"])) in monthly_years
        )
        if gate is not None:
            semantically_projectable: list[dict[str, Any]] = []
            unsafe_mapping_rows = 0
            for row in projection_candidates:
                definition = gate.series[_text(row.get("SourceSeriesCode"))]
                if is_case_count_series(
                    {
                        "metric_type": definition.get("measure"),
                        "series_unit": definition.get("unit", "count"),
                        "observation_unit": row.get("Unit"),
                        "mapping_relation": definition.get("mapping_relation"),
                    }
                ):
                    semantically_projectable.append(row)
                else:
                    unsafe_mapping_rows += 1
            context.stats["projection.unsafe_mapping_rows_excluded"] = (
                unsafe_mapping_rows
            )

            definitions_by_concept: dict[str, dict[str, Mapping[str, Any]]] = (
                defaultdict(dict)
            )
            for row in series_rows:
                series_code = _text(row.get("SourceSeriesCode"))
                definition = gate.series[series_code]
                if is_case_count_series(
                    {
                        "metric_type": definition.get("measure"),
                        "series_unit": definition.get("unit", "count"),
                        "observation_unit": row.get("Unit"),
                        "mapping_relation": definition.get("mapping_relation"),
                    }
                ):
                    definitions_by_concept[_text(row.get("DiseaseFull"))][
                        series_code
                    ] = definition
            source_only_concepts: set[str] = set()
            for concept_id, definitions in definitions_by_concept.items():
                if len(definitions) < 2:
                    continue
                selection = select_series_projection(
                    [
                        {
                            "series_code": series_code,
                            "mapping_relation": definition.get("mapping_relation"),
                            "aggregation_policy": definition.get(
                                "aggregation_policy", "non_additive"
                            ),
                        }
                        for series_code, definition in definitions.items()
                    ]
                )
                if selection.projection_policy == SOURCE_OBSERVATIONS_ONLY_POLICY:
                    source_only_concepts.add(concept_id)
            before_narrower_filter = len(semantically_projectable)
            projection_candidates = [
                row
                for row in semantically_projectable
                if _text(row.get("DiseaseFull")) not in source_only_concepts
            ]
            context.stats["projection.narrower_multi_series_rows_excluded"] = (
                before_narrower_filter - len(projection_candidates)
            )
        # ``disease_records`` has no source-series key. Distinct reviewed
        # categories can intentionally map to one ontology concept (for
        # example historical Hib and all-invasive H. influenzae series), so
        # projecting either one would silently overwrite the other. Preserve
        # every fact in ``series_rows`` and omit all non-representable legacy
        # identities instead of choosing or aggregating a category.
        projection_identity_counts = Counter(
            (
                _text(row.get("DiseaseFull")),
                _text(row.get("Date")),
            )
            for row in projection_candidates
        )
        rows = [
            row
            for row in projection_candidates
            if projection_identity_counts[
                (_text(row.get("DiseaseFull")), _text(row.get("Date")))
            ]
            == 1
        ]
        context.stats["projection.multi_series_identity_rows_excluded"] = (
            len(projection_candidates) - len(rows)
        )

        sort_key = lambda row: (
            _text(row.get("Date")),
            _text(row.get("SourceSeriesCode")),
            _text(row.get("DiseaseCode")),
        )
        rows.sort(key=sort_key)
        series_rows.sort(key=sort_key)
        manifest = {
            "schema": PARSED_MANIFEST_SCHEMA,
            "parser_version": PARSER_VERSION,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "country_code": "IS",
            "source_families": {
                "registry_annual": {
                    "source_id": HISTORY_SOURCE_ID,
                    "frequency": "annual",
                    "measure": "case_notifications",
                    "dash_semantics": "not_applicable_skip",
                    "blank_semantics": "unknown_skip",
                },
                "registry_disease_monthly": {
                    "source_id": HISTORY_SOURCE_ID,
                    "frequency": "monthly",
                    "measure": "case_notifications",
                    "dash_semantics": "published_zero",
                    "blank_semantics": "unknown_skip",
                },
                "legacy_icd_monthly": {
                    "source_id": LEGACY_SOURCE_ID,
                    "frequency": "monthly",
                    "measure": "registered_diagnoses",
                    "dash_semantics": "published_zero",
                    "blank_semantics": "unknown_skip",
                    "legacy_projection": "excluded_reporting_basis_not_comparable",
                },
            },
            "raw_manifest_schema": _text((raw_manifest or {}).get("schema")),
            "raw_landing_sha256": _text((raw_manifest or {}).get("landing_sha256")),
            "files": file_manifest,
            "raw_hashes": dict(sorted(raw_hashes.items())),
            "counts": {
                "projection_rows": len(rows),
                "series_rows": len(series_rows),
                "quarantine_rows": len(context.quarantine),
                **dict(sorted(context.stats.items())),
            },
            "series_counts": dict(
                sorted(Counter(_text(row.get("SourceSeriesCode")) for row in series_rows).items())
            ),
            "quarantine": _quarantine_summary(context.quarantine),
        }
        return IcelandHistoryPreparedResult(
            rows=rows,
            series_rows=series_rows,
            quarantine=context.quarantine,
            manifest=manifest,
            raw_hashes=raw_hashes,
        )

    @staticmethod
    def write_outputs(
        result: IcelandHistoryPreparedResult,
        output_dir: Path | str = DEFAULT_OUTPUT_DIR,
    ) -> dict[str, Path]:
        target = Path(output_dir)
        target.mkdir(parents=True, exist_ok=True)
        outputs: dict[str, Path] = {}
        for filename, rows in (
            ("rows.csv", result.rows),
            ("series_rows.csv", result.series_rows),
            ("quarantine.csv", result.quarantine),
        ):
            path = target / filename
            fieldnames = sorted({key for row in rows for key in row})
            with path.open("w", encoding="utf-8-sig", newline="") as handle:
                if fieldnames:
                    writer = csv.DictWriter(
                        handle, fieldnames=fieldnames, extrasaction="ignore"
                    )
                    writer.writeheader()
                    writer.writerows(rows)
            outputs[filename] = path.resolve()
        manifest_path = target / "manifest.json"
        # ``prepare_files`` resolves workbooks to absolute paths while it is
        # parsing them.  Those paths are useful in-memory, but persisting them
        # would bind the reviewed artifact to one checkout and disclose the
        # local machine layout.  Store paths relative to the parsed manifest so
        # the raw archive and normalized audit bundle remain movable together.
        manifest = {**result.manifest}
        manifest_files: list[dict[str, Any]] = []
        for raw_entry in result.manifest.get("files", []):
            entry = dict(raw_entry)
            raw_path = Path(_text(entry.get("path")))
            if raw_path:
                entry["path"] = Path(
                    os.path.relpath(
                        raw_path.resolve(),
                        start=manifest_path.parent.resolve(),
                    )
                ).as_posix()
            manifest_files.append(entry)
        manifest["files"] = manifest_files
        manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        outputs["manifest.json"] = manifest_path.resolve()
        return outputs

    async def import_rows(
        self,
        db: AsyncSession,
        rows: list[dict[str, Any]],
        *,
        db_latest_date: date | None = None,
        source_latest_date: date | None = None,
        force: bool = False,
    ) -> IcelandHistoryImportResult:
        """Upsert the safe annual/monthly case-notification projection.

        ``legacy_icd_monthly`` is rejected even if a caller bypasses
        :meth:`prepare_files`, because ``disease_records`` cannot encode its
        registered-diagnosis reporting basis.
        """

        del db_latest_date, force
        safe_rows = [
            row
            for row in rows
            if _text(row.get("SourceKind")) in {
                "registry_annual",
                "registry_disease_monthly",
            }
            and _text(row.get("Measure")) == "case_notifications"
        ]
        if not safe_rows:
            return IcelandHistoryImportResult(0, 0, source_latest_date)

        country_result = await db.execute(
            text("SELECT id FROM countries WHERE code = 'IS'")
        )
        country_row = country_result.fetchone()
        if country_row is None:
            raise ValueError("Country not found in database: IS")
        country_id = int(country_row[0])
        concept_ids = sorted({_text(row.get("DiseaseFull")) for row in safe_rows})
        disease_result = await db.execute(
            text(
                "SELECT name, id FROM diseases "
                "WHERE name = ANY(:concept_ids) AND is_active = true"
            ),
            {"concept_ids": concept_ids},
        )
        disease_ids = {str(row[0]): int(row[1]) for row in disease_result.fetchall()}
        current_result = await db.execute(
            text(
                """
                SELECT timezone('UTC', time)::date, disease_id
                FROM disease_records
                WHERE country_id = :country_id
                  AND COALESCE(metadata::jsonb ->> 'legacy_projection', '') =
                      'current_annual_dashboard_only'
                """
            ),
            {"country_id": country_id},
        )
        current_identities = {
            (row[0], int(row[1])) for row in current_result.fetchall()
        }

        upserts: list[dict[str, Any]] = []
        skipped = 0
        skipped_current_precedence = 0
        seen: set[tuple[str, int]] = set()
        for row in safe_rows:
            concept_id = _text(row.get("DiseaseFull"))
            disease_id = disease_ids.get(concept_id)
            if disease_id is None:
                skipped += 1
                continue
            report_date = datetime.strptime(_text(row.get("Date")), "%Y-%m-%d").replace(
                tzinfo=timezone.utc
            )
            if (report_date.date(), disease_id) in current_identities:
                # The current dashboard is authoritative for overlapping
                # annual identities (for example D236/MRSA in 2019).  History
                # remains losslessly available in the source-series table.
                skipped_current_precedence += 1
                continue
            identity = (report_date.isoformat(), disease_id)
            if identity in seen:
                raise ValueError(
                    "Iceland projection contains duplicate disease_records identity: "
                    f"{identity}"
                )
            seen.add(identity)
            upserts.append(
                {
                    "time": report_date,
                    "disease_id": disease_id,
                    "country_id": country_id,
                    "cases": int(row["Cases"]),
                    "data_source": _text(row.get("Source")),
                    "metadata": json.dumps(
                        {
                            "source_kind": row.get("SourceKind"),
                            "source_series_code": row.get("SourceSeriesCode"),
                            "raw_disease_label": row.get("SourceRawDiseaseLabel"),
                            "incidence": row.get("Incidence"),
                            "source_url": row.get("SourceURL"),
                            "source_file": row.get("__source_file"),
                            "source_sha256": row.get("__source_sha256"),
                        },
                        ensure_ascii=False,
                    ),
                    "raw_data": json.dumps(row, ensure_ascii=False),
                }
            )
        if upserts:
            await db.execute(
                text(
                    """
                    INSERT INTO disease_records (
                        time, disease_id, country_id, cases, deaths, data_source,
                        metadata, raw_data, new_cases, new_deaths, recoveries,
                        active_cases, new_recoveries
                    ) VALUES (
                        :time, :disease_id, :country_id, :cases, NULL, :data_source,
                        :metadata, :raw_data, 0, 0, 0, 0, 0
                    )
                    ON CONFLICT (time, disease_id, country_id) DO UPDATE SET
                        cases = EXCLUDED.cases,
                        deaths = EXCLUDED.deaths,
                        data_source = EXCLUDED.data_source,
                        metadata = EXCLUDED.metadata,
                        raw_data = EXCLUDED.raw_data
                    WHERE COALESCE(
                        disease_records.metadata::jsonb ->> 'legacy_projection',
                        ''
                    ) <> 'current_annual_dashboard_only'
                    """
                ),
                upserts,
            )
        latest = source_latest_date or max(
            (datetime.strptime(_text(row["Date"]), "%Y-%m-%d").date() for row in safe_rows),
            default=None,
        )
        return IcelandHistoryImportResult(
            len(upserts),
            skipped,
            latest,
            skipped_current_precedence=skipped_current_precedence,
        )


def raw_files_for_paths(paths: Iterable[Path | str]) -> list[IcelandHistoryRawFile]:
    """Build verified raw-file descriptors for already-downloaded catalogue files."""

    specs_by_filename = {spec.filename.casefold(): spec for spec in OFFICIAL_WORKBOOKS}
    raw_files: list[IcelandHistoryRawFile] = []
    for path_value in paths:
        path = Path(path_value).resolve()
        spec = specs_by_filename.get(path.name.casefold())
        if spec is None:
            raise ValueError(f"Path is not a canonical Iceland catalogue filename: {path.name}")
        raw_files.append(
            IcelandHistoryRawFile(
                key=spec.key,
                source_kind=spec.source_kind,
                filename=spec.filename,
                path=str(path),
                source_url=spec.url,
                sha256=_sha256_path(path),
                size_bytes=path.stat().st_size,
                media_type="application/vnd.ms-excel" if path.suffix.casefold() == ".xls" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                disease_key=spec.disease_key,
                validation_only=spec.validation_only,
            )
        )
    return raw_files


__all__ = [
    "ANNUAL_SERIES",
    "DEFAULT_ONTOLOGY_PATH",
    "DEFAULT_OUTPUT_DIR",
    "HISTORY_SOURCE_ID",
    "HISTORY_SOURCE_NAME",
    "IcelandHistoryImportResult",
    "IcelandHistoryPreparedResult",
    "IcelandHistoryProcessor",
    "IcelandWorkbookLayoutError",
    "LEGACY_SERIES",
    "LEGACY_SOURCE_ID",
    "LEGACY_SOURCE_NAME",
    "MONTHLY_SERIES",
    "PARSER_VERSION",
    "raw_files_for_paths",
]
