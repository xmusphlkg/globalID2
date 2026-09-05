"""Generate deterministic, time-partitioned public download files."""

from __future__ import annotations

import csv
from concurrent.futures import FIRST_COMPLETED, Future, ProcessPoolExecutor, wait
from dataclasses import dataclass
from datetime import date, datetime
import hashlib
from io import BytesIO, StringIO
import json
import multiprocessing
import os
from pathlib import Path
import shutil
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook

from src.services.disease_series_policy import is_case_count_series


DOWNLOAD_FIELDS = [
    "dataset_kind",
    "dataset_id",
    "dataset_slug",
    "dataset_name",
    "record_kind",
    "country_code",
    "country_name",
    "disease_id",
    "disease_name_en",
    "disease_name_zh",
    "category",
    "date",
    "year_month",
    "value",
    "cases",
    "weekly_equiv_cases",
    "deaths",
    "incidence_rate_per_100k",
    "incidence_rate_source",
    "mortality_rate",
    "series_code",
    "source_series_code",
    "source_system",
    "source_label",
    "metric_type",
    "reporting_basis",
    "temporal_granularity",
    "unit",
    "geography_key",
    "dimension_key",
    "mapping_relation",
    "comparability",
    "aggregation_policy",
    "availability_status",
    "missing_value_policy",
    "definition_version",
    "case_definition",
    "case_definition_uri",
    "quality_status",
    "is_selected_series",
    "data_layer",
    "projection_policy",
    "coverage_status",
    "loss_risk",
    "selected_series_codes",
    "available_series_count",
    "provenance_note",
    "primary_source_scope",
    "primary_source_label",
    "primary_source_url",
    "primary_source_type",
    "source_scopes",
    "source_labels",
    "source_urls",
    "source_types",
]

SOURCE_SERIES_FIELDS = (
    "source_series_code",
    "source_system",
    "source_label",
    "metric_type",
    "reporting_basis",
    "temporal_granularity",
    "unit",
    "geography_key",
    "dimension_key",
    "mapping_relation",
    "comparability",
    "aggregation_policy",
    "availability_status",
    "missing_value_policy",
    "definition_version",
    "case_definition",
    "case_definition_uri",
)

PUBLIC_FORMATS = ("csv", "json", "xlsx")
GITHUB_MAX_FILE_BYTES = 100 * 1024 * 1024
DEFAULT_TARGET_FILE_BYTES = 90 * 1024 * 1024
SPECIAL_WINDOW_START = 2020
SPECIAL_WINDOW_END = 2025
BRIDGE_WINDOW_START = 2026
BRIDGE_WINDOW_END = 2029
ROLLING_WINDOW_START = 2030
ROLLING_WINDOW_YEARS = 5
_XLSX_TIMESTAMP = (2000, 1, 1, 0, 0, 0)


def _read_cgroup_memory_limit_bytes() -> int | None:
    """Return the tightest cgroup memory guardrail when the exporter is service-bound."""
    candidates = (
        Path("/sys/fs/cgroup/memory.high"),
        Path("/sys/fs/cgroup/memory.max"),
    )
    values: list[int] = []
    for path in candidates:
        try:
            raw = path.read_text(encoding="utf-8").strip()
        except OSError:
            continue
        if not raw or raw == "max":
            continue
        try:
            values.append(int(raw))
        except ValueError:
            continue
    return min(values) if values else None


def _default_export_workers(cpu_count: int | None, memory_limit_bytes: int | None) -> int:
    cpu_workers = min(4, max(1, cpu_count or 1))
    if not memory_limit_bytes:
        return cpu_workers

    gib = 1024 * 1024 * 1024
    if memory_limit_bytes <= 4 * gib:
        return 1
    if memory_limit_bytes <= 8 * gib:
        return min(cpu_workers, 2)
    return cpu_workers


DEFAULT_EXPORT_WORKERS = _default_export_workers(os.cpu_count(), _read_cgroup_memory_limit_bytes())


def _process_pool_context() -> multiprocessing.context.BaseContext:
    try:
        return multiprocessing.get_context("forkserver")
    except ValueError:
        return multiprocessing.get_context()


@dataclass(frozen=True)
class PartitionSpec:
    """One stable calendar window used as a public file name."""

    key: str
    start: date
    end: date
    label: str
    is_current: bool = False


def _download_url(base_url: str, relative_path: str) -> str:
    return f"{base_url.rstrip('/')}/{relative_path.lstrip('/')}"


def _sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _file_sha256(path: Path) -> str:
    """Hash an existing artifact without loading large downloads into memory."""

    digest = hashlib.sha256()
    with path.open("rb") as source:
        while chunk := source.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _source_columns(source_info: dict) -> dict:
    sources = source_info.get("sources") or []
    return {
        "primary_source_scope": source_info.get("primary_scope"),
        "primary_source_label": source_info.get("primary_label"),
        "primary_source_url": source_info.get("primary_url"),
        "primary_source_type": source_info.get("primary_type"),
        "source_scopes": "; ".join(
            source.get("scope") or "" for source in sources if source.get("scope")
        ),
        "source_labels": "; ".join(
            source.get("label") or "" for source in sources if source.get("label")
        ),
        "source_urls": "; ".join(
            source.get("url") or "" for source in sources if source.get("url")
        ),
        "source_types": "; ".join(
            source.get("type") or "" for source in sources if source.get("type")
        ),
    }


def _projection_provenance_columns(
    series: dict,
    source_series: dict | None = None,
) -> dict:
    selected_codes = [
        str(code) for code in series.get("selected_series_codes") or []
    ]
    data_provenance = series.get("data_provenance") or {}
    columns = {
        "series_code": (
            source_series.get("series_code")
            if source_series
            else (selected_codes[0] if len(selected_codes) == 1 else None)
        ),
        "is_selected_series": (
            source_series.get("series_code") in selected_codes
            if source_series else None
        ),
        "data_layer": series.get("data_layer"),
        "projection_policy": series.get("projection_policy"),
        "coverage_status": series.get("coverage_status"),
        "loss_risk": series.get("loss_risk"),
        "selected_series_codes": "; ".join(selected_codes),
        "available_series_count": series.get("available_series_count") or 0,
        "provenance_note": data_provenance.get("note_en"),
    }
    if source_series:
        columns.update(
            {field: source_series.get(field) for field in SOURCE_SERIES_FIELDS}
        )
        quality_statuses = source_series.get("quality_statuses") or []
        columns["quality_status"] = "; ".join(
            str(value) for value in quality_statuses if value
        )
    else:
        columns["temporal_granularity"] = series.get("period_granularity")
    return columns


def _source_observation_rows(
    series: dict,
    base_columns: dict,
    source_columns: dict,
) -> list[dict]:
    rows: list[dict] = []
    for source_series in series.get("source_series") or []:
        dates = source_series.get("dates") or []
        values = source_series.get("values") or []
        granularity = str(source_series.get("temporal_granularity") or "").lower()
        for index, value_date in enumerate(dates):
            value = values[index] if index < len(values) else None
            case_value = value if is_case_count_series(source_series) else None
            rows.append(
                {
                    **base_columns,
                    "record_kind": "source_series_observation",
                    "date": value_date,
                    "year_month": value_date[:7] if value_date else None,
                    "value": value,
                    "cases": case_value,
                    "weekly_equiv_cases": (
                        float(value)
                        if granularity == "weekly" and case_value is not None
                        else None
                    ),
                    "deaths": None,
                    "incidence_rate_per_100k": None,
                    "incidence_rate_source": None,
                    "mortality_rate": None,
                    **_projection_provenance_columns(series, source_series),
                    **source_columns,
                }
            )
    return rows


def build_country_download_rows(country_data: dict, source_info: dict) -> list[dict]:
    """Flatten a country projection without release-volatile fields."""

    rows: list[dict] = []
    source_columns = _source_columns(source_info)
    dataset_id = str(country_data.get("country_code") or "").lower()
    for series in (country_data.get("disease_series") or {}).values():
        base_columns = {
            "dataset_kind": "country",
            "dataset_id": dataset_id,
            "dataset_slug": dataset_id,
            "dataset_name": country_data.get("country_name"),
            "country_code": country_data.get("country_code"),
            "country_name": country_data.get("country_name"),
            "disease_id": series.get("disease_id"),
            "disease_name_en": series.get("name_en"),
            "disease_name_zh": series.get("name_zh"),
            "category": series.get("category"),
        }
        rows.extend(
            _source_observation_rows(series, base_columns, source_columns)
        )
        dates = series.get("dates") or []
        cases = series.get("cases") or []
        weekly_equiv = series.get("weekly_equiv_cases") or []
        deaths = series.get("deaths") or []
        incidence_rates = series.get("incidence_rates") or []
        incidence_sources = series.get("incidence_sources") or []
        mortality_rates = series.get("mortality_rates") or []
        for index, value_date in enumerate(dates):
            rows.append(
                {
                    **base_columns,
                    "record_kind": "public_projection",
                    "date": value_date,
                    "year_month": value_date[:7] if value_date else None,
                    "cases": cases[index] if index < len(cases) else 0,
                    "value": cases[index] if index < len(cases) else 0,
                    "weekly_equiv_cases": (
                        weekly_equiv[index] if index < len(weekly_equiv) else None
                    ),
                    "deaths": deaths[index] if index < len(deaths) else 0,
                    "incidence_rate_per_100k": (
                        incidence_rates[index] if index < len(incidence_rates) else None
                    ),
                    "incidence_rate_source": (
                        incidence_sources[index] if index < len(incidence_sources) else None
                    ),
                    "mortality_rate": (
                        mortality_rates[index] if index < len(mortality_rates) else None
                    ),
                    **_projection_provenance_columns(series),
                    **source_columns,
                }
            )
    return _sort_rows(rows)


def build_disease_download_rows(
    disease_data: dict,
    source_info_by_country: dict[str, dict],
    country_name_by_code: dict[str, str],
) -> list[dict]:
    """Flatten a disease projection without release-volatile fields."""

    rows: list[dict] = []
    dataset_id = disease_data.get("disease_id")
    for country_code, series in (disease_data.get("country_series") or {}).items():
        dates = series.get("dates") or []
        cases = series.get("cases") or []
        weekly_equiv = series.get("weekly_equiv_cases") or []
        deaths = series.get("deaths") or []
        incidence_rates = series.get("incidence_rates") or []
        incidence_sources = series.get("incidence_sources") or []
        source_columns = _source_columns(
            source_info_by_country.get(country_code, {"sources": []})
        )
        base_columns = {
            "dataset_kind": "disease",
            "dataset_id": dataset_id,
            "dataset_slug": disease_data.get("slug"),
            "dataset_name": disease_data.get("name_en"),
            "country_code": country_code,
            "country_name": country_name_by_code.get(country_code),
            "disease_id": dataset_id,
            "disease_name_en": disease_data.get("name_en"),
            "disease_name_zh": disease_data.get("name_zh"),
            "category": disease_data.get("category"),
        }
        rows.extend(
            _source_observation_rows(series, base_columns, source_columns)
        )
        for index, value_date in enumerate(dates):
            rows.append(
                {
                    **base_columns,
                    "record_kind": "public_projection",
                    "date": value_date,
                    "year_month": value_date[:7] if value_date else None,
                    "cases": cases[index] if index < len(cases) else 0,
                    "value": cases[index] if index < len(cases) else 0,
                    "weekly_equiv_cases": (
                        weekly_equiv[index] if index < len(weekly_equiv) else None
                    ),
                    "deaths": deaths[index] if index < len(deaths) else 0,
                    "incidence_rate_per_100k": (
                        incidence_rates[index] if index < len(incidence_rates) else None
                    ),
                    "incidence_rate_source": (
                        incidence_sources[index] if index < len(incidence_sources) else None
                    ),
                    "mortality_rate": None,
                    **_projection_provenance_columns(series),
                    **source_columns,
                }
            )
    return _sort_rows(rows)


def _sort_rows(rows: list[dict]) -> list[dict]:
    return sorted(
        rows,
        key=lambda row: (
            str(row.get("date") or ""),
            str(row.get("country_code") or ""),
            str(row.get("disease_id") or ""),
            str(row.get("record_kind") or ""),
            str(row.get("series_code") or ""),
        ),
    )


def _window_for_year(year: int, latest_year: int) -> PartitionSpec:
    if year < SPECIAL_WINDOW_START:
        start_year = year - (year % 5)
        end_year = start_year + 4
    elif year <= SPECIAL_WINDOW_END:
        start_year = SPECIAL_WINDOW_START
        end_year = SPECIAL_WINDOW_END
    elif year <= BRIDGE_WINDOW_END:
        start_year = BRIDGE_WINDOW_START
        end_year = BRIDGE_WINDOW_END
    else:
        start_year = year - (year % ROLLING_WINDOW_YEARS)
        end_year = start_year + ROLLING_WINDOW_YEARS - 1

    current = start_year <= latest_year <= end_year
    label = f"{start_year}–now" if current else f"{start_year}–{end_year}"
    return PartitionSpec(
        key=f"{start_year}-{end_year}",
        start=date(start_year, 1, 1),
        end=date(end_year, 12, 31),
        label=label,
        is_current=current,
    )


def partition_rows(rows: list[dict]) -> list[tuple[PartitionSpec, list[dict]]]:
    """Group rows into stable historical windows and one current window."""

    valid_rows = [row for row in rows if str(row.get("date") or "")[:4].isdigit()]
    if not valid_rows:
        return []
    latest_year = max(int(str(row["date"])[:4]) for row in valid_rows)
    grouped: dict[str, tuple[PartitionSpec, list[dict]]] = {}
    for row in valid_rows:
        spec = _window_for_year(int(str(row["date"])[:4]), latest_year)
        grouped.setdefault(spec.key, (spec, []))[1].append(row)
    return sorted(grouped.values(), key=lambda item: item[0].start, reverse=True)


def _csv_bytes(rows: list[dict]) -> bytes:
    stream = StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=DOWNLOAD_FIELDS, lineterminator="\n")
    writer.writeheader()
    writer.writerows(
        {field: row.get(field) for field in DOWNLOAD_FIELDS} for row in rows
    )
    return stream.getvalue().encode("utf-8")


def _json_bytes(metadata: dict, rows: list[dict]) -> bytes:
    return json.dumps(
        {"metadata": metadata, "records": rows},
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")


def _xlsx_cell(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return value


def _xlsx_bytes(metadata: dict, rows: list[dict]) -> bytes:
    workbook = Workbook(write_only=True)
    fixed = datetime(*_XLSX_TIMESTAMP)
    workbook.properties.created = fixed
    workbook.properties.modified = fixed
    records_sheet = workbook.create_sheet("records")
    records_sheet.append(DOWNLOAD_FIELDS)
    for row in rows:
        records_sheet.append([_xlsx_cell(row.get(field)) for field in DOWNLOAD_FIELDS])
    metadata_sheet = workbook.create_sheet("metadata")
    metadata_sheet.append(["field", "value"])
    for key in sorted(metadata):
        metadata_sheet.append([key, _xlsx_cell(metadata[key])])

    original = BytesIO()
    workbook.save(original)
    original.seek(0)
    canonical = BytesIO()
    with ZipFile(original, "r") as source_zip, ZipFile(
        canonical, "w", compression=ZIP_DEFLATED, compresslevel=9
    ) as target_zip:
        for name in sorted(source_zip.namelist()):
            source_info = source_zip.getinfo(name)
            target_info = ZipInfo(name, date_time=_XLSX_TIMESTAMP[:6])
            target_info.compress_type = ZIP_DEFLATED
            target_info.external_attr = source_info.external_attr
            target_info.create_system = source_info.create_system
            target_zip.writestr(target_info, source_zip.read(name))
    return canonical.getvalue()


def _row_provenance_counts(rows: list[dict]) -> dict:
    projection_rows = [
        row for row in rows if row.get("record_kind") == "public_projection"
    ]
    source_rows = [
        row
        for row in rows
        if row.get("record_kind") == "source_series_observation"
    ]
    return {
        "projection_record_count": len(projection_rows),
        "source_observation_count": len(source_rows),
        "source_series_count": len(
            {row.get("series_code") for row in source_rows if row.get("series_code")}
        ),
    }


def _sum_measure(rows: list[dict], field: str) -> int | float:
    total = sum(float(row.get(field) or 0) for row in rows)
    return int(total) if total.is_integer() else total


def _partition_metadata(dataset: dict, spec: PartitionSpec, rows: list[dict]) -> dict:
    dates = [str(row["date"]) for row in rows]
    projection_rows = [
        row for row in rows if row.get("record_kind") == "public_projection"
    ]
    summary_rows = projection_rows or rows
    return {
        "schema_version": 4,
        "dataset_kind": dataset["kind"],
        "dataset_id": dataset["dataset_id"],
        "dataset_slug": dataset["slug"],
        "dataset_name": dataset["name"],
        "partition_id": spec.key,
        "window_start": spec.start.isoformat(),
        "window_end": spec.end.isoformat(),
        "data_start": min(dates),
        "data_end": max(dates),
        "record_count": len(rows),
        **_row_provenance_counts(rows),
        "total_cases": _sum_measure(summary_rows, "cases"),
        "total_deaths": _sum_measure(summary_rows, "deaths"),
        "total_basis": (
            "public_projection"
            if projection_rows
            else "source_series_observation"
        ),
    }


def _partition_bytes(dataset: dict, spec: PartitionSpec, rows: list[dict]) -> dict[str, bytes]:
    metadata = _partition_metadata(dataset, spec, rows)
    return {
        "csv": _csv_bytes(rows),
        "json": _json_bytes(metadata, rows),
        "xlsx": _xlsx_bytes(metadata, rows),
    }


def _split_spec(spec: PartitionSpec, rows: list[dict]) -> tuple[tuple[PartitionSpec, list[dict]], tuple[PartitionSpec, list[dict]]]:
    """Bisect an oversized window at a stable calendar boundary."""

    start_ordinal = spec.start.toordinal()
    end_ordinal = spec.end.toordinal()
    if start_ordinal >= end_ordinal:
        midpoint = len(rows) // 2
        if midpoint <= 0 or midpoint >= len(rows):
            raise RuntimeError(
                f"A single-date partition cannot fit below the configured limit: {spec.key}"
            )
        left_rows = rows[:midpoint]
        right_rows = rows[midpoint:]
        left_spec = PartitionSpec(
            key=f"{spec.key}-part-01",
            start=spec.start,
            end=spec.end,
            label=f"{spec.label} · 1",
        )
        right_spec = PartitionSpec(
            key=f"{spec.key}-part-02",
            start=spec.start,
            end=spec.end,
            label=f"{spec.label} · 2",
            is_current=spec.is_current,
        )
        return (left_spec, left_rows), (right_spec, right_rows)

    midpoint_date = date.fromordinal((start_ordinal + end_ordinal) // 2)
    left_rows = [row for row in rows if str(row["date"])[:10] <= midpoint_date.isoformat()]
    right_rows = [row for row in rows if str(row["date"])[:10] > midpoint_date.isoformat()]
    if not left_rows or not right_rows:
        ordered_dates = sorted({str(row["date"])[:10] for row in rows})
        if len(ordered_dates) < 2:
            return _split_spec(
                PartitionSpec(spec.key, spec.start, spec.start, spec.label, spec.is_current),
                rows,
            )
        boundary = ordered_dates[len(ordered_dates) // 2 - 1]
        midpoint_date = date.fromisoformat(boundary)
        left_rows = [row for row in rows if str(row["date"])[:10] <= boundary]
        right_rows = [row for row in rows if str(row["date"])[:10] > boundary]

    right_start = date.fromordinal(midpoint_date.toordinal() + 1)
    left_key = f"{spec.start.isoformat()}_{midpoint_date.isoformat()}"
    right_key = f"{right_start.isoformat()}_{spec.end.isoformat()}"
    left_spec = PartitionSpec(
        key=left_key,
        start=spec.start,
        end=midpoint_date,
        label=f"{spec.start.isoformat()} – {midpoint_date.isoformat()}",
    )
    right_spec = PartitionSpec(
        key=right_key,
        start=right_start,
        end=spec.end,
        label=(
            f"{right_start.isoformat()} – now"
            if spec.is_current
            else f"{right_start.isoformat()} – {spec.end.isoformat()}"
        ),
        is_current=spec.is_current,
    )
    return (left_spec, left_rows), (right_spec, right_rows)


def _bounded_partition_artifacts(
    dataset: dict,
    spec: PartitionSpec,
    rows: list[dict],
    max_file_bytes: int,
) -> list[tuple[PartitionSpec, list[dict], dict[str, bytes]]]:
    artifacts = _partition_bytes(dataset, spec, rows)
    if max(len(content) for content in artifacts.values()) < max_file_bytes:
        return [(spec, rows, artifacts)]
    left, right = _split_spec(spec, rows)
    return _bounded_partition_artifacts(dataset, *left, max_file_bytes) + _bounded_partition_artifacts(
        dataset, *right, max_file_bytes
    )


def _write_if_changed(path: Path, content: bytes) -> bool:
    if path.exists() and path.read_bytes() == content:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    return True


def _remove_stale_files(root: Path, expected: set[Path]) -> int:
    removed = 0
    if not root.exists():
        return removed
    for path in sorted(root.rglob("*"), reverse=True):
        if path.is_file() and path not in expected:
            path.unlink()
            removed += 1
        elif path.is_dir() and not any(path.iterdir()):
            path.rmdir()
    return removed


def _dataset_parts(
    *,
    dataset: dict,
    rows: list[dict],
    output_dir: Path,
    download_url_base: str,
    max_file_bytes: int,
    expected_paths: set[Path],
    existing_parts: dict[str, dict],
) -> tuple[list[dict], int]:
    parts: list[dict] = []
    changed = 0
    for base_spec, base_rows in partition_rows(rows):
        base_metadata = _partition_metadata(dataset, base_spec, base_rows)
        base_content_sha = _sha256(_json_bytes(base_metadata, base_rows))
        existing = existing_parts.get(base_spec.key)
        existing_files = (existing or {}).get("files") or {}
        can_reuse = bool(
            existing
            and existing.get("content_sha256") == base_content_sha
            and set(existing_files) == set(PUBLIC_FORMATS)
        )
        for format_name in PUBLIC_FORMATS:
            file_meta = existing_files.get(format_name) or {}
            relative_path = str(file_meta.get("relative_path") or "")
            path = output_dir / relative_path
            if (
                not relative_path
                or not path.is_file()
                or path.stat().st_size != int(file_meta.get("bytes") or -1)
                or path.stat().st_size >= max_file_bytes
                or _file_sha256(path) != str(file_meta.get("sha256") or "")
            ):
                can_reuse = False
                break
        if can_reuse:
            dates = [str(row["date"]) for row in base_rows]
            files: dict[str, dict] = {}
            for format_name in PUBLIC_FORMATS:
                file_meta = existing_files[format_name]
                expected_paths.add(output_dir / file_meta["relative_path"])
                # The artifacts are content-addressed and can be reused, but
                # their public locations are tied to the configured Git branch.
                # Never carry an old Raw base URL forward into a new manifest.
                files[format_name] = {
                    **file_meta,
                    "url": _download_url(
                        download_url_base,
                        file_meta["relative_path"],
                    ),
                }
            parts.append(
                {
                    "id": base_spec.key,
                    "label": base_spec.label,
                    "is_current": base_spec.is_current,
                    "window": {
                        "start": base_spec.start.isoformat(),
                        "end": base_spec.end.isoformat(),
                    },
                    "date_range": {"start": min(dates), "end": max(dates)},
                    "record_count": len(base_rows),
                    **_row_provenance_counts(base_rows),
                    "content_sha256": base_content_sha,
                    "files": files,
                }
            )
            continue

        bounded = _bounded_partition_artifacts(
            dataset,
            base_spec,
            base_rows,
            max_file_bytes,
        )
        for spec, part_rows, artifacts in bounded:
            files: dict[str, dict] = {}
            for format_name in PUBLIC_FORMATS:
                relative_path = (
                    f"{dataset['directory']}/{dataset['path_id']}/"
                    f"{spec.key}.{format_name}"
                )
                path = output_dir / relative_path
                expected_paths.add(path)
                content = artifacts[format_name]
                changed += int(_write_if_changed(path, content))
                if len(content) >= GITHUB_MAX_FILE_BYTES:
                    raise RuntimeError(
                        f"Generated file exceeds GitHub's limit: {path} ({len(content)} bytes)"
                    )
                files[format_name] = {
                    "url": _download_url(download_url_base, relative_path),
                    "relative_path": relative_path,
                    "filename": f"globalid-{dataset['path_id']}-{spec.key}.{format_name}",
                    "bytes": len(content),
                    "sha256": _sha256(content),
                }
            dates = [str(row["date"]) for row in part_rows]
            parts.append(
                {
                    "id": spec.key,
                    "label": spec.label,
                    "is_current": spec.is_current,
                    "window": {
                        "start": spec.start.isoformat(),
                        "end": spec.end.isoformat(),
                    },
                    "date_range": {"start": min(dates), "end": max(dates)},
                    "record_count": len(part_rows),
                    **_row_provenance_counts(part_rows),
                    "content_sha256": _sha256(artifacts["json"]),
                    "files": files,
                }
            )
    return parts, changed


def _build_dataset_parts_task(
    *,
    kind: str,
    dataset: dict,
    rows: list[dict],
    output_dir: Path,
    download_url_base: str,
    max_file_bytes: int,
    existing_parts: dict[str, dict],
) -> dict:
    """Build one dataset tree independently so changed partitions can overlap."""

    expected_paths: set[Path] = set()
    parts, changed = _dataset_parts(
        dataset=dataset,
        rows=rows,
        output_dir=output_dir,
        download_url_base=download_url_base,
        max_file_bytes=max_file_bytes,
        expected_paths=expected_paths,
        existing_parts=existing_parts,
    )
    return {
        "kind": kind,
        "dataset": dataset,
        "row_count": len(rows),
        "provenance_counts": _row_provenance_counts(rows),
        "parts": parts,
        "changed": changed,
        "paths": expected_paths,
    }


def _build_country_parts_task(
    country_export: dict,
    output_dir: Path,
    download_url_base: str,
    max_file_bytes: int,
    existing_parts: dict[str, dict],
) -> dict:
    code = country_export["code"]
    path_id = code.lower()
    dataset = {
        "kind": "country",
        "directory": "countries",
        "dataset_id": path_id,
        "path_id": path_id,
        "slug": path_id,
        "name": country_export["country_name"],
    }
    rows = build_country_download_rows(
        country_export["country_data"], country_export["source_info"]
    )
    result = _build_dataset_parts_task(
        kind="country",
        dataset=dataset,
        rows=rows,
        output_dir=output_dir,
        download_url_base=download_url_base,
        max_file_bytes=max_file_bytes,
        existing_parts=existing_parts,
    )
    result["source_info"] = country_export["source_info"]
    return result


def _build_disease_parts_task(
    disease_export: dict,
    country_sources_by_code: dict[str, dict],
    country_names: dict[str, str],
    output_dir: Path,
    download_url_base: str,
    max_file_bytes: int,
    existing_parts: dict[str, dict],
) -> dict:
    did = disease_export["disease_id"]
    path_id = did.lower()
    disease_data = disease_export["disease_data"]
    dataset = {
        "kind": "disease",
        "directory": "diseases",
        "dataset_id": did,
        "path_id": path_id,
        "slug": disease_data.get("slug"),
        "name": disease_data.get("name_en"),
    }
    rows = build_disease_download_rows(
        disease_data,
        country_sources_by_code,
        country_names,
    )
    result = _build_dataset_parts_task(
        kind="disease",
        dataset=dataset,
        rows=rows,
        output_dir=output_dir,
        download_url_base=download_url_base,
        max_file_bytes=max_file_bytes,
        existing_parts=existing_parts,
    )
    result["country_codes"] = sorted(
        (disease_data.get("country_series") or {}).keys()
    )
    result["source_info"] = disease_data.get("source_info") or []
    return result


def build_direct_download_files(
    context: dict,
    output_dir: Path,
    *,
    download_url_base: str,
    max_file_bytes: int = DEFAULT_TARGET_FILE_BYTES,
    workers: int | None = None,
) -> dict:
    """Write deterministic CSV/JSON/XLSX partitions and return the manifest.

    Unchanged windows are reused by content hash.  Changed country and disease
    trees write to disjoint paths, allowing their CSV/JSON/XLSX generation to
    run concurrently without changing manifest order.
    """

    if not download_url_base.startswith("https://raw.githubusercontent.com/"):
        raise ValueError("Public downloads require an absolute GitHub Raw URL base")
    if max_file_bytes >= GITHUB_MAX_FILE_BYTES:
        raise ValueError("The partition target must stay below GitHub's 100 MiB limit")

    output_dir.mkdir(parents=True, exist_ok=True)
    misspelled_country_dir = output_dir / "countrys"
    if misspelled_country_dir.exists():
        shutil.rmtree(misspelled_country_dir)
    existing_manifest_path = output_dir / "manifest.json"
    try:
        existing_manifest = json.loads(existing_manifest_path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        existing_manifest = {}
    existing_country_parts = {
        entry.get("id"): {part.get("id"): part for part in entry.get("parts") or []}
        for entry in existing_manifest.get("countries") or []
    }
    existing_disease_parts = {
        entry.get("disease_id"): {part.get("id"): part for part in entry.get("parts") or []}
        for entry in existing_manifest.get("diseases") or []
    }
    expected_paths: set[Path] = set()
    generated_at = context["generated_at"]
    country_names = {
        country["code"]: country["name"] for country in context["countries_simple"]
    }
    country_entries_by_id = {
        entry["id"]: entry for entry in context["country_download_entries"]
    }
    disease_entries_by_id = {
        entry["disease_id"]: entry for entry in context["disease_download_entries"]
    }
    country_exports = context["country_exports"]
    disease_exports = context["disease_exports"]
    task_count = len(country_exports) + len(disease_exports)
    worker_count = min(max(1, workers or DEFAULT_EXPORT_WORKERS), max(1, task_count))
    results_by_index: dict[int, dict] = {}

    if task_count:
        with ProcessPoolExecutor(
            max_workers=worker_count,
            mp_context=_process_pool_context(),
        ) as executor:
            pending: dict[Future, int] = {}
            next_index = 0
            max_pending = min(task_count, worker_count * 2)

            def submit_next() -> None:
                nonlocal next_index
                index = next_index
                next_index += 1
                if index < len(country_exports):
                    country_export = country_exports[index]
                    path_id = country_export["code"].lower()
                    future = executor.submit(
                        _build_country_parts_task,
                        country_export,
                        output_dir,
                        download_url_base,
                        max_file_bytes,
                        existing_country_parts.get(path_id, {}),
                    )
                else:
                    disease_export = disease_exports[index - len(country_exports)]
                    future = executor.submit(
                        _build_disease_parts_task,
                        disease_export,
                        context["country_sources_by_code"],
                        country_names,
                        output_dir,
                        download_url_base,
                        max_file_bytes,
                        existing_disease_parts.get(disease_export["disease_id"], {}),
                    )
                pending[future] = index

            while next_index < task_count and len(pending) < max_pending:
                submit_next()

            while pending:
                done, _ = wait(pending, return_when=FIRST_COMPLETED)
                for future in done:
                    index = pending.pop(future)
                    results_by_index[index] = future.result()
                    if next_index < task_count:
                        submit_next()

    country_entries: list[dict] = []
    disease_entries: list[dict] = []
    changed_files = 0
    for index in range(task_count):
        result = results_by_index[index]
        kind = result["kind"]
        dataset = result["dataset"]
        expected_paths.update(result["paths"])
        changed_files += result["changed"]
        if kind == "country":
            entry = dict(country_entries_by_id[dataset["path_id"]])
            entry.update(
                {
                    "record_count": result["row_count"],
                    **result["provenance_counts"],
                    "includes_series_provenance": True,
                    "parts": result["parts"],
                    "source_info": result["source_info"],
                }
            )
            country_entries.append(entry)
            continue

        entry = dict(disease_entries_by_id[dataset["dataset_id"]])
        entry.update(
            {
                "record_count": result["row_count"],
                **result["provenance_counts"],
                "includes_series_provenance": True,
                "parts": result["parts"],
                "countries": [
                    {"code": code, "name": country_names.get(code)}
                    for code in result["country_codes"]
                ],
                "source_info": result["source_info"],
            }
        )
        disease_entries.append(entry)

    removed_files = _remove_stale_files(output_dir / "countries", expected_paths)
    removed_files += _remove_stale_files(output_dir / "diseases", expected_paths)
    manifest = {
        "schema_version": 4,
        "generated_at": generated_at,
        "includes_source_info": True,
        "includes_series_provenance": True,
        "formats": list(PUBLIC_FORMATS),
        "partitioning": {
            "strategy": "stable_calendar_windows",
            "historical_years": 5,
            "special_window": "2020-2025",
            "bridge_window": "2026-2029",
            "rolling_anchor_year": ROLLING_WINDOW_START,
            "target_file_bytes": max_file_bytes,
            "github_max_file_bytes": GITHUB_MAX_FILE_BYTES,
        },
        "download_url_base": download_url_base.rstrip("/"),
        "countries": country_entries,
        "diseases": disease_entries,
        "generation": {
            "changed_files": changed_files,
            "removed_files": removed_files,
        },
    }
    manifest_path = output_dir / "manifest.json"
    _write_if_changed(
        manifest_path,
        json.dumps(
            manifest,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
        ).encode("utf-8"),
    )
    return manifest


__all__ = [
    "DEFAULT_TARGET_FILE_BYTES",
    "DOWNLOAD_FIELDS",
    "GITHUB_MAX_FILE_BYTES",
    "PUBLIC_FORMATS",
    "PartitionSpec",
    "build_country_download_rows",
    "build_disease_download_rows",
    "build_direct_download_files",
    "partition_rows",
]
